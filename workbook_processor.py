"""Read, validate, calculate, and return controlled PROWRAP batch workbooks."""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import UTC, datetime
from io import BytesIO
import logging
import math
from pathlib import Path, PurePosixPath
import posixpath
import traceback
import zipfile
from xml.etree.ElementTree import ParseError, iterparse

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Alignment, Border, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

try:  # openpyxl uses lxml when it is available.
    from lxml.etree import XMLSyntaxError
except ImportError:  # pragma: no cover - exercised where lxml is unavailable.
    XMLSyntaxError = ParseError

from batch_adapter import (
    CandidateCalculation,
    RowCalculation,
    calculate_row,
    normalize_audit_scalar,
)
from batch_corrosion import ManualGroupLinks, link_manual_groups
from batch_mechanisms import normalize_upload_mechanism
from batch_schema import (
    DETAIL_INPUT_HEADERS,
    DETAIL_OUTPUT_HEADERS,
    INPUT_HEADERS,
    HISTORICAL_V12_OUTPUT_HEADERS,
    LEGACY_INPUT_HEADERS,
    LEGACY_OUTPUT_HEADERS,
    MAX_DETAIL_ROWS,
    MAX_ROWS,
    MAX_UPLOAD_BYTES,
    OUTPUT_HEADERS,
    BatchInfo,
    ValidatedIndividualDefectRow,
    ValidatedRow,
    ValidationIssue,
)
from batch_status import CalculationStatus
from batch_validation import (
    validate_batch_info,
    validate_individual_defect_row,
    validate_row,
)
from cost_calculation import (
    COST_FIRST_DATA_ROW,
    COST_INPUTS,
    COST_LAST_DATA_ROW,
    COST_SOURCE_HEADERS,
    COST_TABLE_HEADERS,
    cost_formula,
    is_allowed_cost_formula,
    price_formula,
    total_amount_formula,
)
from workbook_template import create_template_workbook
from warning_catalog import format_affected_rows, warning_meaning
from engine.corrosion_defects import ACTUAL_DEFECT_LENGTH, ENTER_MANUALLY


BATCH_ENGINE_VERSION = '1.2.0'
SOURCE_ENGINE_REVISION = '91b68d6'
_COMMON_HEADERS = ('Customer', 'Project Location', 'Report No')
_LEGACY_SHEETS = (
    'Batch Information',
    'Batch Input & Results',
    'Summary',
    'Instructions',
    'Lists',
)
_PREVIOUS_SHEETS = (
    'Batch Information',
    'Batch Input & Results',
    'Warnings',
    'Summary',
    'Instructions',
    'Lists',
)
_LEGACY_COST_SHEETS = (
    'Batch Information',
    'Batch Input & Results',
    'Cost Calculation',
    'Warnings',
    'Summary',
    'Instructions',
    'Lists',
)
_CURRENT_SHEETS = (
    'Batch Information',
    'Batch Input & Results',
    'Individual Defects',
    'Cost Calculation',
    'Warnings',
    'Summary',
    'Instructions',
    'Lists',
)
_PREVIEW_LIMIT = 20
_MAX_ZIP_ENTRIES = 250
_MAX_ZIP_UNCOMPRESSED_BYTES = 25 * 1024 * 1024
_MAX_ZIP_ENTRY_BYTES = 16 * 1024 * 1024
_MAX_ZIP_COMPRESSION_RATIO = 100
# The blank 150-row eight-sheet v1.2 template emits 11,710 worksheet cell
# elements. This ceiling retains more than 70 percent headroom while bounding
# the number of Python cell objects openpyxl may materialize.
_MAX_WORKBOOK_CELLS = 20_000
_SPREADSHEETML_MAIN_NAMESPACES = frozenset({
    'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
    'http://purl.oclc.org/ooxml/spreadsheetml/main',
})
_WORKSHEET_CONTENT_TYPE = (
    'application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml'
)
_OPC_RELATIONSHIPS_NAMESPACES = frozenset({
    'http://schemas.openxmlformats.org/package/2006/relationships',
    'http://purl.oclc.org/ooxml/package/relationships',
})
_OFFICE_DOCUMENT_RELATIONSHIP_TYPES = frozenset({
    'http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument',
    'http://purl.oclc.org/ooxml/officeDocument/relationships/officeDocument',
})
_WORKSHEET_RELATIONSHIP_TYPES = frozenset({
    'http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet',
    'http://purl.oclc.org/ooxml/officeDocument/relationships/worksheet',
})

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class WorkbookInspection:
    """Workbook-level validation outcome and a concise row preview."""

    batch_info: BatchInfo | None
    populated_rows: int
    valid_rows: int
    invalid_rows: int
    workbook_errors: tuple[ValidationIssue, ...]
    preview: tuple[dict[str, object], ...]
    recognized_input_headers: tuple[str, ...]
    missing_input_headers: tuple[str, ...]
    unexpected_headers: tuple[str, ...]
    populated_detail_rows: int
    manual_groups: int
    recognized_detail_input_headers: tuple[str, ...]
    missing_detail_input_headers: tuple[str, ...]
    unexpected_detail_headers: tuple[str, ...]


@dataclass(frozen=True)
class WorkbookContract:
    """One exact upload contract accepted by the v1.2 processor."""

    sheet_order: tuple[str, ...]
    input_headers: tuple[str, ...]
    output_headers: tuple[str, ...]
    has_individual_defects: bool
    is_legacy: bool


@dataclass(frozen=True)
class _PreparedRows:
    main_values: tuple[tuple[int, dict[str, object]], ...]
    main_rows: dict[int, ValidatedRow]
    main_issues: dict[int, tuple[ValidationIssue, ...]]
    detail_values: tuple[tuple[int, dict[str, object]], ...]
    detail_rows: dict[int, ValidatedIndividualDefectRow]
    detail_issues: dict[int, tuple[ValidationIssue, ...]]
    links: ManualGroupLinks


_LEGACY_CONTRACTS = tuple(
    WorkbookContract(order, LEGACY_INPUT_HEADERS, LEGACY_OUTPUT_HEADERS, False, True)
    for order in (_LEGACY_SHEETS, _PREVIOUS_SHEETS, _LEGACY_COST_SHEETS)
)
_CURRENT_CONTRACT = WorkbookContract(
    _CURRENT_SHEETS, INPUT_HEADERS, OUTPUT_HEADERS, True, False,
)
_HISTORICAL_V12_CONTRACT = WorkbookContract(
    _CURRENT_SHEETS, INPUT_HEADERS, HISTORICAL_V12_OUTPUT_HEADERS, True, False,
)
_ACCEPTED_CONTRACTS = _LEGACY_CONTRACTS + (
    _HISTORICAL_V12_CONTRACT,
    _CURRENT_CONTRACT,
)


@dataclass(frozen=True)
class ProcessedBatch:
    """A new in-memory workbook and the row-level status totals it contains."""

    workbook_bytes: bytes
    status_counts: dict[str, int]
    populated_rows: int


class WorkbookProcessingError(ValueError):
    """Raised only when the controlled workbook itself cannot be processed."""

    def __init__(self, issues: tuple[ValidationIssue, ...]):
        self.issues = issues
        super().__init__('; '.join(issue.message for issue in issues))


def inspect_workbook(data: bytes) -> WorkbookInspection:
    """Validate a controlled upload and preview the final row classifications."""
    workbook, errors = _load_controlled_workbook(data)
    if errors:
        return _empty_inspection(errors)

    assert workbook is not None
    formula_errors = _formula_errors(workbook)
    if formula_errors:
        return _empty_inspection(formula_errors)

    header_summary = _input_header_summary(workbook)
    detail_header_summary = _detail_input_header_summary(workbook)
    contract, structure_errors = _validate_structure(workbook)
    if structure_errors:
        return _empty_inspection(
            structure_errors, header_summary, detail_header_summary,
        )
    assert contract is not None

    info_sheet = workbook['Batch Information']
    common_values = {
        header: info_sheet.cell(row, 2).value
        for row, header in enumerate(_COMMON_HEADERS, start=3)
    }
    commercial_input_errors = _commercial_input_errors(workbook)
    if commercial_input_errors:
        return _empty_inspection(
            commercial_input_errors, header_summary, detail_header_summary,
        )
    quantity_errors = _quantity_errors(workbook)
    if quantity_errors:
        return _empty_inspection(
            quantity_errors, header_summary, detail_header_summary,
        )

    batch_info, batch_issues = validate_batch_info(common_values)
    data_sheet = workbook['Batch Input & Results']
    out_of_range_errors = _out_of_range_input_errors(
        data_sheet, contract.input_headers,
    )
    if out_of_range_errors:
        return _empty_inspection(
            out_of_range_errors, header_summary, detail_header_summary,
        )
    if contract.has_individual_defects:
        detail_range_errors = _out_of_range_input_errors(
            workbook['Individual Defects'],
            DETAIL_INPUT_HEADERS,
            max_rows=MAX_DETAIL_ROWS,
            code='DETAIL_ROW_OUT_OF_RANGE',
            label='Individual Defects input values',
        )
        if detail_range_errors:
            return _empty_inspection(
                detail_range_errors, header_summary, detail_header_summary,
            )

    prepared = _prepare_rows(workbook, contract)

    valid_rows = 0
    invalid_rows = 0
    preview: list[dict[str, object]] = []
    for excel_row, values in prepared.main_values:
        row = prepared.main_rows.get(excel_row)
        issues = prepared.main_issues.get(excel_row, ())
        if issues:
            invalid_rows += 1
            status = CalculationStatus.INPUT_ERROR.value
            error_code = _issue_codes(issues)
            error_message = _issue_message(issues)
        elif batch_info is None:
            valid_rows += 1
            status = 'READY'
            error_code = ''
            error_message = ''
        else:
            valid_rows += 1
            calculation = _calculate_one(
                batch_info,
                excel_row,
                values,
                individual_defects=prepared.links.defects_by_main_excel_row.get(
                    excel_row, (),
                ),
            )
            status = calculation.status.value
            error_code = calculation.error_code
            error_message = calculation.error_message
        if len(preview) < _PREVIEW_LIMIT:
            preview.append({
                'Source Excel Row': excel_row,
                'Pipe OD [mm]': values['Pipe OD [mm]'],
                'Mechanism': (
                    row.values['Mechanism'] if row is not None else values['Mechanism']
                ),
                'Defect Location': values['Defect Location'],
                'Defect Length Basis': values['Defect Length Basis'],
                'Repair Group ID': values['Repair Group ID'],
                'Calculation Status': status,
                'Error Code': error_code,
                'Error Message': error_message,
            })

    return WorkbookInspection(
        batch_info=batch_info,
        populated_rows=len(prepared.main_values),
        valid_rows=valid_rows,
        invalid_rows=invalid_rows,
        workbook_errors=tuple(batch_issues),
        preview=tuple(preview),
        recognized_input_headers=header_summary[0],
        missing_input_headers=header_summary[1],
        unexpected_headers=header_summary[2],
        populated_detail_rows=len(prepared.detail_values),
        manual_groups=len({
            row.values['Repair Group ID']
            for row in prepared.main_rows.values()
            if row.values.get('Defect Length Basis') == ENTER_MANUALLY
        }),
        recognized_detail_input_headers=detail_header_summary[0],
        missing_detail_input_headers=detail_header_summary[1],
        unexpected_detail_headers=detail_header_summary[2],
    )


def process_workbook(
    data: bytes,
    processed_at: datetime,
    source_name: str | None = None,
) -> ProcessedBatch:
    """Return a new workbook with calculation results appended to each populated row."""
    inspection = inspect_workbook(data)
    if inspection.workbook_errors:
        raise WorkbookProcessingError(inspection.workbook_errors)
    assert inspection.batch_info is not None

    workbook, load_errors = _load_controlled_workbook(data)
    if load_errors or workbook is None:  # Defensive: bytes were just inspected.
        raise WorkbookProcessingError(load_errors)

    contract, structure_errors = _validate_structure(workbook)
    if structure_errors or contract is None:  # Defensive: bytes were just inspected.
        raise WorkbookProcessingError(structure_errors)

    output_workbook = load_workbook(BytesIO(create_template_workbook()), data_only=False)
    _copy_controlled_inputs(workbook, output_workbook, contract)
    prepared = _prepare_rows(output_workbook, _CURRENT_CONTRACT)
    data_sheet = output_workbook['Batch Input & Results']
    detail_sheet = output_workbook['Individual Defects']
    status_counts: Counter[str] = Counter()
    processed_timestamp = _utc_timestamp(processed_at)
    calculations: dict[int, RowCalculation] = {}
    for excel_row, values in prepared.main_values:
        issues = prepared.main_issues.get(excel_row, ())
        if issues:
            calculation = _input_error_calculation(excel_row, issues)
        else:
            calculation = _calculate_one(
                inspection.batch_info,
                excel_row,
                values,
                individual_defects=prepared.links.defects_by_main_excel_row.get(
                    excel_row, (),
                ),
            )
        calculations[excel_row] = calculation
        _write_result_row(
            data_sheet,
            excel_row,
            calculation,
        )
        status_counts[calculation.status.value] += 1

    candidate_by_detail_row: dict[int, CandidateCalculation] = {}
    for main_excel_row, detail_rows in prepared.links.detail_rows_by_main_excel_row.items():
        calculation = calculations.get(main_excel_row)
        if calculation is None or prepared.main_issues.get(main_excel_row):
            continue
        if len(calculation.candidate_calculations) != len(detail_rows):
            continue
        candidate_by_detail_row.update({
            detail_row.source_excel_row: candidate
            for detail_row, candidate in zip(
                detail_rows, calculation.candidate_calculations, strict=True,
            )
        })

    for detail_excel_row, _ in prepared.detail_values:
        issues = prepared.detail_issues.get(detail_excel_row, ())
        candidate = candidate_by_detail_row.get(detail_excel_row)
        if not issues and candidate is None:
            linked_main_row = _linked_main_row_for_detail(
                detail_excel_row, prepared.links,
            )
            if linked_main_row is not None:
                issues = prepared.main_issues.get(linked_main_row, ())
            if not issues:
                issues = (_issue(
                    'MAIN_ROW_NOT_CALCULATED',
                    'The linked main repair row could not produce this detail assessment.',
                ),)
        _write_detail_result_row(
            detail_sheet, detail_excel_row, issues=issues, candidate=candidate,
        )

    _write_cost_sheet(output_workbook)
    _write_warnings_sheet(output_workbook, calculations)
    _write_summary(
        output_workbook,
        inspection.batch_info,
        processed_timestamp,
        inspection.populated_rows,
        calculations,
        _sanitized_source_name(source_name),
    )
    output_workbook.calculation.calcMode = 'auto'
    output_workbook.calculation.fullCalcOnLoad = True
    output_workbook.calculation.forceFullCalc = True
    output = BytesIO()
    output_workbook.save(output)
    return ProcessedBatch(
        workbook_bytes=output.getvalue(),
        status_counts=dict(status_counts),
        populated_rows=inspection.populated_rows,
    )


def _load_controlled_workbook(data: bytes):
    if len(data) > MAX_UPLOAD_BYTES:
        return None, (_issue('FILE_TOO_LARGE', 'The uploaded workbook exceeds the 10 MB limit.'),)
    try:
        with zipfile.ZipFile(BytesIO(data)) as archive:
            safety_errors = _zip_safety_errors(archive)
            if safety_errors:
                return None, safety_errors
            for entry in archive.infolist():
                if entry.flag_bits & 0x1:
                    return None, (_issue('UNREADABLE_WORKBOOK', 'Password-protected workbooks are not supported.'),)
                if PurePosixPath(entry.filename).name.lower() == 'vbaproject.bin':
                    return None, (_issue('MACROS_NOT_ALLOWED', 'Macro-enabled workbooks are not supported.'),)
        return load_workbook(BytesIO(data), data_only=False, keep_vba=False), ()
    except (
        InvalidFileException,
        KeyError,
        OSError,
        ValueError,
        zipfile.BadZipFile,
        ParseError,
        XMLSyntaxError,
    ):
        return None, (_issue('UNREADABLE_WORKBOOK', 'The uploaded file is not a readable .xlsx workbook.'),)


def _validate_structure(
    workbook,
) -> tuple[WorkbookContract | None, tuple[ValidationIssue, ...]]:
    missing = [sheet for sheet in _LEGACY_SHEETS if sheet not in workbook.sheetnames]
    if missing:
        return None, (_issue('MISSING_WORKSHEET', f'Missing required worksheet: {missing[0]}.'),)
    extras = [sheet for sheet in workbook.sheetnames if sheet not in _CURRENT_SHEETS]
    if extras:
        return None, (_issue('UNEXPECTED_WORKSHEET', f'Unexpected worksheet: {extras[0]}.'),)

    sheet_order = tuple(workbook.sheetnames)
    headings = tuple(cell.value for cell in workbook['Batch Input & Results'][1])
    order_contracts = tuple(
        item for item in _ACCEPTED_CONTRACTS if item.sheet_order == sheet_order
    )
    contract = next(
        (
            item for item in order_contracts
            if headings == item.input_headers + item.output_headers
        ),
        None,
    )
    if sheet_order in {item.sheet_order for item in _LEGACY_CONTRACTS} and headings in {
        INPUT_HEADERS + OUTPUT_HEADERS,
        INPUT_HEADERS + HISTORICAL_V12_OUTPUT_HEADERS,
    }:
        return None, (_issue(
            'MISSING_WORKSHEET',
            'Missing required worksheet: Individual Defects.',
        ),)
    if contract is None:
        if headings in {
            INPUT_HEADERS + OUTPUT_HEADERS,
            INPUT_HEADERS + HISTORICAL_V12_OUTPUT_HEADERS,
        }:
            missing_current = [
                sheet for sheet in _CURRENT_SHEETS if sheet not in workbook.sheetnames
            ]
            if missing_current:
                return None, (_issue(
                    'MISSING_WORKSHEET',
                    f'Missing required worksheet: {missing_current[0]}.',
                ),)
        if order_contracts:
            duplicate = _first_duplicate(headings)
            if duplicate:
                return None, (_issue(
                    'DUPLICATE_INPUT_HEADER',
                    f'Duplicate workbook heading: {duplicate}.',
                ),)
            return None, (_issue(
                'INVALID_INPUT_HEADERS',
                'Batch Input & Results headings do not match the controlled template.',
            ),)
        return None, (_issue(
            'INVALID_WORKSHEET_ORDER',
            'Worksheets do not match a controlled legacy or current template order.',
        ),)

    info_sheet = workbook['Batch Information']
    common_headers = tuple(info_sheet.cell(row, 1).value for row in range(3, 6))
    if common_headers != _COMMON_HEADERS:
        return None, (_issue(
            'INVALID_BATCH_INFORMATION_LABELS',
            'Batch Information must contain Customer, Project Location, and Report No labels.',
        ),)

    duplicate = _first_duplicate(headings)
    if duplicate:
        return None, (_issue('DUPLICATE_INPUT_HEADER', f'Duplicate workbook heading: {duplicate}.'),)
    if contract.has_individual_defects:
        detail_headings = tuple(cell.value for cell in workbook['Individual Defects'][1])
        duplicate = _first_duplicate(detail_headings)
        if duplicate:
            return None, (_issue(
                'DUPLICATE_DETAIL_HEADER',
                f'Duplicate Individual Defects heading: {duplicate}.',
            ),)
        if detail_headings != DETAIL_INPUT_HEADERS + DETAIL_OUTPUT_HEADERS:
            return None, (_issue(
                'INVALID_DETAIL_HEADERS',
                'Individual Defects headings do not match the controlled template.',
            ),)
    if 'Cost Calculation' in workbook.sheetnames:
        cost_headings = tuple(cell.value for cell in workbook['Cost Calculation'][5])
        former_cost_headers = COST_SOURCE_HEADERS + ('Cost', 'Price')
        accepts_former_cost_contract = (
            cost_headings == former_cost_headers
            and (
                contract is _HISTORICAL_V12_CONTRACT
                or contract.sheet_order == _LEGACY_COST_SHEETS
            )
        )
        if cost_headings != COST_TABLE_HEADERS and not accepts_former_cost_contract:
            return None, (_issue(
                'INVALID_COST_HEADERS',
                'Cost Calculation headings do not match the controlled template.',
            ),)
    return contract, ()


def _input_header_summary(workbook) -> tuple[tuple[str, ...], tuple[str, ...], tuple[str, ...]]:
    """Return user-facing input-header recognition details without relaxing validation."""
    if 'Batch Input & Results' not in workbook.sheetnames:
        return (), (), ()

    headings = tuple(cell.value for cell in workbook['Batch Input & Results'][1])
    expected = INPUT_HEADERS + OUTPUT_HEADERS
    recognized = tuple(header for header in INPUT_HEADERS if header in headings)
    missing = tuple(header for header in INPUT_HEADERS if header not in headings)
    unexpected = tuple(
        _display_heading(heading) for heading in headings if heading not in expected
    )
    return recognized, missing, unexpected


def _detail_input_header_summary(
    workbook,
) -> tuple[tuple[str, ...], tuple[str, ...], tuple[str, ...]]:
    """Return detail-table heading diagnostics independently of main headings."""
    if 'Individual Defects' not in workbook.sheetnames:
        return (), DETAIL_INPUT_HEADERS, ()
    headings = tuple(cell.value for cell in workbook['Individual Defects'][1])
    expected = DETAIL_INPUT_HEADERS + DETAIL_OUTPUT_HEADERS
    recognized = tuple(header for header in DETAIL_INPUT_HEADERS if header in headings)
    missing = tuple(header for header in DETAIL_INPUT_HEADERS if header not in headings)
    unexpected = tuple(
        _display_heading(heading) for heading in headings if heading not in expected
    )
    return recognized, missing, unexpected


def _formula_errors(workbook) -> tuple[ValidationIssue, ...]:
    for worksheet in workbook.worksheets:
        for _, cell in _loaded_cells(worksheet):
            if _is_formula_cell(cell):
                if (
                    worksheet.title == 'Cost Calculation'
                    and getattr(cell, 'data_type', None) == 'f'
                    and isinstance(cell.value, str)
                    and is_allowed_cost_formula(cell)
                ):
                    continue
                return (_issue(
                    'FORMULA_NOT_ALLOWED',
                    f'Formula cells are not allowed: {worksheet.title}!{cell.coordinate}.',
                ),)
    return ()


def _commercial_input_errors(workbook) -> tuple[ValidationIssue, ...]:
    if 'Cost Calculation' not in workbook.sheetnames:
        return ()
    worksheet = workbook['Cost Calculation']
    for address, label in COST_INPUTS:
        value = worksheet[address].value
        if _is_blank(value):
            continue
        if (
            isinstance(value, bool)
            or not isinstance(value, (int, float))
            or not math.isfinite(value)
            or value < 0
        ):
            return (_issue(
                'INVALID_COST_INPUT',
                f'{label} must be blank or a non-negative number: {address}.',
            ),)
    return ()


def _quantity_errors(workbook) -> tuple[ValidationIssue, ...]:
    """Reject untrusted Quantity values outside the controlled input contract."""
    if 'Cost Calculation' not in workbook.sheetnames:
        return ()
    worksheet = workbook['Cost Calculation']
    for _, cell in _loaded_cells(worksheet):
        if cell.column != 23 or cell.row < COST_FIRST_DATA_ROW or _is_blank(cell.value):
            continue
        value = cell.value
        if (
            cell.row > COST_LAST_DATA_ROW
            or isinstance(value, bool)
            or not isinstance(value, (int, float))
            or not math.isfinite(value)
            or value < 0
        ):
            return (_issue(
                'INVALID_QUANTITY',
                f'Quantity must be blank or a non-negative number: {cell.coordinate}.',
            ),)
    return ()


def _populated_rows(
    worksheet,
    input_headers: tuple[str, ...] = INPUT_HEADERS,
    *,
    max_rows: int = MAX_ROWS,
) -> list[tuple[int, dict[str, object]]]:
    populated: list[tuple[int, dict[str, object]]] = []
    for excel_row in range(2, max_rows + 2):
        values = {
            header: worksheet.cell(excel_row, column).value
            for column, header in enumerate(input_headers, start=1)
        }
        if any(not _is_blank(value) for value in values.values()):
            populated.append((excel_row, values))
    return populated


def _out_of_range_input_errors(
    worksheet,
    input_headers: tuple[str, ...] = INPUT_HEADERS,
    *,
    max_rows: int = MAX_ROWS,
    code: str = 'INPUT_ROW_OUT_OF_RANGE',
    label: str = 'Input values',
) -> tuple[ValidationIssue, ...]:
    for (excel_row, column), cell in _loaded_cells(worksheet):
        if (
            excel_row >= max_rows + 2
            and column <= len(input_headers)
            and not _is_blank(cell.value)
        ):
            return (_issue(
                code,
                f'{label} are allowed only in rows 2 through {max_rows + 1}: {cell.coordinate}.',
            ),)
    return ()


def _prepare_rows(workbook, contract: WorkbookContract) -> _PreparedRows:
    main_values = tuple(_normalized_main_rows(workbook, contract))
    main_rows: dict[int, ValidatedRow] = {}
    main_issues: dict[int, tuple[ValidationIssue, ...]] = {}
    for excel_row, values in main_values:
        row, issues = validate_row(excel_row, values)
        if row is not None:
            main_rows[excel_row] = row
        if issues:
            main_issues[excel_row] = issues

    detail_values: tuple[tuple[int, dict[str, object]], ...] = ()
    detail_rows: dict[int, ValidatedIndividualDefectRow] = {}
    detail_issues: dict[int, tuple[ValidationIssue, ...]] = {}
    raw_detail_groups: dict[int, str] = {}
    if contract.has_individual_defects:
        detail_values = tuple(_populated_rows(
            workbook['Individual Defects'],
            DETAIL_INPUT_HEADERS,
            max_rows=MAX_DETAIL_ROWS,
        ))
        for excel_row, values in detail_values:
            raw_group = values.get('Repair Group ID')
            if not _is_blank(raw_group):
                raw_detail_groups[excel_row] = str(raw_group).strip()
            row, issues = validate_individual_defect_row(excel_row, values)
            if row is not None:
                detail_rows[excel_row] = row
            if issues:
                detail_issues[excel_row] = issues

    links = link_manual_groups(
        tuple(main_rows.values()),
        tuple(detail_rows.values()),
        detail_issues,
    )
    mutable_main_issues = {
        row: list(issues) for row, issues in main_issues.items()
    }
    for row, issues in links.main_issues.items():
        mutable_main_issues.setdefault(row, []).extend(issues)
    mutable_detail_issues = {
        row: list(issues) for row, issues in links.detail_issues.items()
    }

    manual_mains_by_group: dict[str, list[int]] = {}
    for main_row in main_rows.values():
        if main_row.values.get('Defect Length Basis') == ENTER_MANUALLY:
            manual_mains_by_group.setdefault(
                main_row.values['Repair Group ID'], [],
            ).append(main_row.source_excel_row)
    for detail_excel_row in detail_issues:
        group_id = raw_detail_groups.get(detail_excel_row)
        if not group_id:
            continue
        owners = manual_mains_by_group.get(group_id, [])
        if len(owners) == 1:
            _append_unique_issue(
                mutable_main_issues,
                owners[0],
                _issue(
                    'INVALID_INDIVIDUAL_DEFECTS',
                    f'Repair Group ID {group_id!r} contains invalid individual defect rows.',
                ),
            )
        elif not owners:
            _append_unique_issue(
                mutable_detail_issues,
                detail_excel_row,
                _issue(
                    'ORPHAN_REPAIR_GROUP',
                    f'Repair Group ID {group_id!r} does not link to a manual main row.',
                ),
            )
        else:
            _append_unique_issue(
                mutable_detail_issues,
                detail_excel_row,
                _issue(
                    'AMBIGUOUS_REPAIR_GROUP',
                    f'Repair Group ID {group_id!r} links to multiple manual main rows.',
                ),
            )

    merged_main_issues = {
        row: tuple(_deduplicate_issues(issues))
        for row, issues in mutable_main_issues.items()
    }
    merged_detail_issues = {
        row: tuple(_deduplicate_issues(issues))
        for row, issues in mutable_detail_issues.items()
    }
    merged_links = ManualGroupLinks(
        defects_by_main_excel_row=links.defects_by_main_excel_row,
        detail_rows_by_main_excel_row=links.detail_rows_by_main_excel_row,
        main_issues=merged_main_issues,
        detail_issues=merged_detail_issues,
    )
    return _PreparedRows(
        main_values=main_values,
        main_rows=main_rows,
        main_issues=merged_main_issues,
        detail_values=detail_values,
        detail_rows=detail_rows,
        detail_issues=merged_detail_issues,
        links=merged_links,
    )


def _normalized_main_rows(
    workbook, contract: WorkbookContract,
) -> list[tuple[int, dict[str, object]]]:
    source_rows = _populated_rows(
        workbook['Batch Input & Results'], contract.input_headers,
    )
    normalized: list[tuple[int, dict[str, object]]] = []
    for excel_row, source_values in source_rows:
        values = {header: source_values.get(header) for header in INPUT_HEADERS}
        if contract.is_legacy:
            mechanism = normalize_upload_mechanism(source_values.get('Mechanism'))
            if (
                mechanism == 'Corrosion'
                and source_values.get('Defect Location') == 'External'
            ):
                values['Defect Length Basis'] = ACTUAL_DEFECT_LENGTH
            values['Repair Group ID'] = None
        normalized.append((excel_row, values))
    return normalized


def _append_unique_issue(
    issues_by_row: dict[int, list[ValidationIssue]],
    row_number: int,
    issue: ValidationIssue,
) -> None:
    existing = issues_by_row.setdefault(row_number, [])
    if issue not in existing:
        existing.append(issue)


def _deduplicate_issues(
    issues: list[ValidationIssue],
) -> list[ValidationIssue]:
    result: list[ValidationIssue] = []
    for issue in issues:
        if issue not in result:
            result.append(issue)
    return result


def _calculate_one(
    batch_info: BatchInfo,
    excel_row: int,
    values: dict[str, object],
    *,
    individual_defects=(),
) -> RowCalculation:
    row, issues = validate_row(excel_row, values)
    if issues:
        return _input_error_calculation(excel_row, issues)
    assert row is not None
    try:
        if individual_defects:
            return calculate_row(
                batch_info, row, individual_defects=tuple(individual_defects),
            )
        return calculate_row(batch_info, row)
    except Exception as exc:
        frames = ' -> '.join(
            f'{Path(frame.filename).name}:{frame.lineno} in {frame.name}'
            for frame in traceback.extract_tb(exc.__traceback__)
        ) or '(no traceback frames)'
        logger.error(
            'Unexpected %s for source Excel row %s. Traceback frames: %s',
            type(exc).__name__, excel_row, frames,
        )
        return RowCalculation(
            source_excel_row=excel_row,
            status=CalculationStatus.SYSTEM_ERROR,
            outputs={},
            error_code='SYSTEM_ERROR',
            error_message='Unexpected calculation error. Please contact PROTAP.',
        )


def _input_error_calculation(
    excel_row: int,
    issues: tuple[ValidationIssue, ...],
) -> RowCalculation:
    return RowCalculation(
        source_excel_row=excel_row,
        status=CalculationStatus.INPUT_ERROR,
        outputs={},
        error_code=_issue_codes(issues),
        error_message=_issue_message(issues),
    )


def _write_result_row(
    worksheet,
    excel_row: int,
    calculation: RowCalculation,
) -> None:
    for column, heading in enumerate(OUTPUT_HEADERS, start=len(INPUT_HEADERS) + 1):
        worksheet.cell(excel_row, column).value = calculation.outputs.get(heading)


def _write_detail_result_row(
    worksheet,
    excel_row: int,
    *,
    issues: tuple[ValidationIssue, ...],
    candidate: CandidateCalculation | None,
) -> None:
    outputs: dict[str, object] = {
        'Source Excel Row': excel_row,
        'Calculation Status': (
            CalculationStatus.INPUT_ERROR.value if issues else CalculationStatus.OK.value
        ),
        'Error Code': _issue_codes(issues),
        'Error Message': _issue_message(issues),
    }
    if candidate is not None and not issues:
        outputs.update({
            'B31G Method': candidate.method,
            'B31G d/t': candidate.d_over_t,
            'B31G Length Parameter z': candidate.length_parameter_z,
            'B31G Folias Factor M': candidate.folias_factor,
            'B31G Flow Stress [MPa]': candidate.flow_stress_mpa,
            'B31G Estimated Failure Stress [MPa]': candidate.failure_stress_mpa,
            'B31G Failure Pressure [bar]': candidate.failure_pressure_bar,
            'B31G Safe Pressure [bar]': candidate.safe_pressure_bar,
            'B31G Safety Factor': candidate.safety_factor,
            'B31G Operating Hoop Stress [MPa]': (
                candidate.operating_hoop_stress_mpa
            ),
            'B31G Applicable': candidate.applicable,
            'B31G Acceptable': candidate.acceptable,
            'Credited Safe Pressure [bar]': candidate.credited_safe_pressure_bar,
            'Governing Defect': 'Yes' if candidate.governing else None,
            'Assessment Warning Codes': candidate.warning_codes,
        })
    for column, heading in enumerate(
        DETAIL_OUTPUT_HEADERS, start=len(DETAIL_INPUT_HEADERS) + 1,
    ):
        value = normalize_audit_scalar(outputs.get(heading))
        if heading == 'Assessment Warning Codes' and isinstance(value, (tuple, list)):
            value = ', '.join(str(item) for item in value)
        worksheet.cell(excel_row, column).value = value


def _linked_main_row_for_detail(
    detail_excel_row: int,
    links: ManualGroupLinks,
) -> int | None:
    for main_excel_row, detail_rows in links.detail_rows_by_main_excel_row.items():
        if any(row.source_excel_row == detail_excel_row for row in detail_rows):
            return main_excel_row
    return None


def _write_cost_sheet(workbook) -> None:
    source = workbook['Batch Input & Results']
    cost = workbook['Cost Calculation']
    all_headers = INPUT_HEADERS + OUTPUT_HEADERS
    source_columns = {
        header: all_headers.index(header) + 1 for header in COST_SOURCE_HEADERS
    }
    populated = _populated_rows(source)
    for output_row, (source_row, _) in enumerate(populated, start=COST_FIRST_DATA_ROW):
        for destination_column, header in enumerate(COST_SOURCE_HEADERS, start=1):
            cell = cost.cell(output_row, destination_column)
            cell.value = source.cell(source_row, source_columns[header]).value
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = Border(bottom=source['A2'].border.bottom)
        cost.cell(output_row, 21).value = cost_formula(output_row)
        cost.cell(output_row, 22).value = price_formula(output_row)
        cost.cell(output_row, 24).value = total_amount_formula(output_row)
        for column in (21, 22, 24):
            cell = cost.cell(output_row, column)
            cell.alignment = Alignment(vertical='top')
            cell.border = Border(bottom=source['A2'].border.bottom)
            cell.number_format = '#,##0.00'
    table = cost.tables['CostRows']
    table.ref = f'A5:X{max(COST_FIRST_DATA_ROW, 5 + len(populated))}'
    table.autoFilter.ref = table.ref


def _write_warnings_sheet(
    workbook,
    calculations: dict[int, RowCalculation],
) -> None:
    """Build one consolidated, permanent warning register for the batch."""
    warnings_sheet = workbook['Warnings']
    data_sheet = workbook['Batch Input & Results']
    detail_sheet = workbook['Individual Defects']
    affected_rows: dict[str, dict[str, list[int]]] = {}

    def collect(
        worksheet,
        warning_column: int,
        source_row_column: int,
        location: str,
        input_headers: tuple[str, ...],
        max_rows: int,
    ) -> None:
        for excel_row, _ in _populated_rows(
            worksheet, input_headers, max_rows=max_rows,
        ):
            value = worksheet.cell(excel_row, warning_column).value
            if not isinstance(value, str) or not value.strip():
                continue
            source_row = worksheet.cell(excel_row, source_row_column).value
            for code in (item.strip() for item in value.split(',')):
                if not code:
                    continue
                warning_meaning(code)
                rows = affected_rows.setdefault(code, {'main': [], 'detail': []})[location]
                if source_row not in rows:
                    rows.append(source_row)

    for calculation in calculations.values():
        warning_codes = calculation.outputs.get('Compliance Warnings', ())
        if isinstance(warning_codes, str):
            warning_codes = tuple(
                item.strip() for item in warning_codes.split(',') if item.strip()
            )
        if not isinstance(warning_codes, (tuple, list)):
            continue
        for code in warning_codes:
            if not isinstance(code, str) or not code.strip():
                continue
            code = code.strip()
            warning_meaning(code)
            rows = affected_rows.setdefault(code, {'main': [], 'detail': []})['main']
            if calculation.source_excel_row not in rows:
                rows.append(calculation.source_excel_row)
    collect(
        detail_sheet,
        len(DETAIL_INPUT_HEADERS) + DETAIL_OUTPUT_HEADERS.index('Assessment Warning Codes') + 1,
        len(DETAIL_INPUT_HEADERS) + DETAIL_OUTPUT_HEADERS.index('Source Excel Row') + 1,
        'detail',
        DETAIL_INPUT_HEADERS,
        MAX_DETAIL_ROWS,
    )

    if not affected_rows:
        return

    warnings_sheet['A4'] = None
    for output_row, code in enumerate(sorted(affected_rows), start=4):
        warnings_sheet.cell(output_row, 1, code)
        warnings_sheet.cell(output_row, 2, warning_meaning(code))
        warnings_sheet.cell(
            output_row,
            3,
            format_affected_rows(
                affected_rows[code]['main'], affected_rows[code]['detail'],
            ),
        )
        for column in range(1, 4):
            cell = warnings_sheet.cell(output_row, column)
            cell.alignment = Alignment(vertical='top', wrap_text=column == 2)
            cell.border = Border(bottom=data_sheet['A2'].border.bottom)
            cell.font = Font(name='Calibri', size=11, color='000000')

    last_row = 3 + len(affected_rows)
    table = Table(displayName='WarningRegister', ref=f'A3:C{last_row}')
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    warnings_sheet.add_table(table)


def _write_summary(
    workbook,
    batch_info: BatchInfo,
    timestamp: str,
    populated_rows: int,
    calculations: dict[int, RowCalculation],
    source_name: str,
) -> None:
    summary = workbook['Summary']
    summary['B3'] = batch_info.customer
    summary['B4'] = batch_info.project_location
    summary['B5'] = batch_info.report_no
    summary['B7'] = source_name
    summary['B8'] = timestamp
    summary['B10'] = populated_rows
    status_counts = Counter(
        calculation.status.value for calculation in calculations.values()
    )
    for row, status in enumerate(CalculationStatus, start=13):
        summary.cell(row, 2).value = status_counts.get(status.value, 0)

    methods = [
        calculation.outputs.get('Thickness Calculation Method')
        for calculation in calculations.values()
    ]
    summary['B19'] = sum(_is_type_a(method) for method in methods)
    summary['B20'] = sum(_is_type_b(method) for method in methods)
    summary['B21'] = sum(
        bool(calculation.outputs.get('Compliance Warnings'))
        for calculation in calculations.values()
    )
    summary['B22'] = sum(
        calculation.status.value in {
            CalculationStatus.REVIEW_REQUIRED.value,
            CalculationStatus.NOT_REPAIRABLE.value,
        }
        for calculation in calculations.values()
    )
    summary['B24'] = BATCH_ENGINE_VERSION
    summary['B25'] = SOURCE_ENGINE_REVISION


def _utc_timestamp(value: datetime) -> str:
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.astimezone(UTC).replace(microsecond=0).isoformat().replace('+00:00', 'Z')


def _empty_inspection(
    errors: tuple[ValidationIssue, ...],
    header_summary: tuple[tuple[str, ...], tuple[str, ...], tuple[str, ...]] = ((), (), ()),
    detail_header_summary: tuple[
        tuple[str, ...], tuple[str, ...], tuple[str, ...]
    ] = ((), (), ()),
) -> WorkbookInspection:
    return WorkbookInspection(
        None, 0, 0, 0, errors, (), *header_summary,
        0, 0, *detail_header_summary,
    )


def _issue(code: str, message: str) -> ValidationIssue:
    return ValidationIssue(code=code, message=message)


def _issue_codes(issues: tuple[ValidationIssue, ...]) -> str:
    return '; '.join(issue.code for issue in issues)


def _issue_message(issues: tuple[ValidationIssue, ...]) -> str:
    return ' '.join(issue.message for issue in issues)


def _first_duplicate(values: tuple[object, ...]) -> object | None:
    seen: set[object] = set()
    for value in values:
        if value in seen:
            return value
        seen.add(value)
    return None


def _display_heading(value: object) -> str:
    if value is None or (isinstance(value, str) and not value.strip()):
        return '(blank)'
    return str(value)


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _is_formula(value: object) -> bool:
    return isinstance(value, str) and value.lstrip().startswith('=')


def _is_formula_cell(cell) -> bool:
    return getattr(cell, 'data_type', None) == 'f' or _is_formula(cell.value)


def _loaded_cells(worksheet):
    """Yield parsed cells without expanding sparse worksheet dimensions.

    ``Worksheet.iter_rows()`` allocates every coordinate in the declared
    worksheet dimension.  This private openpyxl collection is deliberately
    isolated here so validation examines only cells that were actually loaded
    from the workbook, including far-away formula and input cells.
    """
    return sorted(worksheet._cells.items())


def _zip_safety_errors(archive: zipfile.ZipFile) -> tuple[ValidationIssue, ...]:
    entries = archive.infolist()
    if len(entries) > _MAX_ZIP_ENTRIES:
        return (_issue('UNREADABLE_WORKBOOK', 'The uploaded workbook has too many compressed entries.'),)
    total_uncompressed = 0
    for entry in entries:
        if entry.file_size > _MAX_ZIP_ENTRY_BYTES:
            return (_issue('UNREADABLE_WORKBOOK', 'The uploaded workbook has an oversized compressed entry.'),)
        total_uncompressed += entry.file_size
        if total_uncompressed > _MAX_ZIP_UNCOMPRESSED_BYTES:
            return (_issue('UNREADABLE_WORKBOOK', 'The uploaded workbook expands beyond the safe processing limit.'),)
        if (
            entry.file_size > 0
            and entry.compress_size > 0
            and entry.file_size / entry.compress_size > _MAX_ZIP_COMPRESSION_RATIO
        ):
            return (_issue('UNREADABLE_WORKBOOK', 'The uploaded workbook has a suspicious compression ratio.'),)
    worksheet_parts, declarations_are_consistent = _worksheet_part_names(
        archive, entries,
    )
    if not declarations_are_consistent:
        return (_issue(
            'UNREADABLE_WORKBOOK',
            'The uploaded workbook has inconsistent worksheet declarations.',
        ),)
    worksheet_cells = 0
    for entry in entries:
        if (
            entry.flag_bits & 0x1
            or entry.filename not in worksheet_parts
        ):
            continue
        with archive.open(entry) as worksheet_xml:
            elements = iterparse(worksheet_xml, events=('start', 'end'))
            try:
                event, root = next(elements)
            except StopIteration:
                continue
            root_namespace, root_local_name = _xml_expanded_name(root.tag)
            if (
                event != 'start'
                or root_local_name != 'worksheet'
                or root_namespace not in _SPREADSHEETML_MAIN_NAMESPACES
            ):
                root.clear()
                continue
            for event, element in elements:
                namespace, local_name = _xml_expanded_name(element.tag)
                if (
                    event == 'end'
                    and local_name == 'c'
                    and namespace == root_namespace
                ):
                    worksheet_cells += 1
                    if worksheet_cells > _MAX_WORKBOOK_CELLS:
                        return (_issue(
                            'UNREADABLE_WORKBOOK',
                            'The uploaded workbook contains too many worksheet cells.',
                        ),)
                element.clear()
    return ()


def _worksheet_part_names(
    archive: zipfile.ZipFile,
    entries: list[zipfile.ZipInfo],
) -> tuple[frozenset[str], bool]:
    """Resolve workbook sheet targets and verify their OPC content types."""
    entry_names = {entry.filename for entry in entries}
    declarations = _content_type_declarations(archive)
    if declarations is None:
        return frozenset(), False
    overrides, defaults = declarations

    root_relationships = _relationship_records(archive, '_rels/.rels')
    if root_relationships is None:
        return frozenset(), False
    workbook_relationships = [
        record for record in root_relationships.values()
        if record[0] in _OFFICE_DOCUMENT_RELATIONSHIP_TYPES
    ]
    if len(workbook_relationships) != 1:
        return frozenset(), False
    _, workbook_target, workbook_target_mode = workbook_relationships[0]
    workbook_part = _resolve_relationship_target('', workbook_target)
    if (
        not _is_internal_target_mode(workbook_target_mode)
        or workbook_part not in entry_names
    ):
        return frozenset(), False

    sheet_relationship_ids = _workbook_sheet_relationship_ids(
        archive, workbook_part,
    )
    relationships_part = _relationships_part_name(workbook_part)
    workbook_part_relationships = _relationship_records(
        archive, relationships_part,
    )
    if workbook_part_relationships is None:
        return frozenset(), False
    worksheet_parts: set[str] = set()
    for relationship_id in sheet_relationship_ids:
        record = workbook_part_relationships.get(relationship_id)
        if record is None:
            return frozenset(), False
        relationship_type, target, target_mode = record
        target_part = _resolve_relationship_target(workbook_part, target)
        if (
            not _is_internal_target_mode(target_mode)
            or relationship_type not in _WORKSHEET_RELATIONSHIP_TYPES
            or target_part not in entry_names
            or _effective_content_type(target_part, overrides, defaults)
            != _WORKSHEET_CONTENT_TYPE
        ):
            return frozenset(), False
        worksheet_parts.add(target_part)
    return frozenset(worksheet_parts), True


def _content_type_declarations(
    archive: zipfile.ZipFile,
) -> tuple[dict[str, str], dict[str, str]] | None:
    entry = archive.getinfo('[Content_Types].xml')
    if entry.flag_bits & 0x1:
        return None
    overrides: dict[str, str] = {}
    defaults: dict[str, str] = {}
    with archive.open(entry) as content_types_xml:
        for _, element in iterparse(content_types_xml, events=('end',)):
            _, local_name = _xml_expanded_name(element.tag)
            content_type = element.attrib.get('ContentType')
            if local_name == 'Override':
                part_name = element.attrib.get('PartName', '').lstrip('/')
                if not part_name or not content_type:
                    return None
                if part_name in overrides and overrides[part_name] != content_type:
                    return None
                overrides[part_name] = content_type
            elif local_name == 'Default':
                extension = element.attrib.get('Extension', '').lower()
                if not extension or not content_type:
                    return None
                if extension in defaults and defaults[extension] != content_type:
                    return None
                defaults[extension] = content_type
            element.clear()
    return overrides, defaults


def _relationship_records(
    archive: zipfile.ZipFile,
    part_name: str,
) -> dict[str, tuple[str, str, str]] | None:
    entry = archive.getinfo(part_name)
    if entry.flag_bits & 0x1:
        return {}
    records: dict[str, tuple[str, str, str]] = {}
    with archive.open(entry) as relationships_xml:
        for _, element in iterparse(relationships_xml, events=('end',)):
            namespace, local_name = _xml_expanded_name(element.tag)
            if (
                namespace in _OPC_RELATIONSHIPS_NAMESPACES
                and local_name == 'Relationship'
            ):
                relationship_id = element.attrib.get('Id', '')
                relationship_type = element.attrib.get('Type', '')
                target = element.attrib.get('Target', '')
                target_mode = element.attrib.get('TargetMode', '')
                if (
                    not relationship_id
                    or not relationship_type
                    or not target
                    or relationship_id in records
                ):
                    return None
                records[relationship_id] = (
                    relationship_type, target, target_mode,
                )
            element.clear()
    return records


def _is_internal_target_mode(target_mode: str) -> bool:
    return not target_mode or target_mode.casefold() == 'internal'


def _workbook_sheet_relationship_ids(
    archive: zipfile.ZipFile,
    workbook_part: str,
) -> tuple[str, ...]:
    entry = archive.getinfo(workbook_part)
    if entry.flag_bits & 0x1:
        return ()
    relationship_ids: list[str] = []
    with archive.open(entry) as workbook_xml:
        for _, element in iterparse(workbook_xml, events=('end',)):
            _, local_name = _xml_expanded_name(element.tag)
            if local_name == 'sheet':
                relationship_id = next((
                    value for attribute, value in element.attrib.items()
                    if _xml_expanded_name(attribute)[1] == 'id'
                ), '')
                if relationship_id:
                    relationship_ids.append(relationship_id)
            element.clear()
    return tuple(relationship_ids)


def _relationships_part_name(source_part: str) -> str:
    source = PurePosixPath(source_part)
    return str(source.parent / '_rels' / f'{source.name}.rels')


def _resolve_relationship_target(
    source_part: str,
    target: str,
) -> str:
    if target.startswith('/'):
        unresolved = target.lstrip('/')
    else:
        unresolved = posixpath.join(posixpath.dirname(source_part), target)
    resolved = posixpath.normpath(unresolved)
    if resolved in {'', '.', '..'} or resolved.startswith('../'):
        return ''
    return resolved


def _effective_content_type(
    part_name: str,
    overrides: dict[str, str],
    defaults: dict[str, str],
) -> str | None:
    if part_name in overrides:
        return overrides[part_name]
    extension = PurePosixPath(part_name).suffix.lstrip('.').lower()
    return defaults.get(extension)


def _xml_expanded_name(tag: object) -> tuple[str, str]:
    if not isinstance(tag, str):
        return '', ''
    if tag.startswith('{'):
        namespace, separator, local_name = tag[1:].partition('}')
        if separator:
            return namespace, local_name
    return '', tag


def _copy_controlled_inputs(
    source_workbook,
    output_workbook,
    contract: WorkbookContract,
) -> None:
    source_info = source_workbook['Batch Information']
    output_info = output_workbook['Batch Information']
    for row in range(3, 6):
        output_info.cell(row, 2).value = source_info.cell(row, 2).value

    source_data = source_workbook['Batch Input & Results']
    output_data = output_workbook['Batch Input & Results']
    source_columns = {
        header: column
        for column, header in enumerate(contract.input_headers, start=1)
    }
    output_columns = {
        header: column for column, header in enumerate(INPUT_HEADERS, start=1)
    }
    for excel_row in range(2, MAX_ROWS + 2):
        populated = any(
            not _is_blank(source_data.cell(excel_row, column).value)
            for column in source_columns.values()
        )
        for header, output_column in output_columns.items():
            source_column = source_columns.get(header)
            value = (
                source_data.cell(excel_row, source_column).value
                if source_column is not None else None
            )
            if header == 'Mechanism':
                value = normalize_upload_mechanism(value)
            elif contract.is_legacy and header == 'Defect Length Basis':
                mechanism = normalize_upload_mechanism(
                    source_data.cell(
                        excel_row, source_columns['Mechanism'],
                    ).value
                )
                location = source_data.cell(
                    excel_row, source_columns['Defect Location'],
                ).value
                value = (
                    ACTUAL_DEFECT_LENGTH
                    if populated and mechanism == 'Corrosion' and location == 'External'
                    else None
                )
            elif contract.is_legacy and header == 'Repair Group ID':
                value = None
            output_data.cell(excel_row, output_column).value = value

    if contract.has_individual_defects:
        source_detail = source_workbook['Individual Defects']
        output_detail = output_workbook['Individual Defects']
        source_detail_columns = {
            header: column
            for column, header in enumerate(DETAIL_INPUT_HEADERS, start=1)
        }
        for excel_row in range(2, MAX_DETAIL_ROWS + 2):
            for header, source_column in source_detail_columns.items():
                output_detail.cell(excel_row, source_column).value = source_detail.cell(
                    excel_row, source_column,
                ).value

    if 'Cost Calculation' in source_workbook.sheetnames:
        source_cost = source_workbook['Cost Calculation']
        output_cost = output_workbook['Cost Calculation']
        for address, _ in COST_INPUTS:
            source_value = source_cost[address].value
            output_cost[address].value = None if _is_blank(source_value) else source_value
        populated_rows = _populated_rows(source_data, contract.input_headers)
        for cost_row, _ in enumerate(populated_rows, start=COST_FIRST_DATA_ROW):
            quantity = source_cost.cell(cost_row, 23).value
            output_cost.cell(cost_row, 23).value = (
                None if _is_blank(quantity) else quantity
            )


def _sanitized_source_name(source_name: str | None) -> str:
    if not source_name:
        return 'PROWRAP Batch Results Workbook'
    filename = Path(str(source_name).replace('\\', '/')).name
    clean = ''.join(character for character in filename if character.isprintable()).strip()
    if not clean:
        return 'PROWRAP Batch Results Workbook'
    if clean.lstrip().startswith(('=', '+', '-', '@')):
        return f"'{clean}"
    return clean


def _is_type_a(method: object) -> bool:
    return isinstance(method, str) and 'type a' in method.lower()


def _is_type_b(method: object) -> bool:
    return isinstance(method, str) and 'type b' in method.lower()
