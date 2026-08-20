"""Create the controlled six-row source workbook for 150-row batch acceptance checks."""

from __future__ import annotations

import argparse
from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook


if __package__ in {None, ''}:
    sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from batch_schema import DETAIL_INPUT_HEADERS, INPUT_HEADERS
from workbook_template import create_template_workbook


_COMMON_VALUES = {
    'Customer': 'Acceptance Customer',
    'Project Location': 'Acceptance Location',
    'Report No': 'ACCEPT-001',
}

_CORROSION_COMPARISON_ROW = {
    'Pipe OD [mm]': 1016.0,
    'Nominal Wall [mm]': 12.0,
    'Pipe Yield [MPa]': 450.0,
    'Design Pressure [bar]': 104.9,
    'Operating Temperature [degC]': 40.0,
    'Mechanism': 'Corrosion',
    'Defect Location': 'External',
    'Defect Length [mm]': 1000.0,
    'Defect Length Basis': 'Actual defect length',
    'Repair Group ID': None,
    'Remaining Wall [mm]': 9.652,
    'Internal Corrosion Rate [mm/year]': None,
    'Design Life [years]': 20,
    'Design Factor': 0.72,
    'Run Type A / Class 3 Check': 'No',
    'Installation Temperature [degC]': 20.0,
    'Component Type': 'Straight',
    'Cyclic Derating Factor': 1.0,
    'Axial Load Case': 0,
    'Prowrap CF Cloth Width [mm]': 500.0,
}

_DENT_ROW = {
    **_CORROSION_COMPARISON_ROW,
    'Pipe OD [mm]': 457.2,
    'Nominal Wall [mm]': 9.53,
    'Pipe Yield [MPa]': 359.0,
    'Design Pressure [bar]': 50.0,
    'Defect Length [mm]': 100.0,
    'Defect Length Basis': None,
    'Remaining Wall [mm]': 9.53,
}


def create_acceptance_workbook(destination: str | Path) -> Path:
    """Write the fixed acceptance input workbook to the requested destination only."""
    path = Path(destination)
    workbook = load_workbook(BytesIO(create_template_workbook()))
    try:
        info = workbook['Batch Information']
        for row, label in enumerate(_COMMON_VALUES, start=3):
            info.cell(row, 2).value = _COMMON_VALUES[label]

        data = workbook['Batch Input & Results']
        for excel_row, values in enumerate(_acceptance_rows(), start=2):
            for column, header in enumerate(INPUT_HEADERS, start=1):
                data.cell(excel_row, column).value = values[header]

        detail = workbook['Individual Defects']
        for excel_row, values in enumerate(_acceptance_details(), start=2):
            for column, header in enumerate(DETAIL_INPUT_HEADERS, start=1):
                detail.cell(excel_row, column).value = values[header]
        workbook.save(path)
    finally:
        workbook.close()
    return path


def _acceptance_rows() -> tuple[dict[str, object], ...]:
    return (
        {
            **_CORROSION_COMPARISON_ROW,
        },
        {
            **_CORROSION_COMPARISON_ROW,
            'Defect Length Basis': 'Independent defects',
        },
        {
            **_CORROSION_COMPARISON_ROW,
            'Defect Length Basis': 'Enter manually',
            'Repair Group ID': 'R-001',
            'Remaining Wall [mm]': None,
        },
        {
            **_CORROSION_COMPARISON_ROW,
            'Defect Length Basis': 'Enter manually',
            'Repair Group ID': 'R-BAD',
            'Remaining Wall [mm]': None,
        },
        {
            **_DENT_ROW,
            'Mechanism': 'Dent no-crack',
        },
        {**_DENT_ROW, 'Mechanism': 'Dent w/crack'},
    )


def _acceptance_details() -> tuple[dict[str, object], ...]:
    return (
        {
            'Repair Group ID': 'R-001',
            'Defect ID': 'D-01',
            'Individual longitudinal length [mm]': 10.0,
            'Remaining wall [mm]': 9.652,
            'Separation exceeds 3t': 'Yes',
        },
        {
            'Repair Group ID': 'R-001',
            'Defect ID': 'D-02',
            'Individual longitudinal length [mm]': 35.0,
            'Remaining wall [mm]': 10.0,
            'Separation exceeds 3t': 'Yes',
        },
        {
            'Repair Group ID': 'R-BAD',
            'Defect ID': 'D-BAD',
            'Individual longitudinal length [mm]': 10.0,
            'Remaining wall [mm]': 9.652,
            'Separation exceeds 3t': 'No',
        },
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description='Create the six-row PROWRAP CalcBatch v1.2 acceptance workbook.',
    )
    parser.add_argument('destination', type=Path, help='Path for the generated .xlsx workbook.')
    args = parser.parse_args()
    create_acceptance_workbook(args.destination)


if __name__ == '__main__':
    main()
