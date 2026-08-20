"""User interface for the standalone PROWRAP batch workbook calculator."""

from __future__ import annotations

from datetime import UTC, datetime
import hashlib
from io import BytesIO

from openpyxl import load_workbook
import streamlit as st

from batch_schema import DETAIL_INPUT_HEADERS
from workbook_processor import WorkbookProcessingError, inspect_workbook, process_workbook
from workbook_template import create_template_workbook


_TEMPLATE_FILENAME = 'PROWRAP_CalcBatch_v1.2_Template.xlsx'
_PROCESSED_BYTES_KEY = 'processed_workbook_bytes'
_PROCESSED_IDENTITY_KEY = 'processed_source_identity'
_PROCESSED_NAME_KEY = 'processed_workbook_name'
_SOURCE_IDENTITY_KEY = 'source_identity'


def _clear_processed_result() -> None:
    for key in (_PROCESSED_BYTES_KEY, _PROCESSED_IDENTITY_KEY, _PROCESSED_NAME_KEY):
        st.session_state.pop(key, None)


def _source_identity(data: bytes, filename: str) -> str:
    """Bind uploaded workbook bytes to the exact filename used for processing."""
    digest = hashlib.sha256()
    digest.update(data)
    digest.update(b'\0')
    digest.update(filename.encode('utf-8'))
    return digest.hexdigest()


def _output_filename(processed_at: datetime) -> str:
    return (
        'PROWRAP_CalcBatch_v1.2_Results_'
        f"{processed_at.strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


def _show_workbook_errors(issues) -> None:
    st.error('This workbook needs correction before it can be calculated.')
    for issue in issues:
        st.write(f'• {issue.message}')


def _show_header_summary(inspection) -> None:
    """Show what the uploaded workbook contains without accepting altered headings."""
    main_recognized = ', '.join(inspection.recognized_input_headers) or 'None'
    main_missing = ', '.join(inspection.missing_input_headers) or 'None'
    main_unexpected = ', '.join(inspection.unexpected_headers) or 'None'
    detail_recognized = ', '.join(inspection.recognized_detail_input_headers) or 'None'
    detail_missing = ', '.join(inspection.missing_detail_input_headers) or 'None'
    detail_unexpected = ', '.join(inspection.unexpected_detail_headers) or 'None'
    st.write(
        'Recognized Batch Input & Results input columns '
        f'({len(inspection.recognized_input_headers)}): {main_recognized}'
    )
    st.write(f'Missing Batch Input & Results input columns: {main_missing}')
    st.write(f'Unexpected Batch Input & Results headings: {main_unexpected}')
    st.write(
        'Recognized Individual Defects input columns '
        f'({len(inspection.recognized_detail_input_headers)}): {detail_recognized}'
    )
    st.write(f'Missing Individual Defects input columns: {detail_missing}')
    st.write(f'Unexpected Individual Defects headings: {detail_unexpected}')


def _show_status_counts(status_counts: dict[str, int]) -> None:
    ordered_statuses = (
        'OK',
        'REVIEW REQUIRED',
        'NOT REPAIRABLE',
        'INPUT ERROR',
        'SYSTEM ERROR',
    )
    columns = st.columns(len(ordered_statuses))
    for column, status in zip(columns, ordered_statuses):
        column.metric(status, status_counts.get(status, 0))


def _detail_preview(data: bytes) -> list[dict[str, object]]:
    """Return the first populated linked-defect rows after controlled inspection."""
    workbook = load_workbook(BytesIO(data), read_only=True, data_only=False)
    try:
        worksheet = workbook['Individual Defects']
        preview: list[dict[str, object]] = []
        for values in worksheet.iter_rows(
            min_row=2,
            max_col=len(DETAIL_INPUT_HEADERS),
            values_only=True,
        ):
            if any(value not in (None, '') for value in values):
                preview.append(dict(zip(DETAIL_INPUT_HEADERS, values)))
            if len(preview) == 20:
                break
        return preview
    finally:
        workbook.close()


def main() -> None:
    st.set_page_config(page_title='PROWRAP CalcBatch v1.2', layout='wide')
    st.title('PROWRAP CalcBatch v1.2')
    st.write(
        'Calculate up to 500 independent pipeline defects from one controlled Excel workbook. '
        'Customer, Project Location, and Report No are entered once for the whole batch.'
    )
    st.info(
        'This is a separate batch calculator. It does not change, replace, or connect to '
        'the existing PROWRAP v1.1 calculator.'
    )

    st.subheader('1. Download template')
    st.download_button(
        'Download Excel Template',
        data=create_template_workbook(),
        file_name=_TEMPLATE_FILENAME,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        help='Use this controlled workbook so the batch calculator can recognize every input.',
    )
    st.caption(
        'Use millimetres, MPa, bar, degC, years, m2, and kg as labelled. '
        'Supported mechanisms: Corrosion, Dent w/crack, Dent no-crack, Leak, and Crack.'
    )
    st.caption(
        'Old controlled batch workbooks containing generic Dent are interpreted '
        'conservatively as Dent w/crack.'
    )
    st.caption(
        'For Enter manually, use a stable Repair Group ID on the main repair row '
        'and matching rows with the same ID on the Individual Defects sheet. '
        'Leave the main Remaining Wall cell blank in this mode.'
    )

    st.subheader('2. Upload workbook')
    uploaded = st.file_uploader(
        'Upload the completed Excel template',
        type=['xlsx'],
        help='Upload one controlled .xlsx workbook, up to 10 MB. '
        'Macros and uncontrolled formulas are rejected. Exact controlled Cost and Price '
        'formulas in previously processed workbooks are accepted.',
    )

    inspection = None
    source_data = None
    source_identity = None
    if uploaded is not None:
        source_data = uploaded.getvalue()
        source_identity = _source_identity(source_data, uploaded.name)
        if st.session_state.get(_SOURCE_IDENTITY_KEY) != source_identity:
            st.session_state[_SOURCE_IDENTITY_KEY] = source_identity
            _clear_processed_result()

        st.caption(f'Uploaded file: {uploaded.name}')
        try:
            inspection = inspect_workbook(source_data)
        except Exception:
            st.error('The workbook could not be read safely. Please download a fresh template and try again.')
        else:
            _show_header_summary(inspection)
            if inspection.workbook_errors:
                _show_workbook_errors(inspection.workbook_errors)
            else:
                main_label = 'row' if inspection.populated_rows == 1 else 'rows'
                detail_label = (
                    'row' if inspection.populated_detail_rows == 1 else 'rows'
                )
                group_label = 'group' if inspection.manual_groups == 1 else 'groups'
                st.success(
                    f'Recognized {inspection.populated_rows} populated repair {main_label}: '
                    f'{inspection.valid_rows} valid and {inspection.invalid_rows} needing correction.'
                )
                st.write(
                    f'{inspection.populated_detail_rows} populated individual-defect {detail_label}; '
                    f'{inspection.manual_groups} manual repair {group_label}.'
                )

    st.subheader('3. Review preview')
    if inspection is None:
        st.write('Upload a workbook to see the first 20 populated rows and their validation status.')
    elif not inspection.workbook_errors:
        if inspection.preview:
            st.caption('Main repair-row preview (first 20 populated rows)')
            st.dataframe(list(inspection.preview), hide_index=True, width='stretch')
        if inspection.populated_detail_rows:
            st.caption('Individual-defect preview (first 20 populated rows)')
            st.dataframe(_detail_preview(source_data), hide_index=True, width='stretch')
        else:
            st.warning('No populated defect rows were found. A processed workbook can still be created, but it will contain no row results.')

    can_calculate = (
        source_data is not None
        and inspection is not None
        and not inspection.workbook_errors
    )
    calculate = st.button(
        'Calculate Batch',
        type='primary',
        disabled=not can_calculate,
        help='Calculation is available when the controlled workbook structure is valid. '
        'Rows with input errors remain individually reported.',
    )
    if calculate and source_data is not None:
        try:
            processed_at = datetime.now(UTC)
            processed = process_workbook(
                source_data,
                processed_at=processed_at,
                source_name=uploaded.name,
            )
        except WorkbookProcessingError as error:
            _show_workbook_errors(error.issues)
        except Exception:
            st.error('The batch could not be calculated safely. Please try a fresh template or contact PROTAP.')
        else:
            st.session_state[_PROCESSED_BYTES_KEY] = processed.workbook_bytes
            st.session_state[_PROCESSED_IDENTITY_KEY] = source_identity
            st.session_state[_PROCESSED_NAME_KEY] = _output_filename(processed_at)

    processed_bytes = st.session_state.get(_PROCESSED_BYTES_KEY)
    if processed_bytes and st.session_state.get(_PROCESSED_IDENTITY_KEY) == source_identity:
        st.subheader('4. Calculate and download')
        st.success('Batch calculation is complete. Review the status counts and download the processed workbook.')
        if inspection is not None and not inspection.workbook_errors:
            # The workbook itself is the source of truth for detailed row outputs.
            # Reprocessing is deliberately avoided; the result is stored only for this session.
            processed_workbook = load_workbook(BytesIO(processed_bytes), read_only=True, data_only=True)
            summary = processed_workbook['Summary']
            _show_status_counts({
                'OK': summary['B13'].value or 0,
                'REVIEW REQUIRED': summary['B14'].value or 0,
                'NOT REPAIRABLE': summary['B15'].value or 0,
                'INPUT ERROR': summary['B16'].value or 0,
                'SYSTEM ERROR': summary['B17'].value or 0,
            })
            processed_workbook.close()
        st.download_button(
            'Download Processed Workbook',
            data=processed_bytes,
            file_name=st.session_state[_PROCESSED_NAME_KEY],
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            type='primary',
        )


if __name__ == '__main__':
    main()
