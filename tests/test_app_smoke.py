from pathlib import Path
from io import BytesIO

from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

import app as batch_app
from batch_schema import INPUT_HEADERS
from engine.corrosion_defects import ENTER_MANUALLY
from tests.helpers import detail_values, valid_row_values, workbook_bytes_with_rows


def test_source_identity_binds_bytes_and_exact_filename():
    """Catch either workbook bytes or the exact upload name being omitted from identity."""
    first = batch_app._source_identity(b'workbook-bytes', 'first.xlsx')

    assert first == 'e6b4a86d252ab928b43095ab4d409b1057eba46ce0ed4aafdebc1fe2c7ec2d58'
    assert batch_app._source_identity(b'changed-bytes', 'first.xlsx') != first
    assert batch_app._source_identity(b'workbook-bytes', 'renamed.xlsx') != first


def test_app_starts_with_template_and_upload_actions():
    """Catch a regression that removes the visible first two workflow stages."""
    app = AppTest.from_file(Path(__file__).parents[1] / 'app.py').run()

    assert not app.exception
    assert any('PROWRAP CalcBatch v1.2' in title.value for title in app.title)
    assert any(
        'Download Excel Template' in button.label for button in app.download_button
    )
    assert batch_app._TEMPLATE_FILENAME == 'PROWRAP_CalcBatch_v1.2_Template.xlsx'
    assert any('Upload workbook' in heading.value for heading in app.subheader)


def test_app_lists_canonical_dents_and_requires_the_current_template():
    """Catches a current release that sends users to an unsupported older template."""
    app = AppTest.from_file(Path(__file__).parents[1] / 'app.py').run()
    captions = [caption.value for caption in app.caption]

    assert any(
        'Supported mechanisms: Corrosion, Dent w/crack, Dent no-crack, Leak, and Crack.'
        in caption
        for caption in captions
    )
    assert not any(
        'Supported mechanisms: Corrosion, Dent, Leak, and Crack.' in caption
        for caption in captions
    )
    assert any(
        'Download and use the current PROWRAP CalcBatch v1.2 150/150 template.'
        in caption
        for caption in captions
    )
    assert not any('older batch workbook' in caption.lower() for caption in captions)


def test_app_states_compact_row_and_commercial_contract():
    """Catches the screen promising obsolete row counts or hiding Quantity totals."""
    app = AppTest.from_file(Path(__file__).parents[1] / 'app.py').run()
    rendered = '\n'.join(item.value for item in (*app.markdown, *app.caption))

    assert 'up to 150 continuous-repair rows and 150 linked individual-defect rows' in rendered
    assert 'Quantity is editable' in rendered
    assert 'Total Amount is a controlled Price x Quantity formula' in rendered


def test_uploader_help_distinguishes_rejected_and_controlled_formulas():
    """Catch guidance that incorrectly says every formula-bearing workbook is rejected."""
    app = AppTest.from_file(Path(__file__).parents[1] / 'app.py').run()
    help_text = app.file_uploader[0].help

    assert 'Macros and uncontrolled formulas are rejected.' in help_text
    assert (
        'Exact controlled Cost and Price formulas in previously processed workbooks are accepted.'
        in help_text
    )


def test_valid_workbook_structure_keeps_calculation_available_with_row_errors():
    """Catch a regression that treats empty or bad data rows as workbook-level errors."""
    app = AppTest.from_file(Path(__file__).parents[1] / 'app.py').run()
    app.file_uploader[0].upload(
        'batch.xlsx',
        workbook_bytes_with_rows([
            valid_row_values(**{'Pipe OD [mm]': None}),
        ]),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    ).run()

    calculate_button = next(button for button in app.button if button.label == 'Calculate Batch')
    assert not calculate_button.proto.disabled
    assert any('1 needing correction' in item.value for item in app.success)


def test_unreadable_upload_has_a_safe_error_without_internal_exception_text():
    """A corrupt file must not leak parser details into the customer-facing screen."""
    app = AppTest.from_file(Path(__file__).parents[1] / 'app.py').run()
    app.file_uploader[0].upload(
        'corrupt.xlsx', b'not an Excel workbook',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    ).run()

    errors = '\n'.join(item.value for item in app.error)
    rendered = '\n'.join(item.value for item in app.markdown)
    assert 'needs correction before it can be calculated' in errors
    assert 'BadZipFile' not in errors + rendered


def test_new_upload_clears_a_previous_processed_download():
    """Catch a regression that offers results belonging to a different upload."""
    app = AppTest.from_file(Path(__file__).parents[1] / 'app.py').run()
    app.file_uploader[0].upload(
        'first.xlsx',
        workbook_bytes_with_rows([valid_row_values()]),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    ).run()
    next(button for button in app.button if button.label == 'Calculate Batch').click().run(timeout=10)

    assert any(
        button.label == 'Download Processed Workbook' for button in app.download_button
    )

    app.file_uploader[0].upload(
        'second.xlsx',
        workbook_bytes_with_rows([valid_row_values(**{'Pipe OD [mm]': 508.0})]),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    ).run()

    assert not any(
        button.label == 'Download Processed Workbook' for button in app.download_button
    )


def test_renamed_identical_upload_clears_old_result_and_records_new_source_name():
    """Catch bytes-only identity reusing results generated under an earlier filename."""
    source = workbook_bytes_with_rows([valid_row_values()])
    app = AppTest.from_file(Path(__file__).parents[1] / 'app.py').run()
    app.file_uploader[0].upload(
        'first.xlsx',
        source,
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    ).run()
    next(button for button in app.button if button.label == 'Calculate Batch').click().run(timeout=10)

    assert any(
        button.label == 'Download Processed Workbook' for button in app.download_button
    )

    app.file_uploader[0].upload(
        'renamed-second.xlsx',
        source,
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    ).run()

    assert not any(
        button.label == 'Download Processed Workbook' for button in app.download_button
    )

    next(button for button in app.button if button.label == 'Calculate Batch').click().run(timeout=10)
    processed = load_workbook(BytesIO(app.session_state['processed_workbook_bytes']))

    assert processed['Summary']['B7'].value == 'renamed-second.xlsx'


def test_header_mismatch_shows_recognized_missing_and_unexpected_columns():
    """Catch a generic upload error that does not tell users how to fix headings."""
    source = workbook_bytes_with_rows([valid_row_values()])
    workbook = load_workbook(BytesIO(source))
    workbook['Batch Input & Results']['A1'] = 'Outside Diameter [mm]'
    changed = BytesIO()
    workbook.save(changed)

    app = AppTest.from_file(Path(__file__).parents[1] / 'app.py').run()
    app.file_uploader[0].upload(
        'changed-heading.xlsx',
        changed.getvalue(),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    ).run()

    rendered = '\n'.join(markdown.value for markdown in app.markdown)
    assert (
        'Recognized Batch Input & Results input columns '
        f'({len(INPUT_HEADERS) - 1})'
    ) in rendered
    assert 'Missing Batch Input & Results input columns: Pipe OD [mm]' in rendered
    assert 'Unexpected Batch Input & Results headings: Outside Diameter [mm]' in rendered
    assert 'Recognized Individual Defects input columns (5)' in rendered


def test_app_previews_linked_manual_counts_and_explains_linkage():
    """Manual repairs need one stable main ID and matching detail-sheet records."""
    source = workbook_bytes_with_rows(
        [valid_row_values(**{
            'Defect Length Basis': ENTER_MANUALLY,
            'Repair Group ID': 'R-001',
            'Remaining Wall [mm]': None,
        })],
        detail_rows=[
            detail_values(group='R-001', defect='D-01', length=10, wall=4.5),
            detail_values(group='R-001', defect='D-02', length=35, wall=4.6),
        ],
    )
    app = AppTest.from_file(Path(__file__).parents[1] / 'app.py').run(timeout=10)
    app.file_uploader[0].upload(
        'linked.xlsx', source,
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    ).run(timeout=10)

    rendered = '\n'.join(markdown.value for markdown in app.markdown)
    captions = '\n'.join(caption.value for caption in app.caption)
    successes = '\n'.join(item.value for item in app.success)
    assert '1 populated repair row' in successes
    assert '2 populated individual-defect rows' in rendered
    assert '1 manual repair group' in rendered
    assert 'Recognized Individual Defects input columns (5)' in rendered
    assert 'stable Repair Group ID' in captions
    assert 'Individual Defects' in captions
    assert len(app.dataframe) == 2


def test_processed_v12_download_uses_v12_results_filename():
    """The v1.2 result must never look like an older CalcBatch download."""
    app = AppTest.from_file(Path(__file__).parents[1] / 'app.py').run(timeout=10)
    app.file_uploader[0].upload(
        'batch.xlsx', workbook_bytes_with_rows([valid_row_values()]),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    ).run(timeout=10)
    next(button for button in app.button if button.label == 'Calculate Batch').click().run(timeout=10)

    assert any(
        button.label == 'Download Processed Workbook'
        for button in app.download_button
    )
    assert app.session_state['processed_workbook_name'].startswith(
        'PROWRAP_CalcBatch_v1.2_Results_'
    )
