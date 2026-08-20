"""End-to-end release acceptance for the separate linked-corrosion batch app."""

from datetime import UTC, datetime
from io import BytesIO
import json
from pathlib import Path

from openpyxl import load_workbook
import pytest

from batch_schema import (
    DETAIL_INPUT_HEADERS,
    DETAIL_OUTPUT_HEADERS,
    INPUT_HEADERS,
    OUTPUT_HEADERS,
)
from tests.helpers import legacy_workbook_bytes_with_rows, valid_row_values
from workbook_processor import process_workbook


FIXED_TIME = datetime(2026, 8, 20, 12, 0, tzinfo=UTC)
EXPECTED_SHEETS = [
    'Batch Information', 'Batch Input & Results', 'Individual Defects',
    'Cost Calculation', 'Warnings', 'Summary', 'Instructions', 'Lists',
]
EXPECTED_COST_SOURCE_HEADERS = (
    'Pipe OD [mm]',
    'Nominal Wall [mm]',
    'Pipe Yield [MPa]',
    'Design Pressure [bar]',
    'Operating Temperature [degC]',
    'Mechanism',
    'Defect Location',
    'Defect Length [mm]',
    'Remaining Wall [mm]',
    'Design Life [years]',
    'Design Factor',
    'Prowrap CF Cloth Width [mm]',
    'Wall Loss [%]',
    'Required Structural Thickness [mm]',
    'Installed Plies',
    'Total Repair Length [mm]',
    'Cloth Band Count',
    'Procurement Axial Length [mm]',
    'Fabric Area [m2]',
    'Epoxy Mass [kg]',
)


def _columns(headers):
    return {header: index for index, header in enumerate(headers, start=1)}


def _formula_cells(workbook):
    return [
        (f'{worksheet.title}!{cell.coordinate}', cell.value)
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == 'f'
    ]


def _main_result_signature(workbook):
    """Return every emitted main engineering output for all six repair rows."""
    worksheet = workbook['Batch Input & Results']
    columns = _columns(INPUT_HEADERS + OUTPUT_HEADERS)
    return tuple(
        tuple(worksheet.cell(row, columns[heading]).value for heading in OUTPUT_HEADERS)
        for row in range(2, 8)
    )


def _detail_result_signature(workbook):
    """Return linked-detail ownership and complete calculated trace cells."""
    worksheet = workbook['Individual Defects']
    columns = _columns(DETAIL_INPUT_HEADERS + DETAIL_OUTPUT_HEADERS)
    headings = DETAIL_INPUT_HEADERS + DETAIL_OUTPUT_HEADERS
    return tuple(
        tuple(worksheet.cell(row, columns[heading]).value for heading in headings)
        for row in range(2, 5)
    )


def _warning_signature(workbook):
    worksheet = workbook['Warnings']
    return tuple(
        tuple(worksheet.cell(row, column).value for column in range(1, 4))
        for row in range(4, worksheet.max_row + 1)
        if worksheet.cell(row, 1).value
    )


def _summary_identity_signature(workbook):
    worksheet = workbook['Summary']
    return tuple(worksheet[address].value for address in ('B3', 'B4', 'B5', 'B24', 'B25'))


def test_release_documentation_uses_current_template_and_emitted_provenance():
    """Release instructions must not direct users to stale output identity."""
    root = Path(__file__).resolve().parents[1]
    readme = (root / 'README.md').read_text()
    engine_source = (root / 'ENGINE_SOURCE.md').read_text()

    assert readme.startswith('# PROWRAP CalcBatch v1.2\n')
    assert 'PROWRAP_CalcBatch_v1.2_Template.xlsx' in readme
    assert 'Processed workbooks record the short released revision `746f3b3`.' not in engine_source
    assert 'processor revision update is deferred' not in engine_source
    assert '`91b68d6`' in engine_source


def test_linked_corrosion_release_acceptance_workbook(tmp_path):
    """Exercise all v1.2 modes through the production template and processor."""
    from scripts.create_acceptance_workbook import create_acceptance_workbook

    source_path = tmp_path / 'acceptance-input.xlsx'
    create_acceptance_workbook(source_path)
    input_book = load_workbook(source_path, data_only=False)
    main_input = input_book['Batch Input & Results']
    detail_input = input_book['Individual Defects']
    main_columns = _columns(INPUT_HEADERS + OUTPUT_HEADERS)
    detail_columns = _columns(DETAIL_INPUT_HEADERS + DETAIL_OUTPUT_HEADERS)

    assert input_book.sheetnames == EXPECTED_SHEETS
    assert input_book.properties.title == 'PROWRAP CalcBatch v1.2'
    assert input_book['Batch Information']['A1'].value == 'PROWRAP CalcBatch v1.2'
    assert input_book['Instructions']['A1'].value == (
        'PROWRAP CalcBatch v1.2 — Instructions'
    )
    assert [input_book['Batch Information'].cell(row, 2).value for row in (3, 4, 5)] == [
        'Acceptance Customer', 'Acceptance Location', 'ACCEPT-001',
    ]
    expected_rows = (
        ('Actual defect length', None, 'Corrosion', 'External'),
        ('Independent defects', None, 'Corrosion', 'External'),
        ('Enter manually', 'R-001', 'Corrosion', 'External'),
        ('Enter manually', 'R-BAD', 'Corrosion', 'External'),
        (None, None, 'Dent no-crack', 'External'),
        (None, None, 'Dent w/crack', 'External'),
    )
    assert [
        tuple(main_input.cell(row, main_columns[header]).value for header in (
            'Defect Length Basis', 'Repair Group ID', 'Mechanism', 'Defect Location',
        )) for row in range(2, 8)
    ] == list(expected_rows)
    for row in range(2, 5):
        assert [main_input.cell(row, main_columns[header]).value for header in (
            'Pipe OD [mm]', 'Nominal Wall [mm]', 'Design Pressure [bar]',
            'Defect Length [mm]', 'Prowrap CF Cloth Width [mm]',
        )] == [1016.0, 12.0, 104.9, 1000.0, 500.0]
    assert main_input.cell(4, main_columns['Remaining Wall [mm]']).value is None
    assert [main_input.cell(5, main_columns[header]).value for header in (
        'Pipe OD [mm]', 'Nominal Wall [mm]', 'Design Pressure [bar]',
        'Defect Length [mm]', 'Defect Length Basis', 'Repair Group ID',
        'Remaining Wall [mm]', 'Prowrap CF Cloth Width [mm]',
    )] == [1016.0, 12.0, 104.9, 1000.0, 'Enter manually', 'R-BAD', None, 500.0]
    assert [
        tuple(detail_input.cell(row, detail_columns[header]).value for header in DETAIL_INPUT_HEADERS)
        for row in range(2, 5)
    ] == [
        ('R-001', 'D-01', 10.0, 9.652, 'Yes'),
        ('R-001', 'D-02', 35.0, 10.0, 'Yes'),
        ('R-BAD', 'D-BAD', 10.0, 9.652, 'No'),
    ]

    processed = process_workbook(source_path.read_bytes(), processed_at=FIXED_TIME)
    result_book = load_workbook(BytesIO(processed.workbook_bytes), data_only=False)
    main = result_book['Batch Input & Results']
    detail = result_book['Individual Defects']

    assert result_book.sheetnames == EXPECTED_SHEETS
    assert result_book['Lists'].sheet_state == 'hidden'
    assert main.protection.sheet is True
    assert detail.protection.sheet is True
    assert result_book['Summary'].protection.sheet is True
    assert all(result_book['Summary'][address].protection.locked for address in (
        'B3', 'B7', 'B24', 'B25',
    ))
    assert (main.tables['BatchRows'].ref, main.tables['BatchRows'].autoFilter.ref) == (
        'A1:BG501', 'A1:BG501',
    )
    assert (
        detail.tables['IndividualDefects'].ref,
        detail.tables['IndividualDefects'].autoFilter.ref,
    ) == ('A1:X2001', 'A1:X2001')
    assert (main.protection.autoFilter, main.protection.selectLockedCells,
            main.protection.selectUnlockedCells) == (False, False, False)
    assert (detail.protection.autoFilter, detail.protection.selectLockedCells,
            detail.protection.selectUnlockedCells) == (False, False, False)
    assert (main['A2'].protection.locked, main['T2'].protection.locked,
            main['U2'].protection.locked, main['BG2'].protection.locked) == (
        False, False, True, True,
    )
    assert (detail['A2'].protection.locked, detail['E2'].protection.locked,
            detail['F2'].protection.locked, detail['X2'].protection.locked) == (
        False, False, True, True,
    )
    assert _summary_identity_signature(result_book) == (
        'Acceptance Customer', 'Acceptance Location', 'ACCEPT-001', '1.2.0', '91b68d6',
    )
    assert [main.cell(row, main_columns['Calculation Status']).value for row in range(2, 8)] == [
        'REVIEW REQUIRED', 'REVIEW REQUIRED', 'REVIEW REQUIRED', 'INPUT ERROR', 'OK', 'OK',
    ]
    assert [main.cell(row, main_columns['Effective Pipe Capacity [bar]']).value / 10.0 for row in (2, 3, 4)] == pytest.approx([
        7.571542406120033, 8.82257484144555, 8.783461911867068,
    ])
    assert [main.cell(row, main_columns['Installed Plies']).value for row in (2, 3, 4)] == [12, 7, 7]
    assert [main.cell(row, main_columns['Repair Zone Length [mm]']).value for row in (2, 3, 4)] == [1000.0, 1000.0, 1000.0]
    assert main.cell(4, main_columns['Governing Defect ID']).value == 'D-02'
    assert main.cell(4, main_columns['Governing B31G Length [mm]']).value == 35.0
    assert main.cell(4, main_columns['Governing B31G Remaining Wall [mm]']).value == 10.0
    assert [main.cell(row, main_columns['B31G Candidate Count']).value for row in (2, 3, 4)] == [1, 1, 2]
    b31g_references = [
        json.loads(main.cell(row, main_columns['B31G Detail']).value)
        for row in (2, 3, 4)
    ]
    assert [reference['detail_schema_version'] for reference in b31g_references] == [
        '2', '2', '2',
    ]
    assert [reference['inline_candidate'] is not None for reference in b31g_references] == [
        True, True, False,
    ]
    assert b31g_references[0]['inline_candidate']['length_mm'] == 1000.0
    assert b31g_references[1]['inline_candidate']['length_mm'] == 10.0
    assert b31g_references[2]['detail_excel_row_range'] == '2:3'
    detail_signature = _detail_result_signature(result_book)
    assert detail_signature[0] == (
        'R-001', 'D-01', 10.0, 9.652, 'Yes', 2, 'OK', None, None,
        'modified', pytest.approx(0.19566666666666674),
        pytest.approx(0.008202099737532808), pytest.approx(1.0025699928354461),
        pytest.approx(519.0), pytest.approx(518.7347244738818),
        pytest.approx(122.53576168674373), pytest.approx(88.2257484144555),
        pytest.approx(1.3888888888888888), pytest.approx(444.07666666666677),
        True, False, pytest.approx(88.2257484144555), None, 'W013',
    )
    assert detail_signature[1] == (
        'R-001', 'D-02', 35.0, 10.0, 'Yes', 3, 'OK', None, None,
        'modified', pytest.approx(0.16666666666666666),
        pytest.approx(0.1004757217847769), pytest.approx(1.0310259179787589),
        pytest.approx(519.0), pytest.approx(516.4350290773692),
        pytest.approx(121.99252655370927), pytest.approx(87.83461911867067),
        pytest.approx(1.3888888888888888), pytest.approx(444.07666666666677),
        True, False, pytest.approx(87.83461911867067), 'Yes', 'W013',
    )
    assert detail_signature[2][:9] == (
        'R-BAD', 'D-BAD', 10.0, 9.652, 'No', 4, 'INPUT ERROR',
        'INVALID_SELECTION', 'Separation exceeds 3t: must be exactly Yes.',
    )
    assert detail_signature[2][9:] == (None,) * 15

    warning_rows = {
        result_book['Warnings'].cell(row, 1).value: result_book['Warnings'].cell(row, 3).value
        for row in range(4, result_book['Warnings'].max_row + 1)
        if result_book['Warnings'].cell(row, 1).value
    }
    assert warning_rows['W013'] == 'Main 2, 3, 4; Individual Defects 2, 3'

    cost = result_book['Cost Calculation']
    assert tuple(cost.cell(5, column).value for column in range(1, 21)) == (
        EXPECTED_COST_SOURCE_HEADERS
    )
    assert (cost['U5'].value, cost['V5'].value) == ('Cost', 'Price')
    assert (cost.tables['CostRows'].ref, cost.tables['CostRows'].autoFilter.ref) == ('A5:V11', 'A5:V11')
    assert cost.protection.sheet is True
    assert all(not cost[address].protection.locked for address in ('B3', 'E3', 'H3'))
    assert all(cost[address].protection.locked for address in ('A3', 'D3', 'G3'))
    expected_formulas = [
        (f'Cost Calculation!{column}{row}', formula)
        for row in range(6, 12)
        for column, formula in (
            ('U', f'=IF(OR($B$3="",$E$3="",S{row}="",T{row}=""),"",S{row}*$B$3+T{row}*$E$3)'),
            ('V', f'=IF(OR(U{row}="",$H$3=""),"",U{row}*$H$3)'),
        )
    ]
    assert _formula_cells(result_book) == expected_formulas
    source_columns = _columns(INPUT_HEADERS + OUTPUT_HEADERS)
    for cost_row, main_row in zip(range(6, 12), range(2, 8), strict=True):
        assert [cost.cell(cost_row, column).value for column in range(1, 21)] == [
            main.cell(main_row, source_columns[header]).value
            for header in EXPECTED_COST_SOURCE_HEADERS
        ]

    main_signature = _main_result_signature(result_book)
    warning_signature = _warning_signature(result_book)
    summary_identity = _summary_identity_signature(result_book)

    cost['B3'], cost['E3'], cost['H3'] = 25.0, 8.0, 1.4
    reupload = BytesIO()
    result_book.save(reupload)
    rebuilt_book = load_workbook(
        BytesIO(process_workbook(reupload.getvalue(), processed_at=FIXED_TIME).workbook_bytes),
        data_only=False,
    )
    assert rebuilt_book.sheetnames == EXPECTED_SHEETS
    assert [rebuilt_book['Cost Calculation'][address].value for address in ('B3', 'E3', 'H3')] == [25.0, 8.0, 1.4]
    assert _formula_cells(rebuilt_book) == expected_formulas
    rebuilt_main = rebuilt_book['Batch Input & Results']
    assert [rebuilt_main.cell(row, main_columns['Calculation Status']).value for row in range(2, 8)] == [
        'REVIEW REQUIRED', 'REVIEW REQUIRED', 'REVIEW REQUIRED', 'INPUT ERROR', 'OK', 'OK',
    ]
    assert _main_result_signature(rebuilt_book) == main_signature
    assert _detail_result_signature(rebuilt_book) == detail_signature
    assert _warning_signature(rebuilt_book) == warning_signature
    assert _summary_identity_signature(rebuilt_book) == summary_identity
    assert rebuilt_book['Summary'].protection.sheet is True
    assert all(rebuilt_book['Summary'][address].protection.locked for address in (
        'B3', 'B7', 'B24', 'B25',
    ))


@pytest.mark.parametrize('sheet_count', (5, 6, 7))
def test_legacy_controlled_layouts_upgrade_to_the_v12_eight_sheet_contract(sheet_count):
    """Old controlled five/six/seven-sheet downloads remain safe input files."""
    legacy = legacy_workbook_bytes_with_rows([valid_row_values()], sheet_count=sheet_count)
    upgraded = process_workbook(legacy, processed_at=FIXED_TIME)
    workbook = load_workbook(BytesIO(upgraded.workbook_bytes), data_only=False)
    main_columns = _columns(INPUT_HEADERS + OUTPUT_HEADERS)

    assert workbook.sheetnames == EXPECTED_SHEETS
    assert workbook['Batch Input & Results'].cell(2, main_columns['Defect Length Basis']).value == 'Actual defect length'
    assert workbook['Batch Input & Results'].cell(2, main_columns['Calculation Status']).value == 'OK'
