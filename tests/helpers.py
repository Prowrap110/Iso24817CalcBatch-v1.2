def valid_engine_inputs(**overrides):
    values = dict(
        customer='Batch Customer', location='Batch Location', report_no='B-001',
        od=457.2, wall=9.53, pressure=50.0, temp=40.0,
        defect_type='Corrosion', defect_loc='External', length=100.0,
        rem_wall=4.5, yield_strength=359.0, design_factor=0.72,
        design_life=20, internal_corrosion_rate=0.0,
        installation_temp=20.0, component_type='Straight',
        cyclic_derating_factor=1.0, axial_load_case=0,
        cloth_width_mm=300.0,
    )
    values.update(overrides)
    return values


def valid_row_values(**overrides):
    values = {
        'Pipe OD [mm]': 457.2,
        'Nominal Wall [mm]': 9.53,
        'Pipe Yield [MPa]': 359.0,
        'Design Pressure [bar]': 50.0,
        'Operating Temperature [degC]': 40.0,
        'Mechanism': 'Corrosion',
        'Defect Location': 'External',
        'Defect Length [mm]': 100.0,
        'Defect Length Basis': 'Actual defect length',
        'Repair Group ID': None,
        'Remaining Wall [mm]': 4.5,
        'Internal Corrosion Rate [mm/year]': None,
        'Design Life [years]': 20,
        'Design Factor': 0.72,
        'Run Type A / Class 3 Check': 'No',
        'Installation Temperature [degC]': 20.0,
        'Component Type': 'Straight',
        'Cyclic Derating Factor': 1.0,
        'Axial Load Case': 0,
        'Prowrap CF Cloth Width [mm]': 300.0,
    }
    values.update(overrides)
    if (
        values['Mechanism'] != 'Corrosion'
        or values['Defect Location'] != 'External'
    ) and 'Defect Length Basis' not in overrides:
        values['Defect Length Basis'] = None
    return values


def batch_info():
    from batch_schema import BatchInfo

    return BatchInfo('Batch Customer', 'Batch Location', 'B-001')


def validated_row(**overrides):
    from batch_validation import validate_row

    row, issues = validate_row(2, valid_row_values(**overrides))
    assert not issues
    assert row is not None
    return row


def valid_detail_row(
    excel_row, *, group='R-001', defect='D-01', length=10.0, wall=4.5,
    separation='Yes',
):
    from batch_validation import validate_individual_defect_row

    row, issues = validate_individual_defect_row(excel_row, {
        'Repair Group ID': group,
        'Defect ID': defect,
        'Individual longitudinal length [mm]': length,
        'Remaining wall [mm]': wall,
        'Separation exceeds 3t': separation,
    })
    assert issues == ()
    assert row is not None
    return row


def detail_values(
    *, group='R-001', defect='D-01', length=10.0, wall=4.5,
    separation='Yes',
):
    return {
        'Repair Group ID': group,
        'Defect ID': defect,
        'Individual longitudinal length [mm]': length,
        'Remaining wall [mm]': wall,
        'Separation exceeds 3t': separation,
    }


def workbook_bytes_with_rows(rows, *, detail_rows=(), commercial_inputs=None):
    """Create a controlled template populated with test defect rows."""
    from io import BytesIO

    from openpyxl import load_workbook

    from batch_schema import DETAIL_INPUT_HEADERS, INPUT_HEADERS
    from workbook_template import create_template_workbook

    workbook = load_workbook(BytesIO(create_template_workbook()))
    info = workbook['Batch Information']
    info['B3'] = 'Batch Customer'
    info['B4'] = 'Batch Location'
    info['B5'] = 'B-001'
    data = workbook['Batch Input & Results']
    for excel_row, values in enumerate(rows, start=2):
        for column, header in enumerate(INPUT_HEADERS, start=1):
            data.cell(excel_row, column, values.get(header))
    detail = workbook['Individual Defects']
    for excel_row, values in enumerate(detail_rows, start=2):
        for column, header in enumerate(DETAIL_INPUT_HEADERS, start=1):
            detail.cell(excel_row, column, values.get(header))
    if commercial_inputs:
        cost = workbook['Cost Calculation']
        for address, value in commercial_inputs.items():
            cost[address] = value
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def legacy_workbook_bytes_with_rows(
    rows, *, sheet_count=7, former_cost_contract=False,
):
    """Create an exact controlled pre-v1.2 workbook in a supported layout."""
    from io import BytesIO

    from openpyxl import load_workbook

    from batch_schema import LEGACY_INPUT_HEADERS, LEGACY_OUTPUT_HEADERS

    workbook = load_workbook(BytesIO(workbook_bytes_with_rows(rows)))
    del workbook['Individual Defects']
    if sheet_count < 7:
        del workbook['Cost Calculation']
    if sheet_count < 6:
        del workbook['Warnings']
    main = workbook['Batch Input & Results']
    main.delete_cols(9, 2)
    for column, header in enumerate(
        LEGACY_INPUT_HEADERS + LEGACY_OUTPUT_HEADERS, start=1,
    ):
        main.cell(1, column).value = header
    if former_cost_contract:
        _freeze_former_cost_contract(workbook)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def historical_v12_workbook_bytes_with_rows(rows, *, former_cost_contract=False):
    """Create the former wide eight-sheet v1.2 contract for upgrade tests."""
    from io import BytesIO

    from openpyxl import load_workbook

    from batch_schema import HISTORICAL_V12_OUTPUT_HEADERS, INPUT_HEADERS

    workbook = load_workbook(BytesIO(workbook_bytes_with_rows(rows)))
    main = workbook['Batch Input & Results']
    for column, header in enumerate(
        INPUT_HEADERS + HISTORICAL_V12_OUTPUT_HEADERS, start=1,
    ):
        main.cell(1, column).value = header
    if former_cost_contract:
        _freeze_former_cost_contract(workbook)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _freeze_former_cost_contract(workbook) -> None:
    """Make a faithful processed Cost sheet from before Quantity existed."""
    from cost_calculation import cost_formula, price_formula

    cost = workbook['Cost Calculation']
    cost.unmerge_cells('A1:X1')
    cost.merge_cells('A1:V1')
    cost.delete_cols(23, 2)
    table = cost.tables['CostRows']
    table.ref = 'A5:V6'
    table.autoFilter.ref = table.ref
    cost['U6'] = cost_formula(6)
    cost['V6'] = price_formula(6)
