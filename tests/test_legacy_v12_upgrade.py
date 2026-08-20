from datetime import UTC, datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

import batch_schema
from batch_schema import (
    INPUT_HEADERS,
    LEGACY_INPUT_HEADERS,
    LEGACY_OUTPUT_HEADERS,
    OUTPUT_HEADERS,
)
from cost_calculation import (
    COST_TABLE_HEADERS,
    cost_formula,
    price_formula,
    total_amount_formula,
)
from engine.corrosion_defects import ACTUAL_DEFECT_LENGTH
from tests.helpers import (
    _freeze_former_cost_contract,
    historical_v12_workbook_bytes_with_rows,
    legacy_workbook_bytes_with_rows,
    valid_row_values,
    workbook_bytes_with_rows,
)
from workbook_processor import inspect_workbook, process_workbook


FIXED_TIME = datetime(2026, 8, 14, 12, 0, tzinfo=UTC)
CURRENT_SHEETS = [
    'Batch Information',
    'Batch Input & Results',
    'Individual Defects',
    'Cost Calculation',
    'Warnings',
    'Summary',
    'Instructions',
    'Lists',
]


def _saved(workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _workbook_from_bytes(data: bytes):
    return load_workbook(BytesIO(data), data_only=False)


def _legacy_source(sheet_count: int) -> bytes:
    """Freeze the former v1.1 header contract in each supported sheet layout."""
    workbook = load_workbook(BytesIO(legacy_workbook_bytes_with_rows(
        [valid_row_values()], sheet_count=sheet_count,
    )))
    main = workbook['Batch Input & Results']
    assert tuple(cell.value for cell in main[1]) == (
        LEGACY_INPUT_HEADERS + LEGACY_OUTPUT_HEADERS
    )
    installed_plies_column = (
        len(LEGACY_INPUT_HEADERS) + LEGACY_OUTPUT_HEADERS.index('Installed Plies') + 1
    )
    main.cell(2, installed_plies_column).value = 999
    return _saved(workbook)


@pytest.mark.parametrize('sheet_count', [5, 6, 7])
def test_supported_legacy_workbooks_upgrade_to_current_contract(sheet_count):
    """Catches stranding any controlled pre-v1.2 workbook generation."""
    source = _legacy_source(sheet_count)

    inspection = inspect_workbook(source)
    processed = process_workbook(source, FIXED_TIME, f'legacy-{sheet_count}.xlsx')
    workbook = load_workbook(BytesIO(processed.workbook_bytes), data_only=False)
    main = workbook['Batch Input & Results']
    headings = tuple(cell.value for cell in main[1])

    assert inspection.workbook_errors == ()
    assert workbook.sheetnames == CURRENT_SHEETS
    assert headings == INPUT_HEADERS + OUTPUT_HEADERS
    assert main.cell(2, headings.index('Defect Length Basis') + 1).value == (
        ACTUAL_DEFECT_LENGTH
    )
    assert main.cell(2, headings.index('Repair Group ID') + 1).value is None
    assert main.cell(2, headings.index('Installed Plies') + 1).value == 3
    detail = workbook['Individual Defects']
    assert all(cell.value is None for cell in detail[2])


def test_legacy_upgrade_sets_actual_basis_only_for_eligible_external_corrosion():
    """Catches adding a corrosion-only choice to unrelated legacy repair rows."""
    workbook = load_workbook(BytesIO(_legacy_source(7)))
    main = workbook['Batch Input & Results']
    headings = tuple(cell.value for cell in main[1])
    main.cell(2, headings.index('Mechanism') + 1).value = 'Leak'

    processed = process_workbook(_saved(workbook), FIXED_TIME, 'legacy-leak.xlsx')
    rebuilt = load_workbook(BytesIO(processed.workbook_bytes), data_only=False)
    rebuilt_main = rebuilt['Batch Input & Results']
    rebuilt_headings = tuple(cell.value for cell in rebuilt_main[1])

    assert rebuilt_main.cell(
        2, rebuilt_headings.index('Defect Length Basis') + 1,
    ).value is None


def test_historical_wide_v12_workbook_upgrades_without_legacy_normalization():
    """Catches an old eight-sheet v1.2 upload being misclassified as pre-v1.2."""
    source = historical_v12_workbook_bytes_with_rows([
        valid_row_values(**{'Defect Length Basis': 'Independent defects'}),
    ])

    inspection = inspect_workbook(source)
    processed = process_workbook(source, FIXED_TIME, 'historical-v12.xlsx')
    workbook = _workbook_from_bytes(processed.workbook_bytes)
    main = workbook['Batch Input & Results']

    assert inspection.workbook_errors == ()
    assert tuple(cell.value for cell in main[1]) == INPUT_HEADERS + OUTPUT_HEADERS
    assert tuple(cell.value for cell in _workbook_from_bytes(source)[
        'Batch Input & Results'
    ][1]) == INPUT_HEADERS + batch_schema.HISTORICAL_V12_OUTPUT_HEADERS
    assert main.cell(2, INPUT_HEADERS.index('Defect Length Basis') + 1).value == (
        'Independent defects'
    )


@pytest.mark.parametrize('source_factory', [
    lambda: historical_v12_workbook_bytes_with_rows(
        [valid_row_values(**{'Defect Length Basis': 'Independent defects'})],
        former_cost_contract=True,
    ),
    lambda: legacy_workbook_bytes_with_rows(
        [valid_row_values()], sheet_count=7, former_cost_contract=True,
    ),
])
def test_former_cost_contract_upgrades_only_recognized_historical_sources(source_factory):
    """Catches rejecting a faithful pre-Quantity controlled Cost sheet."""
    source = source_factory()
    source_cost = _workbook_from_bytes(source)['Cost Calculation']

    assert tuple(cell.value for cell in source_cost[5]) == (
        COST_TABLE_HEADERS[:-2]
    )
    assert source_cost.tables['CostRows'].ref == 'A5:V6'
    assert (source_cost['U6'].value, source_cost['V6'].value) == (
        cost_formula(6), price_formula(6),
    )

    inspection = inspect_workbook(source)
    assert inspection.workbook_errors == ()
    rebuilt = _workbook_from_bytes(process_workbook(source, FIXED_TIME).workbook_bytes)
    cost = rebuilt['Cost Calculation']

    assert tuple(cell.value for cell in cost[5]) == COST_TABLE_HEADERS
    assert cost['W6'].value is None
    assert (cost['U6'].value, cost['V6'].value, cost['X6'].value) == (
        cost_formula(6), price_formula(6), total_amount_formula(6),
    )


def test_current_compact_source_rejects_former_cost_contract():
    """Catches broad acceptance of A:V Cost headers for current compact files."""
    source = workbook_bytes_with_rows([valid_row_values()])
    workbook = _workbook_from_bytes(source)
    _freeze_former_cost_contract(workbook)

    inspection = inspect_workbook(_saved(workbook))

    assert [issue.code for issue in inspection.workbook_errors] == [
        'INVALID_COST_HEADERS',
    ]
