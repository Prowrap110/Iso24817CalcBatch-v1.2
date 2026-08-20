from io import BytesIO

from openpyxl import load_workbook

from batch_schema import (
    DETAIL_INPUT_HEADERS,
    DETAIL_OUTPUT_HEADERS,
    INPUT_HEADERS,
    MAX_DETAIL_ROWS,
    MAX_ROWS,
    OUTPUT_HEADERS,
)


def _template_workbook():
    from workbook_template import create_template_workbook

    return load_workbook(BytesIO(create_template_workbook()))


def test_template_has_common_info_and_row_table():
    """Catches a template that places batch fields in the defect table."""
    workbook = _template_workbook()

    assert workbook.sheetnames == [
        'Batch Information', 'Batch Input & Results', 'Individual Defects',
        'Cost Calculation', 'Warnings', 'Summary', 'Instructions', 'Lists',
    ]
    info = workbook['Batch Information']
    assert [info['A3'].value, info['A4'].value, info['A5'].value] == [
        'Customer', 'Project Location', 'Report No',
    ]
    assert [info['B3'].value, info['B4'].value, info['B5'].value] == [None, None, None]

    data = workbook['Batch Input & Results']
    assert data['A1'].value == 'Pipe OD [mm]'
    assert data.freeze_panes == 'B2'
    assert workbook['Lists'].sheet_state == 'hidden'


def test_template_uses_exact_calc_batch_v12_product_identity():
    workbook = _template_workbook()

    assert workbook.properties.title == 'PROWRAP CalcBatch v1.2'
    assert workbook['Batch Information']['A1'].value == 'PROWRAP CalcBatch v1.2'
    assert workbook['Instructions']['A1'].value == (
        'PROWRAP CalcBatch v1.2 — Instructions'
    )


def test_template_has_blank_editable_cost_sheet_in_new_controlled_order():
    workbook = _template_workbook()

    assert workbook.sheetnames == [
        'Batch Information', 'Batch Input & Results', 'Individual Defects',
        'Cost Calculation', 'Warnings', 'Summary', 'Instructions', 'Lists',
    ]
    cost = workbook['Cost Calculation']
    assert [cost[address].value for address in ('B3', 'E3', 'H3')] == [None, None, None]
    assert all(
        cost[address].protection.locked is False
        for address in ('B3', 'E3', 'H3')
    )
    assert [cost.cell(5, column).value for column in range(1, 25)][-4:] == [
        'Cost', 'Price', 'Quantity', 'Total Amount',
    ]
    assert cost['W6'].protection.locked is False
    assert cost['X6'].protection.locked is True
    assert cost['W6'].fill.fgColor.rgb == '00FFF2CC'
    assert any(
        validation.type == 'decimal' and str(validation.sqref) == 'W6:W155'
        for validation in cost.data_validations.dataValidation
    )
    assert cost.freeze_panes == 'A6'
    assert cost.protection.sheet is True
    assert cost.protection.selectUnlockedCells is False


def test_template_visibly_highlights_cost_assumption_value_cells():
    """Catches editable commercial inputs blending into the white worksheet."""
    cost = _template_workbook()['Cost Calculation']
    inputs = [cost[address] for address in ('B3', 'E3', 'H3')]

    assert [cell.fill.fill_type for cell in inputs] == ['solid'] * 3
    assert [cell.fill.fgColor.rgb for cell in inputs] == ['00FFF2CC'] * 3
    assert [cell.number_format for cell in inputs] == ['#,##0.00'] * 3
    assert [cell.protection.locked for cell in inputs] == [False] * 3
    assert {str(validation.sqref) for validation in cost.data_validations.dataValidation} == {
        'B3 E3 H3', 'W6:W155',
    }


def test_cost_quantity_columns_use_integer_display_without_populating_template_rows():
    """Catch plies or cloth-band counts being displayed as fractional quantities."""
    cost = _template_workbook()['Cost Calculation']
    headers = [cost.cell(5, column).value for column in range(1, 23)]
    count_columns = [
        headers.index(header) + 1
        for header in ('Installed Plies', 'Cloth Band Count')
    ]

    assert [cost.cell(5, column).value for column in count_columns] == [
        'Installed Plies', 'Cloth Band Count',
    ]
    for row in (6, MAX_ROWS + 5):
        assert [cost.cell(row, column).value for column in count_columns] == [
            None, None,
        ]
        assert [cost.cell(row, column).number_format for column in count_columns] == [
            '#,##0', '#,##0',
        ]
        assert cost.cell(row, 11).number_format == '#,##0.00'


def test_template_has_a_visible_protected_warning_register_with_empty_state():
    """Catches a template that leaves long warning explanations in result rows."""
    workbook = _template_workbook()
    warnings = workbook['Warnings']

    assert warnings.sheet_state == 'visible'
    assert warnings['A1'].value == 'Compliance Warning Register'
    assert [warnings.cell(3, column).value for column in range(1, 4)] == [
        'Warning Code',
        'Warning Meaning / Required Action',
        'Affected Excel Rows',
    ]
    assert warnings['A4'].value == 'No compliance warnings were generated.'
    assert warnings.freeze_panes == 'A4'
    assert warnings.protection.sheet is True
    assert not warnings.tables


def test_warning_register_subtitle_spans_the_table_without_clipping():
    """Catches a wrapped subtitle constrained to the narrow code column."""
    warnings = _template_workbook()['Warnings']

    assert 'A2:C2' in {str(cell_range) for cell_range in warnings.merged_cells.ranges}
    assert warnings.row_dimensions[2].height >= 32


def test_compact_output_columns_do_not_include_diagnostic_json():
    """Catches removed audit diagnostics returning to the public main table."""
    data = _template_workbook()['Batch Input & Results']

    assert tuple(cell.value for cell in data[1]) == INPUT_HEADERS + OUTPUT_HEADERS
    assert not {
        'Calculation Status', 'Error Code', 'Error Message', 'Compliance Warnings',
        'B31G Detail', 'Type A Detail', 'Type B Detail',
    }.intersection(OUTPUT_HEADERS)


def test_template_uses_canonical_headings_and_a_filterable_compact_table():
    """Catches a workbook whose headers or accepted row extent diverge from parsing."""
    workbook = _template_workbook()
    data = workbook['Batch Input & Results']
    headings = [cell.value for cell in data[1]]

    assert headings == list(INPUT_HEADERS + OUTPUT_HEADERS)
    assert len(data.tables) == 1
    table = next(iter(data.tables.values()))
    assert table.ref == 'A1:AC151'
    assert table.autoFilter.ref == table.ref


def test_v12_template_has_linked_detail_sheet_in_controlled_order():
    """Catches a template that omits the protected linked-defect input table."""
    workbook = _template_workbook()

    assert workbook.sheetnames == [
        'Batch Information', 'Batch Input & Results', 'Individual Defects',
        'Cost Calculation', 'Warnings', 'Summary', 'Instructions', 'Lists',
    ]
    detail = workbook['Individual Defects']
    assert tuple(cell.value for cell in detail[1]) == DETAIL_INPUT_HEADERS + DETAIL_OUTPUT_HEADERS
    assert detail.freeze_panes == 'B2'
    assert detail.tables['IndividualDefects'].ref == 'A1:X151'
    assert detail.protection.sheet is True
    assert detail.protection.autoFilter is False
    assert detail['A2'].protection.locked is False
    assert detail.cell(2, len(DETAIL_INPUT_HEADERS) + 1).protection.locked is True


def test_v12_template_has_exact_basis_and_yes_dropdowns():
    """Catches linked-corrosion selections that bypass the controlled lists."""
    workbook = _template_workbook()
    main = workbook['Batch Input & Results']
    detail = workbook['Individual Defects']

    validations = {
        item.formula1: str(item.sqref)
        for item in main.data_validations.dataValidation
    }
    assert validations['=DefectLengthBasisChoices'] == 'I2:I151'
    detail_validations = {
        item.formula1: str(item.sqref)
        for item in detail.data_validations.dataValidation
    }
    assert detail_validations['=SeparationChoices'] == 'E2:E151'


def test_template_adds_dropdowns_for_every_selection_through_row_151():
    """Catches a controlled selection that can be entered unchecked in later rows."""
    workbook = _template_workbook()
    data = workbook['Batch Input & Results']
    validations = {
        validation.formula1: str(validation.sqref)
        for validation in data.data_validations.dataValidation
    }

    expected_choices = {
        f'={name}': (
            f'{data.cell(1, INPUT_HEADERS.index(header) + 1).column_letter}2:'
            f'{data.cell(1, INPUT_HEADERS.index(header) + 1).column_letter}{MAX_ROWS + 1}'
        )
        for name, header in {
            'MechanismChoices': 'Mechanism',
            'DefectLocationChoices': 'Defect Location',
            'TypeACheckChoices': 'Run Type A / Class 3 Check',
            'ComponentTypeChoices': 'Component Type',
            'AxialLoadCaseChoices': 'Axial Load Case',
            'DefectLengthBasisChoices': 'Defect Length Basis',
        }.items()
    }
    assert validations == expected_choices


def test_template_mechanism_choices_and_guidance_distinguish_dent_routes():
    """Catch a new template that restores ambiguous Dent or hides route limits."""
    workbook = _template_workbook()
    data = workbook['Batch Input & Results']
    lists = workbook['Lists']

    mechanism_values = [lists.cell(row, 1).value for row in range(1, 6)]
    assert mechanism_values == [
        'Corrosion', 'Dent w/crack', 'Dent no-crack', 'Leak', 'Crack',
    ]
    assert 'Dent' not in mechanism_values
    assert data['F1'].comment is not None
    mechanism_note = data['F1'].comment.text.lower()
    assert 'dent w/crack' in mechanism_note
    assert 'full-pressure laminate' in mechanism_note
    assert 'dent no-crack' in mechanism_note
    assert 'component-pipe substrate load sharing' in mechanism_note

    instruction_text = ' '.join(
        str(cell.value).lower()
        for row in workbook['Instructions'].iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert 'dent w/crack' in instruction_text
    assert 'full-pressure laminate' in instruction_text
    assert 'dent no-crack' in instruction_text
    assert 'component-pipe substrate load sharing' in instruction_text
    assert 'legacy dent' in instruction_text
    assert 'older batch workbook' in instruction_text
    assert 'becomes dent w/crack' in instruction_text
    assert 'not a complete dent integrity or fatigue acceptance assessment' in instruction_text


def test_long_dent_instruction_has_room_for_every_wrapped_line():
    """Catches the final legacy-migration sentence being clipped in row 15."""
    instructions = _template_workbook()['Instructions']

    assert instructions['A15'].alignment.wrap_text is True
    assert instructions.row_dimensions[15].height >= 48


def test_template_dropdowns_reject_invalid_selections_but_allow_unused_blank_rows():
    """Catches validations that block a blank unused row or silently accept bad selections."""
    workbook = _template_workbook()
    data = workbook['Batch Input & Results']
    validations = data.data_validations.dataValidation

    assert all(validation.showErrorMessage is True for validation in validations)
    assert all(validation.errorStyle == 'stop' for validation in validations)
    assert all(validation.allow_blank is True for validation in validations)


def test_template_marks_inputs_editable_and_outputs_protected_with_clear_headers():
    """Catches output cells that are editable or visually indistinguishable from inputs."""
    workbook = _template_workbook()
    data = workbook['Batch Input & Results']
    info = workbook['Batch Information']
    input_count = len(INPUT_HEADERS)

    assert data['A1'].fill.fgColor.rgb != data.cell(1, input_count + 1).fill.fgColor.rgb
    assert info['B3'].protection.locked is False
    assert data['A2'].protection.locked is False
    assert data.cell(2, input_count + 1).protection.locked is True
    assert data.protection.sheet is True
    defect_length_basis_column = INPUT_HEADERS.index('Defect Length Basis') + 1
    corrosion_rate_column = INPUT_HEADERS.index('Internal Corrosion Rate [mm/year]') + 1
    assert data.cell(1, defect_length_basis_column).comment is not None
    assert 'required only for internal corrosion' in data.cell(
        1, corrosion_rate_column,
    ).comment.text.lower()


def test_template_protects_summary_outputs_and_keeps_selection_consistent():
    """Catches a template whose summary/provenance values can be overwritten."""
    summary = _template_workbook()['Summary']

    assert summary.protection.sheet is True
    assert summary.protection.selectLockedCells is False
    assert summary.protection.selectUnlockedCells is False
    assert all(summary[address].protection.locked for address in (
        'A3', 'B3', 'A7', 'B7', 'A24', 'B24', 'A25', 'B25',
    ))


def test_protected_main_sheet_keeps_inputs_selectable_and_table_filterable():
    """Catches inverted openpyxl protection flags after workbook serialization."""
    workbook = _template_workbook()
    data = workbook['Batch Input & Results']
    input_count = len(INPUT_HEADERS)
    table = data.tables['BatchRows']

    assert data.protection.sheet is True
    assert data.protection.selectUnlockedCells is False
    assert data.protection.autoFilter is False
    assert data['A2'].protection.locked is False
    assert data.cell(2, input_count + 1).protection.locked is True
    assert table.autoFilter.ref == table.ref


def test_template_keeps_status_colors_only_for_individual_defect_audits():
    """Catches status formatting returning after its main result column was removed."""
    workbook = _template_workbook()
    data = workbook['Batch Input & Results']
    detail = workbook['Individual Defects']

    main_rules = [
        rule
        for rules in data.conditional_formatting._cf_rules.values()
        for rule in rules
    ]
    detail_rules = [
        rule
        for rules in detail.conditional_formatting._cf_rules.values()
        for rule in rules
    ]
    status_rules = [rule for rule in detail_rules if rule.type == 'containsText']

    assert not [rule for rule in main_rules if rule.type == 'containsText']
    assert {rule.text for rule in status_rules} == {
        'OK', 'REVIEW REQUIRED', 'NOT REPAIRABLE', 'INPUT ERROR', 'SYSTEM ERROR',
    }
    assert all(rule.dxf.fill.fgColor.rgb for rule in status_rules)


def test_template_contains_no_formulas_and_has_user_guidance():
    """Catches executable workbook logic or a template missing the review disclaimer."""
    workbook = _template_workbook()

    formulas = [
        cell.coordinate
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith('=')
    ]
    assert formulas == []
    instructions = workbook['Instructions']
    summary = workbook['Summary']
    assert 'blank rows are ignored' in ' '.join(
        str(cell.value).lower()
        for row in instructions.iter_rows()
        for cell in row
        if cell.value is not None
    )
    instruction_text = ' '.join(
        str(cell.value).lower()
        for row in instructions.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert 'warnings worksheet' in instruction_text
    assert 'up to 150 populated main rows and 150 individual defects rows' in instruction_text
    assert 'wall loss [%], required structural thickness [mm], installed plies, total repair length [mm], cloth band count, procurement axial length [mm], fabric area [m2], epoxy mass [kg], and repair zone length [mm]' in instruction_text
    assert 'quantity is editable' in instruction_text
    assert 'total amount = price x quantity' in instruction_text
    assert '300 mm and 500 mm' in instruction_text
    assert 'tg = 110' in instruction_text
    assert 'b3 (cf cost / m2), e3 (epoxy cost / kg), and h3 (price multiplier)' in instruction_text
    assert 'may be blank or retain values from a previously processed workbook' in instruction_text
    assert 'actual defect length = continuous or interacting b31g length.' in instruction_text
    assert 'independent defects = 10 x 10 mm, each separated by more than 3t.' in instruction_text
    assert 't means nominal pipe wall thickness.' in instruction_text
    assert 'enter manually = leave main remaining wall blank and link detail rows with repair group id.' in instruction_text
    assert 'defect length remains the complete outer-to-outer continuous repair-zone span.' in instruction_text
    assert 'one main row per continuous repair' in instruction_text
    assert 'one independent defect per row' not in instruction_text
    assert 'five-sheet, six-sheet, and seven-sheet' in instruction_text
    assert 'preliminary screening' in ' '.join(
        str(cell.value).lower()
        for row in summary.iter_rows()
        for cell in row
        if cell.value is not None
    )
