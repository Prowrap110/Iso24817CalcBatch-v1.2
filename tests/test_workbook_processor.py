from datetime import UTC, datetime
from io import BytesIO
import json
import zipfile

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula

from batch_adapter import RowCalculation
from batch_schema import (
    DETAIL_INPUT_HEADERS,
    DETAIL_OUTPUT_HEADERS,
    INPUT_HEADERS,
    MAX_DETAIL_ROWS,
    MAX_ROWS,
    MAX_UPLOAD_BYTES,
    OUTPUT_HEADERS,
)
from batch_status import CalculationStatus
from cost_calculation import (
    COST_SOURCE_HEADERS,
    COST_TABLE_HEADERS,
    cost_formula,
    price_formula,
)
from engine.corrosion_defects import ENTER_MANUALLY
from tests.helpers import (
    detail_values,
    legacy_workbook_bytes_with_rows,
    valid_row_values,
    workbook_bytes_with_rows,
)
from workbook_processor import (
    WorkbookProcessingError,
    _commercial_input_errors,
    _output_value,
    inspect_workbook,
    process_workbook,
)


FIXED_TIME = datetime(2026, 8, 14, 12, 0, tzinfo=UTC)


def _workbook(data: bytes):
    return load_workbook(BytesIO(data), data_only=False)


def _saved(workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_inspects_a_valid_template_and_previews_populated_rows():
    inspection = inspect_workbook(workbook_bytes_with_rows([valid_row_values()]))

    assert inspection.workbook_errors == ()
    assert inspection.batch_info is not None
    assert inspection.batch_info.customer == 'Batch Customer'
    assert inspection.populated_rows == 1
    assert inspection.valid_rows == 1
    assert inspection.invalid_rows == 0
    assert inspection.recognized_input_headers == INPUT_HEADERS
    assert inspection.missing_input_headers == ()
    assert inspection.unexpected_headers == ()
    assert inspection.preview == ({
        'Source Excel Row': 2,
        'Pipe OD [mm]': 457.2,
        'Mechanism': 'Corrosion',
        'Defect Location': 'External',
        'Defect Length Basis': 'Actual defect length',
        'Repair Group ID': None,
        'Calculation Status': 'OK',
        'Error Code': '',
        'Error Message': '',
    },)


def test_inspection_reports_main_detail_and_manual_group_counts():
    """Catches a v1.2 inspection that ignores the linked detail table."""
    source = workbook_bytes_with_rows(
        [valid_row_values(**{
            'Defect Length Basis': ENTER_MANUALLY,
            'Repair Group ID': 'R-001',
            'Remaining Wall [mm]': None,
        })],
        detail_rows=[
            detail_values(group='R-001', defect='D-01', length=10, wall=9.0),
            detail_values(group='R-001', defect='D-02', length=35, wall=9.1),
        ],
    )

    inspection = inspect_workbook(source)

    assert inspection.workbook_errors == ()
    assert inspection.populated_rows == 1
    assert inspection.populated_detail_rows == 2
    assert inspection.manual_groups == 1
    assert inspection.recognized_detail_input_headers == DETAIL_INPUT_HEADERS
    assert inspection.missing_detail_input_headers == ()
    assert inspection.unexpected_detail_headers == ()
    assert inspection.preview[0]['Defect Length Basis'] == ENTER_MANUALLY
    assert inspection.preview[0]['Repair Group ID'] == 'R-001'


def _manual_row(group='R-001'):
    return valid_row_values(**{
        'Pipe OD [mm]': 1016.0,
        'Nominal Wall [mm]': 12.0,
        'Pipe Yield [MPa]': 450.0,
        'Design Pressure [bar]': 104.9,
        'Defect Length [mm]': 1000.0,
        'Defect Length Basis': ENTER_MANUALLY,
        'Repair Group ID': group,
        'Remaining Wall [mm]': None,
        'Prowrap CF Cloth Width [mm]': 500.0,
        'Run Type A / Class 3 Check': 'Yes',
    })


def _result_signature(data: bytes):
    workbook = _workbook(data)
    main = workbook['Batch Input & Results']
    detail = workbook['Individual Defects']
    return (
        tuple(main.cell(2, column).value for column in range(
            len(INPUT_HEADERS) + 1,
            len(INPUT_HEADERS) + len(OUTPUT_HEADERS) + 1,
        )),
        tuple(
            tuple(detail.cell(row, column).value for column in range(
                len(DETAIL_INPUT_HEADERS) + 1,
                len(DETAIL_INPUT_HEADERS) + len(DETAIL_OUTPUT_HEADERS) + 1,
            ))
            for row in (2, 3)
        ),
    )


def test_manual_rows_calculate_with_ordered_detail_results_and_one_governing_row():
    """Catches pair reordering or loss of candidate results at the workbook boundary."""
    source = workbook_bytes_with_rows(
        [_manual_row()],
        detail_rows=[
            detail_values(group='R-001', defect='D-01', length=10, wall=9.652),
            detail_values(group='R-001', defect='D-02', length=35, wall=10.0),
        ],
    )

    processed = process_workbook(source, FIXED_TIME, 'manual.xlsx')
    workbook = _workbook(processed.workbook_bytes)
    main = workbook['Batch Input & Results']
    detail = workbook['Individual Defects']
    main_headings = tuple(cell.value for cell in main[1])
    detail_headings = tuple(cell.value for cell in detail[1])

    assert processed.status_counts == {'REVIEW REQUIRED': 1}
    assert main.cell(2, main_headings.index('Governing Defect ID') + 1).value == 'D-02'
    assert [detail.cell(row, detail_headings.index('Calculation Status') + 1).value for row in (2, 3)] == [
        'OK', 'OK',
    ]
    assert [detail.cell(row, detail_headings.index('Source Excel Row') + 1).value for row in (2, 3)] == [
        2, 3,
    ]
    assert [detail.cell(row, detail_headings.index('Governing Defect') + 1).value for row in (2, 3)] == [
        None, 'Yes',
    ]
    assert [detail.cell(row, detail_headings.index('Credited Safe Pressure [bar]') + 1).value for row in (2, 3)] == pytest.approx([
        88.2257484144555, 87.83461911867067,
    ])


def test_actual_and_independent_workbook_audits_are_inline_bounded_and_stable():
    """Catches single-candidate traces being replaced by reference metadata only."""
    source = workbook_bytes_with_rows([
        valid_row_values(**{'Defect Length Basis': 'Actual defect length'}),
        valid_row_values(**{'Defect Length Basis': 'Independent defects'}),
    ])

    first = process_workbook(source, FIXED_TIME, 'single-candidates.xlsx')
    first_workbook = _workbook(first.workbook_bytes)
    main = first_workbook['Batch Input & Results']
    headings = tuple(cell.value for cell in main[1])
    b31g_column = headings.index('B31G Detail') + 1
    references = [json.loads(main.cell(row, b31g_column).value) for row in (2, 3)]

    assert [reference['governing_defect_id'] for reference in references] == [
        'Actual/combined defect', 'Independent 10x10 mm defects',
    ]
    assert [reference['inline_candidate']['length_mm'] for reference in references] == [
        100.0, 10.0,
    ]
    assert [reference['inline_candidate']['remaining_wall_mm'] for reference in references] == [
        4.5, 4.5,
    ]
    inline_keys = {
        'defect_id', 'length_mm', 'remaining_wall_mm', 'method', 'd_over_t',
        'length_parameter_z', 'folias_factor', 'flow_stress_mpa',
        'failure_stress_mpa', 'failure_pressure_bar', 'safe_pressure_bar',
        'safety_factor', 'operating_hoop_stress_mpa', 'applicable',
        'acceptable', 'credited_safe_pressure_bar', 'governing', 'warning_codes',
    }
    assert all(set(reference['inline_candidate']) == inline_keys for reference in references)
    for row, reference in zip((2, 3), references, strict=True):
        value = main.cell(row, b31g_column).value
        assert len(value) < 2000
        assert value == json.dumps(reference, sort_keys=True, separators=(',', ':'))
        assert reference['inline_candidate']['credited_safe_pressure_bar'] == pytest.approx(
            main.cell(row, headings.index('Effective Pipe Capacity [bar]') + 1).value
        )

    second = process_workbook(
        first.workbook_bytes, FIXED_TIME, 'single-candidates-processed.xlsx',
    )
    second_main = _workbook(second.workbook_bytes)['Batch Input & Results']
    assert [second_main.cell(row, b31g_column).value for row in (2, 3)] == [
        main.cell(row, b31g_column).value for row in (2, 3)
    ]


def test_high_smys_actual_b31g_detail_is_strict_bounded_json():
    """Catches Original-B31G limiting Folias leaking as bare JSON Infinity."""
    source = workbook_bytes_with_rows([valid_row_values(**{
        'Pipe OD [mm]': 1016.0,
        'Nominal Wall [mm]': 12.0,
        'Pipe Yield [MPa]': 550.0,
        'Defect Length [mm]': 1000.0,
    })])

    result = process_workbook(source, FIXED_TIME, 'actual-high-smys.xlsx')
    workbook = _workbook(result.workbook_bytes)
    main = workbook['Batch Input & Results']
    headings = tuple(cell.value for cell in main[1])
    value = main.cell(2, headings.index('B31G Detail') + 1).value

    def reject_nonstandard_constant(constant):
        pytest.fail(f'B31G Detail contains non-standard JSON constant {constant}')

    reference = json.loads(value, parse_constant=reject_nonstandard_constant)
    inline = reference['inline_candidate']

    assert result.status_counts == {'REVIEW REQUIRED': 1}
    assert len(value) < 2000
    assert inline['method'] == 'original'
    assert inline['length_parameter_z'] == pytest.approx(82.02099737532808)
    assert inline['folias_factor'] == 'Infinity'
    assert isinstance(inline['safe_pressure_bar'], float)


def test_b31g_detail_serializer_rejects_an_unnormalized_nonfinite_float():
    """Catches a future audit field bypassing normalization into invalid JSON."""
    with pytest.raises(ValueError, match='Out of range float values'):
        _output_value('B31G Detail', {'folias_factor': float('inf')})


def test_high_smys_manual_folias_is_explicit_and_stable_on_reupload():
    """Catches a limiting Manual Folias value becoming a blank detail cell."""
    source = workbook_bytes_with_rows(
        [valid_row_values(**{
            'Pipe OD [mm]': 1016.0,
            'Nominal Wall [mm]': 12.0,
            'Pipe Yield [MPa]': 550.0,
            'Defect Length [mm]': 1000.0,
            'Defect Length Basis': ENTER_MANUALLY,
            'Repair Group ID': 'R-LONG',
            'Remaining Wall [mm]': None,
        })],
        detail_rows=[detail_values(
            group='R-LONG', defect='D-LONG', length=1000.0, wall=4.5,
        )],
    )

    first = process_workbook(source, FIXED_TIME, 'manual-high-smys.xlsx')
    first_workbook = _workbook(first.workbook_bytes)
    first_main = first_workbook['Batch Input & Results']
    first_detail = first_workbook['Individual Defects']
    main_headings = tuple(cell.value for cell in first_main[1])
    detail_headings = tuple(cell.value for cell in first_detail[1])
    method_column = detail_headings.index('B31G Method') + 1
    z_column = detail_headings.index('B31G Length Parameter z') + 1
    folias_column = detail_headings.index('B31G Folias Factor M') + 1
    b31g_column = main_headings.index('B31G Detail') + 1

    assert first.status_counts == {'REVIEW REQUIRED': 1}
    assert first_detail.cell(2, method_column).value == 'original'
    assert first_detail.cell(2, z_column).value == pytest.approx(82.02099737532808)
    assert first_detail.cell(2, folias_column).value == 'Infinity'

    second = process_workbook(
        first.workbook_bytes, FIXED_TIME, 'manual-high-smys-reupload.xlsx',
    )
    second_workbook = _workbook(second.workbook_bytes)
    second_main = second_workbook['Batch Input & Results']
    second_detail = second_workbook['Individual Defects']

    assert second_detail.cell(2, method_column).value == 'original'
    assert second_detail.cell(2, z_column).value == first_detail.cell(2, z_column).value
    assert second_detail.cell(2, folias_column).value == 'Infinity'
    assert second_main.cell(2, b31g_column).value == first_main.cell(2, b31g_column).value


def test_legacy_actual_upgrade_preserves_inline_candidate_audit_on_reupload():
    legacy = legacy_workbook_bytes_with_rows([valid_row_values()])

    first = process_workbook(legacy, FIXED_TIME, 'legacy.xlsx')
    first_workbook = _workbook(first.workbook_bytes)
    main = first_workbook['Batch Input & Results']
    headings = tuple(cell.value for cell in main[1])
    b31g_column = headings.index('B31G Detail') + 1
    reference = json.loads(main.cell(2, b31g_column).value)

    assert reference['detail_schema_version'] == '2'
    assert reference['detail_excel_row_range'] is None
    assert reference['inline_candidate']['defect_id'] == 'Actual/combined defect'
    assert reference['inline_candidate']['length_mm'] == 100.0
    assert reference['inline_candidate']['remaining_wall_mm'] == 4.5

    second = process_workbook(first.workbook_bytes, FIXED_TIME, 'legacy-reupload.xlsx')
    assert _workbook(second.workbook_bytes)['Batch Input & Results'].cell(
        2, b31g_column,
    ).value == main.cell(2, b31g_column).value


def test_warning_register_scans_detail_rows_501_502_and_2001():
    """Catches detail warnings being scanned with the 500-row main-table limit."""
    workbook = _workbook(workbook_bytes_with_rows([_manual_row()]))
    detail = workbook['Individual Defects']
    for excel_row in (501, 502, 2001):
        values = detail_values(
            group='R-001', defect=f'D-{excel_row}', length=10, wall=9.652,
        )
        for column, header in enumerate(DETAIL_INPUT_HEADERS, start=1):
            detail.cell(excel_row, column, values[header])

    result = process_workbook(_saved(workbook), FIXED_TIME, 'detail-boundaries.xlsx')
    warnings = _workbook(result.workbook_bytes)['Warnings']
    affected_rows = {
        warnings.cell(row, 1).value: warnings.cell(row, 3).value
        for row in range(4, warnings.max_row + 1)
    }

    assert affected_rows['W013'] == (
        'Main 2; Individual Defects 501, 502, 2001'
    )


def test_501_detail_candidates_use_bounded_valid_json_and_scalar_audit_rows():
    """Catches Excel truncating a large opaque candidate collection to invalid JSON."""
    source = workbook_bytes_with_rows(
        [_manual_row()],
        detail_rows=[
            detail_values(
                group='R-001', defect=f'D-{index:04d}', length=10, wall=9.652,
            )
            for index in range(1, 502)
        ],
    )

    result = process_workbook(source, FIXED_TIME, '501-details.xlsx')
    workbook = _workbook(result.workbook_bytes)
    main = workbook['Batch Input & Results']
    detail = workbook['Individual Defects']
    main_headings = tuple(cell.value for cell in main[1])
    detail_headings = tuple(cell.value for cell in detail[1])
    b31g_detail = main.cell(2, main_headings.index('B31G Detail') + 1).value

    try:
        reference = json.loads(b31g_detail)
    except json.JSONDecodeError as error:
        pytest.fail(
            f'B31G Detail must be bounded valid JSON; got {len(b31g_detail)} '
            f'characters and decode error {error}'
        )
    assert len(b31g_detail) < 1000
    assert reference == {
        'candidate_count': 501,
        'detail_excel_row_range': '2:502',
        'detail_schema': 'Individual Defects',
        'detail_schema_version': '2',
        'governing_defect_id': 'D-0001',
        'inline_candidate': None,
    }
    last_audit = {
        heading: detail.cell(502, detail_headings.index(heading) + 1).value
        for heading in DETAIL_OUTPUT_HEADERS
    }
    assert last_audit['Source Excel Row'] == 502
    assert last_audit['Calculation Status'] == 'OK'
    assert last_audit['B31G d/t'] == pytest.approx(0.19566666666666674)
    assert last_audit['B31G Safe Pressure [bar]'] == pytest.approx(88.2257484144555)
    assert last_audit['B31G Operating Hoop Stress [MPa]'] == pytest.approx(
        444.07666666666677
    )
    assert last_audit['Assessment Warning Codes'] == 'W013'


def test_full_2000_detail_audit_is_bounded_and_stable_on_reupload():
    """Catches maximum-size linked audits overflowing or changing on rebuild."""
    main_row = _manual_row()
    main_row['Design Pressure [bar]'] = 50.0
    main_row['Run Type A / Class 3 Check'] = 'No'
    source = workbook_bytes_with_rows(
        [main_row],
        detail_rows=[
            detail_values(
                group='R-001', defect=f'D-{index:04d}', length=10, wall=9.652,
            )
            for index in range(1, MAX_DETAIL_ROWS + 1)
        ],
    )

    first = process_workbook(source, FIXED_TIME, '2000-details.xlsx')
    first_workbook = _workbook(first.workbook_bytes)
    main = first_workbook['Batch Input & Results']
    detail = first_workbook['Individual Defects']
    main_headings = tuple(cell.value for cell in main[1])
    detail_headings = tuple(cell.value for cell in detail[1])
    b31g_column = main_headings.index('B31G Detail') + 1
    first_reference_text = main.cell(2, b31g_column).value
    first_reference = json.loads(first_reference_text)

    assert len(first_reference_text) < 1000
    assert first_reference == {
        'candidate_count': 2000,
        'detail_excel_row_range': '2:2001',
        'detail_schema': 'Individual Defects',
        'detail_schema_version': '2',
        'governing_defect_id': 'D-0001',
        'inline_candidate': None,
    }
    assert 'candidates' not in first_reference
    audit_headings = (
        'Source Excel Row',
        'Calculation Status',
        'B31G d/t',
        'B31G Length Parameter z',
        'B31G Folias Factor M',
        'B31G Flow Stress [MPa]',
        'B31G Estimated Failure Stress [MPa]',
        'B31G Failure Pressure [bar]',
        'B31G Safe Pressure [bar]',
        'B31G Safety Factor',
        'B31G Operating Hoop Stress [MPa]',
        'B31G Applicable',
        'B31G Acceptable',
        'Credited Safe Pressure [bar]',
        'Governing Defect',
    )
    first_last_audit = tuple(
        detail.cell(2001, detail_headings.index(heading) + 1).value
        for heading in audit_headings
    )
    assert first_last_audit[0:2] == (2001, 'OK')
    assert all(value is not None for value in first_last_audit[2:-1])
    assert first_last_audit[-1] is None

    second = process_workbook(
        first.workbook_bytes, FIXED_TIME, '2000-details-processed.xlsx',
    )
    second_workbook = _workbook(second.workbook_bytes)
    second_main = second_workbook['Batch Input & Results']
    second_detail = second_workbook['Individual Defects']

    assert second_main.cell(2, b31g_column).value == first_reference_text
    assert tuple(
        second_detail.cell(2001, detail_headings.index(heading) + 1).value
        for heading in audit_headings
    ) == first_last_audit


def test_partial_detail_with_trimmed_group_invalidates_only_its_manual_owner():
    """Catches losing the owner ID when a partially populated detail row fails parsing."""
    source = workbook_bytes_with_rows(
        [_manual_row(), valid_row_values(**{'Pipe OD [mm]': 508.0})],
        detail_rows=[detail_values(
            group='  R-001  ', defect='D-01', length=None, wall=9.0,
        )],
    )

    processed = process_workbook(source, FIXED_TIME, 'partial-detail.xlsx')
    workbook = _workbook(processed.workbook_bytes)
    main = workbook['Batch Input & Results']
    detail = workbook['Individual Defects']
    main_headings = tuple(cell.value for cell in main[1])
    detail_headings = tuple(cell.value for cell in detail[1])

    assert processed.status_counts == {'INPUT ERROR': 1, 'OK': 1}
    assert main.cell(2, main_headings.index('Error Code') + 1).value.find(
        'INVALID_INDIVIDUAL_DEFECTS'
    ) >= 0
    assert main.cell(3, main_headings.index('Calculation Status') + 1).value == 'OK'
    assert detail.cell(2, detail_headings.index('Calculation Status') + 1).value == (
        'INPUT ERROR'
    )
    assert 'REQUIRED_VALUE' in detail.cell(
        2, detail_headings.index('Error Code') + 1,
    ).value


def test_invalid_orphan_detail_remains_local_and_does_not_block_main_rows():
    """Catches an unowned detail error leaking into an unrelated repair result."""
    source = workbook_bytes_with_rows(
        [valid_row_values()],
        detail_rows=[detail_values(
            group='ORPHAN', defect='D-X', length=None, wall=4.0,
        )],
    )

    processed = process_workbook(source, FIXED_TIME, 'orphan-detail.xlsx')
    workbook = _workbook(processed.workbook_bytes)
    detail = workbook['Individual Defects']
    headings = tuple(cell.value for cell in detail[1])

    assert processed.status_counts == {'OK': 1}
    assert detail.cell(2, headings.index('Calculation Status') + 1).value == 'INPUT ERROR'
    assert set(detail.cell(2, headings.index('Error Code') + 1).value.split('; ')) == {
        'REQUIRED_VALUE', 'ORPHAN_REPAIR_GROUP',
    }


def test_processed_manual_workbook_reuploads_without_changing_results():
    """Catches trusting stale outputs or changing linked tuple order on re-upload."""
    source = workbook_bytes_with_rows(
        [_manual_row()],
        detail_rows=[
            detail_values(group='R-001', defect='D-01', length=10, wall=9.652),
            detail_values(group='R-001', defect='D-02', length=35, wall=10.0),
        ],
    )

    first = process_workbook(source, FIXED_TIME, 'manual.xlsx')
    second = process_workbook(first.workbook_bytes, FIXED_TIME, 'manual-processed.xlsx')

    assert _result_signature(second.workbook_bytes) == _result_signature(first.workbook_bytes)


def test_manual_system_error_marks_details_without_breaking_row_continuation(
    monkeypatch,
):
    """Catches strict candidate mapping turning a row-local failure into a batch crash."""
    import workbook_processor

    def explode(*_args, **_kwargs):
        raise RuntimeError('controlled test failure')

    monkeypatch.setattr(workbook_processor, 'calculate_row', explode)
    source = workbook_bytes_with_rows(
        [_manual_row(), valid_row_values()],
        detail_rows=[
            detail_values(group='R-001', defect='D-01', length=10, wall=9.652),
            detail_values(group='R-001', defect='D-02', length=35, wall=10.0),
        ],
    )

    processed = process_workbook(source, FIXED_TIME, 'system-error.xlsx')
    workbook = _workbook(processed.workbook_bytes)
    detail = workbook['Individual Defects']
    headings = tuple(cell.value for cell in detail[1])

    assert processed.status_counts == {'SYSTEM ERROR': 2}
    assert [detail.cell(row, headings.index('Calculation Status') + 1).value for row in (2, 3)] == [
        'INPUT ERROR', 'INPUT ERROR',
    ]
    assert [detail.cell(row, headings.index('Error Code') + 1).value for row in (2, 3)] == [
        'MAIN_ROW_NOT_CALCULATED', 'MAIN_ROW_NOT_CALCULATED',
    ]


def _capture_successful_mechanisms(monkeypatch):
    import workbook_processor

    received = []

    def calculate_successfully(_batch_info, row):
        received.append(row.values['Mechanism'])
        return RowCalculation(
            source_excel_row=row.source_excel_row,
            status=CalculationStatus.OK,
            outputs={},
        )

    monkeypatch.setattr(workbook_processor, 'calculate_row', calculate_successfully)
    return received


def test_legacy_dent_is_canonical_through_preview_processing_and_reupload(monkeypatch):
    """Catches old Dent uploads leaking the legacy name past the workbook boundary."""
    received = _capture_successful_mechanisms(monkeypatch)
    source = workbook_bytes_with_rows(
        [valid_row_values(Mechanism='Dent')],
        commercial_inputs={'B3': 50.0, 'E3': 20.0, 'H3': 1.5},
    )

    inspection = inspect_workbook(source)
    first = process_workbook(source, processed_at=FIXED_TIME)
    first_workbook = _workbook(first.workbook_bytes)
    reinspection = inspect_workbook(first.workbook_bytes)
    second = process_workbook(first.workbook_bytes, processed_at=FIXED_TIME)
    second_workbook = _workbook(second.workbook_bytes)

    assert inspection.preview[0]['Mechanism'] == 'Dent w/crack'
    assert reinspection.preview[0]['Mechanism'] == 'Dent w/crack'
    assert received and set(received) == {'Dent w/crack'}
    assert first_workbook['Batch Input & Results']['F2'].value == 'Dent w/crack'
    assert first_workbook['Cost Calculation']['F6'].value == 'Dent w/crack'
    assert second_workbook['Batch Input & Results']['F2'].value == 'Dent w/crack'
    assert second_workbook['Cost Calculation']['F6'].value == 'Dent w/crack'
    assert first_workbook['Batch Input & Results']['A2'].value == 457.2
    main_headings = tuple(
        cell.value for cell in first_workbook['Batch Input & Results'][1]
    )
    assert first_workbook['Batch Input & Results'].cell(
        2, main_headings.index('Prowrap CF Cloth Width [mm]') + 1,
    ).value == 300.0
    assert first_workbook['Batch Information']['B3'].value == 'Batch Customer'
    assert [second_workbook['Cost Calculation'][address].value for address in (
        'B3', 'E3', 'H3',
    )] == [50.0, 20.0, 1.5]


def test_dent_no_crack_is_stable_through_preview_processing_and_reupload(monkeypatch):
    """Catches a current uncracked-dent choice being changed at the workbook boundary."""
    received = _capture_successful_mechanisms(monkeypatch)
    source = workbook_bytes_with_rows([valid_row_values(Mechanism='Dent no-crack')])

    inspection = inspect_workbook(source)
    first = process_workbook(source, processed_at=FIXED_TIME)
    first_workbook = _workbook(first.workbook_bytes)
    reinspection = inspect_workbook(first.workbook_bytes)
    second = process_workbook(first.workbook_bytes, processed_at=FIXED_TIME)
    second_workbook = _workbook(second.workbook_bytes)

    assert inspection.preview[0]['Mechanism'] == 'Dent no-crack'
    assert reinspection.preview[0]['Mechanism'] == 'Dent no-crack'
    assert received and set(received) == {'Dent no-crack'}
    assert first_workbook['Batch Input & Results']['F2'].value == 'Dent no-crack'
    assert first_workbook['Cost Calculation']['F6'].value == 'Dent no-crack'
    assert second_workbook['Batch Input & Results']['F2'].value == 'Dent no-crack'
    assert second_workbook['Cost Calculation']['F6'].value == 'Dent no-crack'


def test_preview_uses_the_same_qualification_review_status_as_processing():
    source = workbook_bytes_with_rows([valid_row_values(**{
        'Operating Temperature [degC]': 150.0,
    })])

    inspection = inspect_workbook(source)
    processed = process_workbook(source, processed_at=FIXED_TIME)

    assert inspection.preview[0]['Calculation Status'] == 'REVIEW REQUIRED'
    assert processed.status_counts == {'REVIEW REQUIRED': 1}


def test_missing_batch_information_is_a_workbook_error():
    source = workbook_bytes_with_rows([valid_row_values()])
    workbook = _workbook(source)
    workbook['Batch Information']['B3'] = None

    inspection = inspect_workbook(_saved(workbook))

    assert inspection.batch_info is None
    assert [issue.code for issue in inspection.workbook_errors] == ['REQUIRED_VALUE']


@pytest.mark.parametrize('worksheet', [
    'Batch Information', 'Batch Input & Results', 'Individual Defects',
    'Summary', 'Instructions', 'Lists',
])
def test_missing_required_worksheet_is_a_workbook_error(worksheet):
    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    del workbook[worksheet]

    inspection = inspect_workbook(_saved(workbook))

    assert [issue.code for issue in inspection.workbook_errors] == ['MISSING_WORKSHEET']


def test_extra_worksheet_is_a_workbook_error():
    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    workbook.create_sheet('Unexpected')

    inspection = inspect_workbook(_saved(workbook))

    assert [issue.code for issue in inspection.workbook_errors] == ['UNEXPECTED_WORKSHEET']


def test_missing_or_duplicate_headings_are_workbook_errors():
    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    data = workbook['Batch Input & Results']
    data['A1'] = 'Outside Diameter [mm]'

    inspection = inspect_workbook(_saved(workbook))
    assert [issue.code for issue in inspection.workbook_errors] == ['INVALID_INPUT_HEADERS']

    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    data = workbook['Batch Input & Results']
    data.cell(1, 2).value = INPUT_HEADERS[0]

    inspection = inspect_workbook(_saved(workbook))
    assert [issue.code for issue in inspection.workbook_errors] == ['DUPLICATE_INPUT_HEADER']


def test_header_inspection_identifies_recognized_missing_and_unexpected_headings():
    """Catch a generic header failure that leaves users unable to correct their workbook."""
    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    data = workbook['Batch Input & Results']
    data['A1'] = 'Outside Diameter [mm]'

    inspection = inspect_workbook(_saved(workbook))

    assert [issue.code for issue in inspection.workbook_errors] == ['INVALID_INPUT_HEADERS']
    assert inspection.recognized_input_headers == INPUT_HEADERS[1:]
    assert inspection.missing_input_headers == ('Pipe OD [mm]',)
    assert inspection.unexpected_headers == ('Outside Diameter [mm]',)


def test_oversized_and_macro_enabled_workbooks_are_rejected_before_processing():
    oversized = b'x' * (MAX_UPLOAD_BYTES + 1)
    assert [issue.code for issue in inspect_workbook(oversized).workbook_errors] == ['FILE_TOO_LARGE']

    source = workbook_bytes_with_rows([valid_row_values()])
    macro = BytesIO()
    with zipfile.ZipFile(BytesIO(source)) as input_zip, zipfile.ZipFile(macro, 'w') as output_zip:
        for entry in input_zip.infolist():
            output_zip.writestr(entry, input_zip.read(entry.filename))
        output_zip.writestr('xl/vbaProject.bin', b'not-a-real-macro')

    assert [issue.code for issue in inspect_workbook(macro.getvalue()).workbook_errors] == ['MACROS_NOT_ALLOWED']


def test_formula_inputs_and_corrupt_or_encrypted_data_are_rejected_safely():
    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    workbook['Batch Information']['B3'] = '=CONCAT("Batch", " Customer")'
    inspection = inspect_workbook(_saved(workbook))
    assert [issue.code for issue in inspection.workbook_errors] == ['FORMULA_NOT_ALLOWED']

    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    workbook['Batch Input & Results']['A2'] = '=1+1'
    inspection = inspect_workbook(_saved(workbook))
    assert [issue.code for issue in inspection.workbook_errors] == ['FORMULA_NOT_ALLOWED']

    for data in (b'not an xlsx', _encrypted_zip_bytes()):
        inspection = inspect_workbook(data)
        assert [issue.code for issue in inspection.workbook_errors] == ['UNREADABLE_WORKBOOK']


@pytest.mark.parametrize(('worksheet_name', 'coordinate'), [
    ('Batch Input & Results', 'AY501'),
    ('Summary', 'B30'),
])
def test_formula_anywhere_in_the_controlled_workbook_is_rejected(worksheet_name, coordinate):
    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    workbook[worksheet_name][coordinate] = '=1+1'

    inspection = inspect_workbook(_saved(workbook))

    assert [issue.code for issue in inspection.workbook_errors] == ['FORMULA_NOT_ALLOWED']
    assert coordinate in inspection.workbook_errors[0].message


@pytest.mark.parametrize('formula', [
    ArrayFormula(ref='A2', text='=1+1'),
    DataTableFormula(ref='A2'),
])
def test_non_string_excel_formula_objects_are_rejected(formula):
    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    workbook['Batch Input & Results']['A2'].value = formula

    inspection = inspect_workbook(_saved(workbook))

    assert [issue.code for issue in inspection.workbook_errors] == ['FORMULA_NOT_ALLOWED']
    assert 'A2' in inspection.workbook_errors[0].message


def test_formula_scan_uses_loaded_cells_without_dense_worksheet_iteration(monkeypatch):
    import workbook_processor

    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    workbook['Summary'].cell(1_048_576, 16_384).value = '=1+1'
    for worksheet in workbook.worksheets:
        monkeypatch.setattr(
            worksheet,
            'iter_rows',
            lambda: (_ for _ in ()).throw(AssertionError('dense iteration used')),
        )

    issues = workbook_processor._formula_errors(workbook)

    assert [issue.code for issue in issues] == ['FORMULA_NOT_ALLOWED']
    assert 'Summary!XFD1048576' in issues[0].message


def test_zip_valid_workbook_with_malformed_xml_is_rejected_safely():
    source = workbook_bytes_with_rows([valid_row_values()])
    malformed = BytesIO()
    with zipfile.ZipFile(BytesIO(source)) as input_zip, zipfile.ZipFile(malformed, 'w') as output_zip:
        for entry in input_zip.infolist():
            content = input_zip.read(entry.filename)
            if entry.filename == 'xl/workbook.xml':
                content = b'<workbook><sheets>'
            output_zip.writestr(entry, content)

    inspection = inspect_workbook(malformed.getvalue())

    assert [issue.code for issue in inspection.workbook_errors] == ['UNREADABLE_WORKBOOK']


def test_zip_expansion_bomb_is_rejected_before_openpyxl_parsing():
    source = workbook_bytes_with_rows([valid_row_values()])
    compressed = BytesIO()
    with zipfile.ZipFile(BytesIO(source)) as input_zip, zipfile.ZipFile(
        compressed, 'w', compression=zipfile.ZIP_DEFLATED,
    ) as output_zip:
        for entry in input_zip.infolist():
            output_zip.writestr(entry, input_zip.read(entry.filename))
        output_zip.writestr('xl/large-compressed-payload.bin', b'x' * (2 * 1024 * 1024))

    inspection = inspect_workbook(compressed.getvalue())

    assert [issue.code for issue in inspection.workbook_errors] == ['UNREADABLE_WORKBOOK']


def test_dense_worksheet_is_rejected_before_openpyxl_materializes_cells(monkeypatch):
    """Catches small compressed uploads that amplify into excessive cell objects."""
    import workbook_processor

    source = _dense_workbook_bytes(100_001)
    assert len(source) < MAX_UPLOAD_BYTES

    def fail_if_loaded(*_args, **_kwargs):
        raise AssertionError('openpyxl must not parse an over-limit worksheet')

    monkeypatch.setattr(workbook_processor, 'load_workbook', fail_if_loaded)

    inspection = workbook_processor.inspect_workbook(source)

    assert [issue.code for issue in inspection.workbook_errors] == [
        'UNREADABLE_WORKBOOK',
    ]
    assert inspection.workbook_errors[0].message == (
        'The uploaded workbook contains too many worksheet cells.'
    )


def test_relocated_dense_worksheet_is_rejected_before_openpyxl(monkeypatch):
    """Catches a worksheet relationship moved outside xl/worksheets bypassing the cap."""
    import workbook_processor

    source = _relocated_dense_workbook_bytes(100_001)
    assert len(source) < MAX_UPLOAD_BYTES

    def fail_if_loaded(*_args, **_kwargs):
        raise AssertionError('openpyxl reached through relocated worksheet bypass')

    monkeypatch.setattr(workbook_processor, 'load_workbook', fail_if_loaded)

    inspection = workbook_processor.inspect_workbook(source)

    assert [issue.code for issue in inspection.workbook_errors] == [
        'UNREADABLE_WORKBOOK',
    ]
    assert inspection.workbook_errors[0].message == (
        'The uploaded workbook contains too many worksheet cells.'
    )


def test_relationship_worksheet_with_nonworksheet_content_type_rejects_before_openpyxl(
    monkeypatch,
):
    """Catches content-type relabeling hiding a relationship-reachable dense sheet."""
    import workbook_processor

    source = _relabeled_dense_worksheet_bytes(100_001)
    assert len(source) < MAX_UPLOAD_BYTES

    def fail_if_loaded(*_args, **_kwargs):
        raise AssertionError('openpyxl reached content-type-inconsistent worksheet')

    monkeypatch.setattr(workbook_processor, 'load_workbook', fail_if_loaded)

    inspection = workbook_processor.inspect_workbook(source)

    assert [issue.code for issue in inspection.workbook_errors] == [
        'UNREADABLE_WORKBOOK',
    ]
    assert inspection.workbook_errors[0].message == (
        'The uploaded workbook has inconsistent worksheet declarations.'
    )


def test_foreign_namespace_duplicate_relationship_cannot_shadow_dense_target(
    monkeypatch,
):
    """A foreign duplicate must not replace the genuine OPC relationship."""
    import workbook_processor

    source = _dense_workbook_with_duplicate_relationship(
        100_001,
        relationship_namespace='urn:protap:foreign-relationships',
    )

    def fail_if_loaded(*_args, **_kwargs):
        raise AssertionError('openpyxl reached foreign-namespace shadow bypass')

    monkeypatch.setattr(workbook_processor, 'load_workbook', fail_if_loaded)

    inspection = workbook_processor.inspect_workbook(source)

    assert [issue.code for issue in inspection.workbook_errors] == [
        'UNREADABLE_WORKBOOK',
    ]
    assert inspection.workbook_errors[0].message == (
        'The uploaded workbook has inconsistent worksheet declarations.'
    )


def test_duplicate_opc_relationship_id_is_rejected_before_openpyxl(monkeypatch):
    """Ambiguous genuine OPC relationship IDs are invalid package metadata."""
    import workbook_processor

    source = _dense_workbook_with_duplicate_relationship(100_001)

    def fail_if_loaded(*_args, **_kwargs):
        raise AssertionError('openpyxl reached duplicate OPC relationship bypass')

    monkeypatch.setattr(workbook_processor, 'load_workbook', fail_if_loaded)

    inspection = workbook_processor.inspect_workbook(source)

    assert [issue.code for issue in inspection.workbook_errors] == [
        'UNREADABLE_WORKBOOK',
    ]
    assert inspection.workbook_errors[0].message == (
        'The uploaded workbook has inconsistent worksheet declarations.'
    )


def test_explicit_internal_relationship_target_mode_is_accepted():
    source = _explicit_internal_relationship_workbook_bytes()

    inspection = inspect_workbook(source)

    assert inspection.workbook_errors == ()
    assert inspection.populated_rows == 1


def test_arbitrary_suffix_dense_worksheet_is_rejected_before_openpyxl(monkeypatch):
    """Catches a valid worksheet content type using a non-XML part suffix."""
    import workbook_processor

    source = _relocated_dense_workbook_bytes(
        100_001,
        new_path='xl/custom/dense.data',
    )
    assert len(source) < MAX_UPLOAD_BYTES

    def fail_if_loaded(*_args, **_kwargs):
        raise AssertionError('openpyxl reached through worksheet-suffix bypass')

    monkeypatch.setattr(workbook_processor, 'load_workbook', fail_if_loaded)

    inspection = workbook_processor.inspect_workbook(source)

    assert [issue.code for issue in inspection.workbook_errors] == [
        'UNREADABLE_WORKBOOK',
    ]
    assert inspection.workbook_errors[0].message == (
        'The uploaded workbook contains too many worksheet cells.'
    )


def test_malformed_arbitrary_suffix_worksheet_is_rejected_before_openpyxl(monkeypatch):
    """Catches recognized non-XML-suffix worksheets escaping malformed handling."""
    import workbook_processor

    source = _relocated_worksheet_workbook_bytes(
        (
            b'<worksheet '
            b'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            b'<sheetData><row>'
        ),
        new_path='xl/custom/malformed.data',
    )

    def fail_if_loaded(*_args, **_kwargs):
        raise AssertionError('openpyxl reached malformed recognized worksheet')

    monkeypatch.setattr(workbook_processor, 'load_workbook', fail_if_loaded)

    inspection = workbook_processor.inspect_workbook(source)

    assert [issue.code for issue in inspection.workbook_errors] == [
        'UNREADABLE_WORKBOOK',
    ]


def test_unrelated_binary_custom_part_is_not_parsed_as_xml():
    """Catches suffix-independent scanning that probes unrelated binary package parts."""
    source = _workbook_with_binary_custom_part()

    inspection = inspect_workbook(source)

    assert inspection.workbook_errors == ()
    assert inspection.populated_rows == 1


def test_strict_spreadsheetml_dense_worksheet_is_still_bounded(monkeypatch):
    """Catches namespace filtering that disables the cap for ISO Strict sheets."""
    import workbook_processor

    source = _dense_workbook_bytes(
        100_001,
        namespace='http://purl.oclc.org/ooxml/spreadsheetml/main',
    )

    def fail_if_loaded(*_args, **_kwargs):
        raise AssertionError('openpyxl reached through strict-namespace bypass')

    monkeypatch.setattr(workbook_processor, 'load_workbook', fail_if_loaded)

    inspection = workbook_processor.inspect_workbook(source)

    assert [issue.code for issue in inspection.workbook_errors] == [
        'UNREADABLE_WORKBOOK',
    ]
    assert inspection.workbook_errors[0].message == (
        'The uploaded workbook contains too many worksheet cells.'
    )


def test_foreign_custom_xml_with_worksheet_names_is_ignored():
    """Catches local-name scanning that treats unrelated customer XML as a sheet."""
    source = _foreign_custom_xml_workbook_bytes(100_001)
    assert len(source) < MAX_UPLOAD_BYTES

    inspection = inspect_workbook(source)

    assert inspection.workbook_errors == ()
    assert inspection.populated_rows == 1


def test_foreign_cell_elements_inside_a_real_worksheet_extension_are_ignored():
    """Catches foreign extension elements being counted as SpreadsheetML cells."""
    source = _worksheet_with_foreign_extension_bytes(100_001)
    assert len(source) < MAX_UPLOAD_BYTES

    with pytest.warns(UserWarning, match='Unknown extension is not supported'):
        inspection = inspect_workbook(source)

    assert inspection.workbook_errors == ()
    assert inspection.populated_rows == 1


def test_more_than_500_populated_rows_is_a_workbook_error():
    source = workbook_bytes_with_rows([valid_row_values() for _ in range(501)])

    inspection = inspect_workbook(source)

    assert inspection.populated_rows == 0
    assert [issue.code for issue in inspection.workbook_errors] == ['INPUT_ROW_OUT_OF_RANGE']


@pytest.mark.parametrize('excel_row', [502, 1000])
def test_populated_input_beyond_controlled_rows_is_a_workbook_error(excel_row):
    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    workbook['Batch Input & Results'].cell(excel_row, 1).value = 508.0

    inspection = inspect_workbook(_saved(workbook))

    assert [issue.code for issue in inspection.workbook_errors] == ['INPUT_ROW_OUT_OF_RANGE']
    assert f'A{excel_row}' in inspection.workbook_errors[0].message


def test_far_input_scan_uses_loaded_cells_without_max_row_iteration(monkeypatch):
    import workbook_processor

    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    data = workbook['Batch Input & Results']
    data.cell(1_048_576, 1).value = 508.0
    monkeypatch.setattr(type(data), 'max_row', property(lambda _worksheet: MAX_ROWS + 1))

    issues = workbook_processor._out_of_range_input_errors(data)

    assert [issue.code for issue in issues] == ['INPUT_ROW_OUT_OF_RANGE']
    assert 'A1048576' in issues[0].message


def test_populated_row_scan_stays_within_controlled_input_rectangle(monkeypatch):
    import workbook_processor

    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    data = workbook['Batch Input & Results']
    data.cell(1_048_576, len(INPUT_HEADERS) + len(OUTPUT_HEADERS)).value = 'harmless output note'
    original_cell = data.cell

    def controlled_cell(row=None, column=None, *args, **kwargs):
        if row is not None and row > MAX_ROWS + 1:
            raise AssertionError('uncontrolled row scan')
        return original_cell(row=row, column=column, *args, **kwargs)

    monkeypatch.setattr(data, 'cell', controlled_cell)

    populated = workbook_processor._populated_rows(data)

    assert [excel_row for excel_row, _ in populated] == [2]


def test_formula_issue_has_priority_over_far_input_row_issue():
    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    workbook['Batch Input & Results'].cell(1_048_576, 1).value = 508.0
    workbook['Summary'].cell(1_048_576, 16_384).value = '=1+1'

    inspection = inspect_workbook(_saved(workbook))

    assert [issue.code for issue in inspection.workbook_errors] == ['FORMULA_NOT_ALLOWED']


def test_processed_cost_formulas_and_commercial_inputs_are_safe_to_reupload():
    """Catches deny-all formula scanning or a rebuild that drops cost assumptions."""
    first = process_workbook(
        workbook_bytes_with_rows([valid_row_values()]),
        processed_at=FIXED_TIME,
    )
    workbook = _workbook(first.workbook_bytes)
    cost = workbook['Cost Calculation']
    cost['B3'], cost['E3'], cost['H3'] = 50.0, 20.0, 1.5

    second = process_workbook(_saved(workbook), processed_at=FIXED_TIME)
    regenerated = _workbook(second.workbook_bytes)['Cost Calculation']

    assert [regenerated[address].value for address in ('B3', 'E3', 'H3')] == [
        50.0, 20.0, 1.5,
    ]


def test_whitespace_only_commercial_input_is_rebuilt_as_a_true_blank():
    """Catches whitespace bypassing blank validation and breaking Excel IF checks."""
    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    workbook['Cost Calculation']['B3'] = '   '

    result = process_workbook(_saved(workbook), processed_at=FIXED_TIME)
    rebuilt = _workbook(result.workbook_bytes)['Cost Calculation']

    assert rebuilt['B3'].value is None
    assert rebuilt['U6'].value == cost_formula(6)


def test_altered_cost_formula_is_rejected():
    """Catches broad formula allowlisting in the controlled cost range."""
    result = process_workbook(
        workbook_bytes_with_rows([valid_row_values()]),
        processed_at=FIXED_TIME,
    )
    workbook = _workbook(result.workbook_bytes)
    workbook['Cost Calculation']['U6'] = '=1+1'

    inspection = inspect_workbook(_saved(workbook))

    assert [issue.code for issue in inspection.workbook_errors] == ['FORMULA_NOT_ALLOWED']
    assert 'Cost Calculation!U6' in inspection.workbook_errors[0].message


@pytest.mark.parametrize(('address', 'value'), [
    ('B3', -0.01),
    ('E3', 'twenty'),
    ('H3', float('inf')),
    ('B3', float('nan')),
    ('E3', True),
])
def test_invalid_commercial_inputs_are_rejected(address, value):
    """Catches unsafe or unusable values copied into the trusted output."""
    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    workbook['Cost Calculation'][address] = value

    issues = _commercial_input_errors(workbook)

    assert [issue.code for issue in issues] == ['INVALID_COST_INPUT']
    assert address in issues[0].message


def test_formula_commercial_input_keeps_formula_error_priority():
    """Catches formula injection being downgraded to a generic cost-input error."""
    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    workbook['Cost Calculation']['B3'] = '=1+1'
    workbook['Batch Input & Results'].cell(1_048_576, 1).value = 508.0

    inspection = inspect_workbook(_saved(workbook))

    assert [issue.code for issue in inspection.workbook_errors] == ['FORMULA_NOT_ALLOWED']


def test_changed_cost_heading_is_a_workbook_error():
    """Catches accepting a structurally altered commercial table."""
    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    workbook['Cost Calculation']['A5'] = 'Outside Diameter [mm]'

    inspection = inspect_workbook(_saved(workbook))

    assert [issue.code for issue in inspection.workbook_errors] == [
        'INVALID_COST_HEADERS',
    ]


def test_exactly_500_controlled_rows_remain_valid():
    inspection = inspect_workbook(
        workbook_bytes_with_rows([valid_row_values() for _ in range(500)])
    )

    assert inspection.populated_rows == 500
    assert inspection.workbook_errors == ()


def test_exactly_2000_detail_rows_are_kept_and_first_row_beyond_is_rejected():
    """Catches an off-by-one detail bound or an unbounded detail scan."""
    exact = workbook_bytes_with_rows(
        [valid_row_values()],
        detail_rows=[
            detail_values(group='ORPHAN', defect=f'D-{index}', length=10, wall=4.0)
            for index in range(MAX_DETAIL_ROWS)
        ],
    )
    assert inspect_workbook(exact).workbook_errors == ()

    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    workbook['Individual Defects'].cell(MAX_DETAIL_ROWS + 2, 1).value = 'ORPHAN'
    inspection = inspect_workbook(_saved(workbook))

    assert [issue.code for issue in inspection.workbook_errors] == [
        'DETAIL_ROW_OUT_OF_RANGE',
    ]
    assert f'A{MAX_DETAIL_ROWS + 2}' in inspection.workbook_errors[0].message


def test_one_invalid_row_does_not_stop_valid_rows_and_inputs_are_preserved():
    source = workbook_bytes_with_rows([
        valid_row_values(),
        valid_row_values(**{'Remaining Wall [mm]': 12.0}),
        valid_row_values(),
    ])

    result = process_workbook(source, processed_at=FIXED_TIME)

    assert result.status_counts == {'OK': 2, 'INPUT ERROR': 1}
    assert result.populated_rows == 3
    processed = _workbook(result.workbook_bytes)
    data = processed['Batch Input & Results']
    assert [data.cell(3, column).value for column in range(1, len(INPUT_HEADERS) + 1)] == [
        _workbook(source)['Batch Input & Results'].cell(3, column).value
        for column in range(1, len(INPUT_HEADERS) + 1)
    ]
    headings = tuple(cell.value for cell in data[1])
    status_column = headings.index('Calculation Status') + 1
    error_code_column = headings.index('Error Code') + 1
    error_message_column = headings.index('Error Message') + 1
    assert data.cell(2, status_column).value == 'OK'
    assert data.cell(3, status_column).value == 'INPUT ERROR'
    assert data.cell(3, error_code_column).value == 'OUT_OF_RANGE'
    assert 'Remaining Wall [mm]' in data.cell(3, error_message_column).value
    assert data.cell(4, status_column).value == 'OK'


def test_processed_warning_sheet_consolidates_codes_and_affected_rows():
    """Catches repeated long warning text or one legend entry per defect row."""
    source = workbook_bytes_with_rows([
        valid_row_values(**{'Prowrap CF Cloth Width [mm]': 250.0}),
        valid_row_values(**{'Prowrap CF Cloth Width [mm]': 250.0}),
    ])

    result = process_workbook(source, processed_at=FIXED_TIME)
    workbook = _workbook(result.workbook_bytes)
    data = workbook['Batch Input & Results']
    warnings = workbook['Warnings']

    headings = tuple(cell.value for cell in data[1])
    warning_column = headings.index('Compliance Warnings') + 1
    assert data.cell(2, warning_column).value == 'W018'
    assert data.cell(3, warning_column).value == 'W018'
    assert warnings['A4'].value == 'W018'
    assert '300 mm or 500 mm' in warnings['B4'].value
    assert warnings['C4'].value == '2, 3'
    assert warnings['A4'].font.italic is False
    assert list(warnings.tables) == ['WarningRegister']
    assert warnings.tables['WarningRegister'].ref == 'A3:C4'


def test_processed_workbook_keeps_clear_no_warning_register_state():
    """Catches an empty warning sheet that looks broken or unfinished."""
    result = process_workbook(
        workbook_bytes_with_rows([valid_row_values()]),
        processed_at=FIXED_TIME,
    )
    warnings = _workbook(result.workbook_bytes)['Warnings']

    assert warnings['A4'].value == 'No compliance warnings were generated.'
    assert not warnings.tables


def test_processed_warning_register_remains_filterable_while_protected():
    """Catches sheet protection disabling the warning table filter controls."""
    result = process_workbook(
        workbook_bytes_with_rows([
            valid_row_values(**{'Prowrap CF Cloth Width [mm]': 250.0}),
        ]),
        processed_at=FIXED_TIME,
    )
    warnings = _workbook(result.workbook_bytes)['Warnings']

    assert warnings.protection.sheet is True
    assert warnings.protection.autoFilter is False


def test_previous_five_sheet_template_is_accepted_and_upgraded():
    """Catches a release that strands users holding the previous template."""
    source = legacy_workbook_bytes_with_rows(
        [valid_row_values()], sheet_count=5,
    )
    inspection = inspect_workbook(source)
    result = process_workbook(source, processed_at=FIXED_TIME)

    assert inspection.workbook_errors == ()
    assert _workbook(result.workbook_bytes).sheetnames == [
        'Batch Information', 'Batch Input & Results', 'Individual Defects',
        'Cost Calculation', 'Warnings', 'Summary', 'Instructions', 'Lists',
    ]


def test_previous_six_sheet_template_is_accepted_and_upgraded():
    """Catches a release that strands users holding the warning-register template."""
    source = legacy_workbook_bytes_with_rows(
        [valid_row_values()], sheet_count=6,
    )
    result = process_workbook(source, processed_at=FIXED_TIME)

    assert _workbook(result.workbook_bytes).sheetnames == [
        'Batch Information', 'Batch Input & Results', 'Individual Defects',
        'Cost Calculation', 'Warnings', 'Summary', 'Instructions', 'Lists',
    ]


def test_processed_cost_sheet_maps_requested_values_and_formulas():
    """Catches wrong source-column order or missing controlled formulas."""
    result = process_workbook(
        workbook_bytes_with_rows([valid_row_values()]),
        processed_at=FIXED_TIME,
    )
    workbook = _workbook(result.workbook_bytes)
    source = workbook['Batch Input & Results']
    cost = workbook['Cost Calculation']
    source_headings = tuple(cell.value for cell in source[1])

    assert tuple(cell.value for cell in cost[5]) == COST_TABLE_HEADERS
    assert [cost.cell(6, column).value for column in range(1, 21)] == [
        source.cell(2, source_headings.index(header) + 1).value
        for header in COST_SOURCE_HEADERS
    ]
    assert cost['U6'].value == cost_formula(6)
    assert cost['V6'].value == price_formula(6)
    assert cost.tables['CostRows'].ref == 'A5:V6'


def test_uploaded_cost_table_values_are_never_trusted():
    """Catches user-edited or tampered commercial rows being copied into output."""
    first = process_workbook(
        workbook_bytes_with_rows([valid_row_values()]),
        processed_at=FIXED_TIME,
    )
    edited = _workbook(first.workbook_bytes)
    edited['Cost Calculation']['A6'] = 999999.0
    edited['Cost Calculation']['S6'] = 999999.0

    second = process_workbook(_saved(edited), processed_at=FIXED_TIME)
    regenerated = _workbook(second.workbook_bytes)

    assert regenerated['Cost Calculation']['A6'].value == 457.2
    main = regenerated['Batch Input & Results']
    headings = tuple(cell.value for cell in main[1])
    assert regenerated['Cost Calculation']['S6'].value == main.cell(
        2, headings.index(COST_SOURCE_HEADERS[18]) + 1,
    ).value


def test_processed_cost_sheet_uses_one_compact_row_per_populated_defect():
    """Catches sparse source row numbers leaking into compact cost-table layout."""
    workbook = _workbook(workbook_bytes_with_rows([
        valid_row_values(),
        valid_row_values(**{'Pipe OD [mm]': 508.0}),
        valid_row_values(**{'Pipe OD [mm]': 610.0}),
    ]))
    data = workbook['Batch Input & Results']
    for column in range(1, len(INPUT_HEADERS) + 1):
        data.cell(3, column).value = None

    result = process_workbook(_saved(workbook), processed_at=FIXED_TIME)
    cost = _workbook(result.workbook_bytes)['Cost Calculation']

    assert [cost['A6'].value, cost['A7'].value, cost['A8'].value] == [
        457.2, 610.0, None,
    ]
    assert cost['U7'].value == cost_formula(7)
    assert cost.tables['CostRows'].ref == 'A5:V7'


def test_processed_cost_table_filter_covers_every_compact_row():
    """Catches the table filter retaining the one-row template range."""
    result = process_workbook(
        workbook_bytes_with_rows([
            valid_row_values(),
            valid_row_values(**{'Pipe OD [mm]': 508.0}),
            valid_row_values(**{'Pipe OD [mm]': 610.0}),
        ]),
        processed_at=FIXED_TIME,
    )
    table = _workbook(result.workbook_bytes)['Cost Calculation'].tables['CostRows']

    assert table.ref == 'A5:V8'
    assert table.autoFilter.ref == table.ref


def test_cleared_processed_defect_ignores_stale_exact_cost_formulas():
    """Catches valid generated formulas blocking a safe processed-workbook rebuild."""
    first = process_workbook(
        workbook_bytes_with_rows([valid_row_values()]),
        processed_at=FIXED_TIME,
    )
    edited = _workbook(first.workbook_bytes)
    data = edited['Batch Input & Results']
    for column in range(1, len(INPUT_HEADERS) + 1):
        data.cell(2, column).value = None

    second = process_workbook(_saved(edited), processed_at=FIXED_TIME)
    cost = _workbook(second.workbook_bytes)['Cost Calculation']

    assert second.populated_rows == 0
    assert [cost.cell(6, column).value for column in range(1, 23)] == [None] * 22
    assert cost.tables['CostRows'].ref == 'A5:V6'


def test_processed_workbook_requests_full_automatic_recalculation():
    """Catches cost formulas remaining stale after users edit commercial assumptions."""
    result = process_workbook(
        workbook_bytes_with_rows([valid_row_values()]),
        processed_at=FIXED_TIME,
    )
    workbook = _workbook(result.workbook_bytes)

    assert workbook.calculation.calcMode == 'auto'
    assert workbook.calculation.fullCalcOnLoad is True
    assert workbook.calculation.forceFullCalc is True


def test_processed_workbook_updates_summary_and_uses_stable_diagnostic_json():
    source = workbook_bytes_with_rows([valid_row_values(**{
        'Mechanism': 'Leak',
        'Design Pressure [bar]': 150.0,
    })])

    result = process_workbook(source, processed_at=FIXED_TIME)
    workbook = _workbook(result.workbook_bytes)
    data = workbook['Batch Input & Results']
    summary = workbook['Summary']

    assert result.status_counts == {'NOT REPAIRABLE': 1}
    headings = tuple(cell.value for cell in data[1])
    assert data.cell(2, headings.index('Calculation Status') + 1).value == (
        'NOT REPAIRABLE'
    )
    detail_json = data.cell(2, headings.index('Type B Detail') + 1).value
    assert detail_json.startswith('{')
    assert detail_json == _stable_json(detail_json)
    assert summary['B3'].value == 'Batch Customer'
    assert summary['B4'].value == 'Batch Location'
    assert summary['B5'].value == 'B-001'
    assert summary['B8'].value == '2026-08-14T12:00:00Z'
    assert summary['B10'].value == 1
    assert summary['B15'].value == 1
    assert summary['B24'].value == '1.2.0'
    assert summary['B25'].value == '91b68d6'


def test_processed_workbook_records_the_sanitized_uploaded_source_name():
    source = workbook_bytes_with_rows([valid_row_values()])

    result = process_workbook(
        source,
        processed_at=FIXED_TIME,
        source_name='../../Customer Batch.xlsx',
    )

    assert _workbook(result.workbook_bytes)['Summary']['B7'].value == 'Customer Batch.xlsx'


@pytest.mark.parametrize('source_name', [
    '=1+1.xlsx',
    '+1+1.xlsx',
    '-1+1.xlsx',
    '@danger.xlsx',
    '../../\x00  @unsafe.xlsx',
])
def test_processed_workbook_neutralizes_formula_like_source_names(source_name):
    result = process_workbook(
        workbook_bytes_with_rows([valid_row_values()]),
        processed_at=FIXED_TIME,
        source_name=source_name,
    )

    source_cell = _workbook(result.workbook_bytes)['Summary']['B7']

    assert source_cell.data_type == 's'
    assert source_cell.value.startswith("'")
    assert '\x00' not in source_cell.value


def test_processing_regenerates_a_clean_workbook_after_a_row_is_cleared():
    first = process_workbook(workbook_bytes_with_rows([valid_row_values()]), processed_at=FIXED_TIME)
    edited = _workbook(first.workbook_bytes)
    data = edited['Batch Input & Results']
    for column in range(1, len(INPUT_HEADERS) + 1):
        data.cell(2, column).value = None

    second = process_workbook(_saved(edited), processed_at=FIXED_TIME)
    cleaned = _workbook(second.workbook_bytes)
    output_values = [
        cleaned['Batch Input & Results'].cell(2, column).value
        for column in range(
            len(INPUT_HEADERS) + 1,
            len(INPUT_HEADERS) + len(OUTPUT_HEADERS) + 1,
        )
    ]

    assert second.populated_rows == 0
    assert output_values == [None] * len(output_values)


def test_processing_restores_the_trusted_summary_and_instruction_content():
    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    workbook['Summary']['A27'] = 'TAMPERED DISCLAIMER'
    workbook['Instructions']['A3'] = 'TAMPERED INSTRUCTIONS'
    workbook['Lists'].sheet_state = 'visible'

    result = process_workbook(_saved(workbook), processed_at=FIXED_TIME)
    processed = _workbook(result.workbook_bytes)

    assert 'preliminary screening outputs' in processed['Summary']['A27'].value
    assert processed['Instructions']['A3'].value.startswith('1. Complete Customer')
    assert processed['Lists'].sheet_state == 'hidden'
    assert processed['Batch Input & Results'].protection.sheet is True


def test_process_rejects_workbook_level_errors():
    workbook = _workbook(workbook_bytes_with_rows([valid_row_values()]))
    workbook['Batch Information']['B3'] = None
    with pytest.raises(WorkbookProcessingError, match='Customer'):
        process_workbook(_saved(workbook), processed_at=FIXED_TIME)


def test_unexpected_row_exception_is_logged_with_source_row_only(caplog, monkeypatch):
    import workbook_processor

    sensitive_message = 'customer=Top Secret; pipe_od=999'

    def explode(*_args, **_kwargs):
        raise RuntimeError(sensitive_message)

    monkeypatch.setattr(workbook_processor, 'calculate_row', explode)
    with caplog.at_level('ERROR', logger='workbook_processor'):
        result = process_workbook(workbook_bytes_with_rows([valid_row_values()]), processed_at=FIXED_TIME)

    assert result.status_counts == {'SYSTEM ERROR': 1}
    assert 'source Excel row 2' in caplog.text
    assert 'RuntimeError' in caplog.text
    assert 'explode' in caplog.text
    assert sensitive_message not in caplog.text
    assert 'Top Secret' not in caplog.text
    assert 'pipe_od=999' not in caplog.text
    assert 'Batch Customer' not in caplog.text
    assert '457.2' not in caplog.text


def _encrypted_zip_bytes() -> bytes:
    output = BytesIO()
    with zipfile.ZipFile(output, 'w') as archive:
        entry = zipfile.ZipInfo('xl/workbook.xml')
        entry.flag_bits |= 0x1
        archive.writestr(entry, b'<workbook/>')
    return output.getvalue()


def _dense_workbook_bytes(
    cell_count: int,
    namespace: str = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
) -> bytes:
    """Return a valid controlled XLSX package with one deliberately dense sheet."""
    worksheet_xml = _dense_worksheet_xml(cell_count, namespace)
    source = workbook_bytes_with_rows([valid_row_values()])
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(source)) as input_zip, zipfile.ZipFile(output, 'w') as output_zip:
        for entry in input_zip.infolist():
            content = input_zip.read(entry.filename)
            if entry.filename == 'xl/worksheets/sheet5.xml':
                replacement = zipfile.ZipInfo(entry.filename)
                replacement.compress_type = zipfile.ZIP_STORED
                output_zip.writestr(replacement, worksheet_xml)
            else:
                output_zip.writestr(entry, content)
    return output.getvalue()


def _relocated_dense_workbook_bytes(
    cell_count: int,
    new_path: str = 'xl/custom/dense.xml',
) -> bytes:
    """Return a controlled XLSX whose dense worksheet part uses a custom path."""
    return _relocated_worksheet_workbook_bytes(
        _dense_worksheet_xml(cell_count),
        new_path=new_path,
    )


def _relabeled_dense_worksheet_bytes(cell_count: int) -> bytes:
    """Keep a worksheet relationship but relabel its target as generic XML."""
    worksheet_content_type = (
        b'application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml'
    )
    source = _dense_workbook_bytes(cell_count)
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(source)) as input_zip, zipfile.ZipFile(
        output, 'w',
    ) as output_zip:
        for entry in input_zip.infolist():
            content = input_zip.read(entry.filename)
            if entry.filename == '[Content_Types].xml':
                target = (
                    b'<Override PartName="/xl/worksheets/sheet5.xml" ContentType="'
                    + worksheet_content_type
                    + b'"/>'
                )
                replacement = (
                    b'<Override PartName="/xl/worksheets/sheet5.xml" '
                    b'ContentType="application/xml"/>'
                )
                assert target in content
                content = content.replace(target, replacement)
            output_zip.writestr(entry, content)
    return output.getvalue()


def _dense_workbook_with_duplicate_relationship(
    cell_count: int,
    relationship_namespace: str | None = None,
) -> bytes:
    """Append a safe-target decoy after the genuine dense-sheet relationship."""
    source = _relabeled_dense_worksheet_bytes(cell_count)
    namespace = (
        f' xmlns="{relationship_namespace}"'.encode()
        if relationship_namespace is not None
        else b''
    )
    decoy = (
        b'<Relationship'
        + namespace
        + b' Type="http://schemas.openxmlformats.org/officeDocument/2006/'
        b'relationships/worksheet" Target="/xl/worksheets/sheet6.xml" '
        b'Id="rId5"/>'
    )
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(source)) as input_zip, zipfile.ZipFile(
        output, 'w',
    ) as output_zip:
        for entry in input_zip.infolist():
            content = input_zip.read(entry.filename)
            if entry.filename == 'xl/_rels/workbook.xml.rels':
                assert content.endswith(b'</Relationships>')
                content = content.replace(
                    b'</Relationships>', decoy + b'</Relationships>',
                )
            output_zip.writestr(entry, content)
    return output.getvalue()


def _explicit_internal_relationship_workbook_bytes() -> bytes:
    """Return a valid workbook with explicitly internal OPC relationships."""
    source = workbook_bytes_with_rows([valid_row_values()])
    office_document = (
        b'<Relationship Type="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/relationships/officeDocument" '
        b'Target="xl/workbook.xml" Id="rId1"/>'
    )
    explicit_office_document = office_document.replace(
        b'/>', b' TargetMode="Internal"/>',
    )
    worksheet = (
        b'<Relationship Type="http://schemas.openxmlformats.org/'
        b'officeDocument/2006/relationships/worksheet" '
        b'Target="/xl/worksheets/sheet1.xml" Id="rId1"/>'
    )
    explicit_worksheet = worksheet.replace(
        b'/>', b' TargetMode="iNtErNaL"/>',
    )
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(source)) as input_zip, zipfile.ZipFile(
        output, 'w',
    ) as output_zip:
        for entry in input_zip.infolist():
            content = input_zip.read(entry.filename)
            if entry.filename == '_rels/.rels':
                assert office_document in content
                content = content.replace(
                    office_document, explicit_office_document,
                )
            elif entry.filename == 'xl/_rels/workbook.xml.rels':
                assert worksheet in content
                content = content.replace(worksheet, explicit_worksheet)
            output_zip.writestr(entry, content)
    return output.getvalue()


def _relocated_worksheet_workbook_bytes(
    worksheet_xml: bytes,
    new_path: str,
) -> bytes:
    """Relocate one valid worksheet part and update its OPC declarations."""
    old_path = b'/xl/worksheets/sheet5.xml'
    new_part_name = f'/{new_path}'.encode()
    source = workbook_bytes_with_rows([valid_row_values()])
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(source)) as input_zip, zipfile.ZipFile(output, 'w') as output_zip:
        for entry in input_zip.infolist():
            if entry.filename == 'xl/worksheets/sheet5.xml':
                continue
            content = input_zip.read(entry.filename)
            if entry.filename in {'[Content_Types].xml', 'xl/_rels/workbook.xml.rels'}:
                content = content.replace(old_path, new_part_name)
            output_zip.writestr(entry, content)
        replacement = zipfile.ZipInfo(new_path)
        replacement.compress_type = zipfile.ZIP_STORED
        output_zip.writestr(replacement, worksheet_xml)
    return output.getvalue()


def _workbook_with_binary_custom_part() -> bytes:
    """Add a related binary part with a valid OPC content-type declaration."""
    default_type = (
        b'<Default Extension="data" ContentType="application/octet-stream"/>'
    )
    relationship = (
        b'<Relationship Type="urn:protap:binary-attachment" '
        b'Target="custom/blob.data" Id="rId4"/>'
    )
    source = workbook_bytes_with_rows([valid_row_values()])
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(source)) as input_zip, zipfile.ZipFile(output, 'w') as output_zip:
        for entry in input_zip.infolist():
            content = input_zip.read(entry.filename)
            if entry.filename == '[Content_Types].xml':
                content = content.replace(b'</Types>', default_type + b'</Types>')
            elif entry.filename == '_rels/.rels':
                content = content.replace(b'</Relationships>', relationship + b'</Relationships>')
            output_zip.writestr(entry, content)
        binary_entry = zipfile.ZipInfo('custom/blob.data')
        binary_entry.compress_type = zipfile.ZIP_STORED
        output_zip.writestr(binary_entry, b'\x00\xffnot-xml' * 10_000)
    return output.getvalue()


def _foreign_custom_xml_workbook_bytes(cell_count: int) -> bytes:
    """Add a related custom XML part whose local names resemble a worksheet."""
    custom_xml = (
        b'<?xml version="1.0" encoding="UTF-8"?>'
        b'<worksheet xmlns="urn:customer-custom-data">'
        + b'<c/>' * cell_count
        + b'</worksheet>'
    )
    relationship = (
        b'<Relationship '
        b'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/customXml" '
        b'Target="customXml/item1.xml" Id="rId4"/>'
    )
    source = workbook_bytes_with_rows([valid_row_values()])
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(source)) as input_zip, zipfile.ZipFile(output, 'w') as output_zip:
        for entry in input_zip.infolist():
            content = input_zip.read(entry.filename)
            if entry.filename == '_rels/.rels':
                content = content.replace(b'</Relationships>', relationship + b'</Relationships>')
            output_zip.writestr(entry, content)
        custom_entry = zipfile.ZipInfo('customXml/item1.xml')
        custom_entry.compress_type = zipfile.ZIP_STORED
        output_zip.writestr(custom_entry, custom_xml)
    return output.getvalue()


def _worksheet_with_foreign_extension_bytes(cell_count: int) -> bytes:
    """Add many foreign c elements inside a valid worksheet extension container."""
    extension = (
        b'<extLst><ext uri="{PROTAP-FOREIGN-CELL-TEST}" '
        b'xmlns:f="urn:customer-custom-data">'
        + b'<f:c/>' * cell_count
        + b'</ext></extLst>'
    )
    source = workbook_bytes_with_rows([valid_row_values()])
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(source)) as input_zip, zipfile.ZipFile(output, 'w') as output_zip:
        for entry in input_zip.infolist():
            content = input_zip.read(entry.filename)
            if entry.filename == 'xl/worksheets/sheet5.xml':
                content = content.replace(b'</worksheet>', extension + b'</worksheet>')
                replacement = zipfile.ZipInfo(entry.filename)
                replacement.compress_type = zipfile.ZIP_STORED
                output_zip.writestr(replacement, content)
            else:
                output_zip.writestr(entry, content)
    return output.getvalue()


def _dense_worksheet_xml(
    cell_count: int,
    namespace: str = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
) -> bytes:
    cells_per_row = 100
    rows = []
    for first_cell in range(0, cell_count, cells_per_row):
        row_number = first_cell // cells_per_row + 1
        row_cell_count = min(cells_per_row, cell_count - first_cell)
        cells = ''.join(
            f'<c r="{_column_name(column)}{row_number}" t="n"><v>1</v></c>'
            for column in range(1, row_cell_count + 1)
        )
        rows.append(f'<row r="{row_number}">{cells}</row>')
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<worksheet xmlns="{namespace}">'
        f'<sheetData>{"".join(rows)}</sheetData></worksheet>'
    ).encode()


def _column_name(column: int) -> str:
    name = ''
    while column:
        column, remainder = divmod(column - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _stable_json(value: str) -> str:
    import json

    return json.dumps(json.loads(value), sort_keys=True, separators=(',', ':'))
