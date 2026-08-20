"""Generate the controlled, formula-free PROWRAP batch input workbook."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName

from batch_mechanisms import CANONICAL_MECHANISMS
from batch_schema import (
    DETAIL_INPUT_HEADERS,
    DETAIL_OUTPUT_HEADERS,
    INPUT_HEADERS,
    MAX_DETAIL_ROWS,
    MAX_ROWS,
    OUTPUT_HEADERS,
)
from cost_calculation import COST_INPUTS, COST_TABLE_HEADERS
from engine.corrosion_defects import DEFECT_LENGTH_BASES
from workbook_formatting import (
    HEADER_HEIGHT,
    INPUT_ERROR_COLOR,
    INPUT_HEADER_COLOR,
    NOT_REPAIRABLE_COLOR,
    OK_COLOR,
    OUTPUT_HEADER_COLOR,
    REVIEW_REQUIRED_COLOR,
    SYSTEM_ERROR_COLOR,
    apply_common_field_style,
    apply_header_style,
    header_column_letter,
    set_capped_column_widths,
    unlock_cells,
)


_CHOICES = {
    'MechanismChoices': ('Mechanism', CANONICAL_MECHANISMS),
    'DefectLocationChoices': ('Defect Location', ('External', 'Internal')),
    'TypeACheckChoices': ('Run Type A / Class 3 Check', ('Yes', 'No')),
    'ComponentTypeChoices': (
        'Component Type', ('Straight', 'Bend', 'Tee', 'Flange', 'Reducer'),
    ),
    'AxialLoadCaseChoices': ('Axial Load Case', (0, 1)),
    'DefectLengthBasisChoices': ('Defect Length Basis', DEFECT_LENGTH_BASES),
    'SeparationChoices': ('Separation exceeds 3t', ('Yes', None)),
}

_MAIN_DROPDOWN_NAMES = (
    'MechanismChoices',
    'DefectLocationChoices',
    'TypeACheckChoices',
    'ComponentTypeChoices',
    'AxialLoadCaseChoices',
    'DefectLengthBasisChoices',
)
_DETAIL_DROPDOWN_NAMES = ('SeparationChoices',)

_HEADER_NOTES = {
    'Pipe OD [mm]': 'Required. Enter the outside diameter in millimetres; value must be positive.',
    'Nominal Wall [mm]': 'Required. Enter the nominal pipe wall in millimetres; value must be positive.',
    'Pipe Yield [MPa]': 'Required. Enter the specified pipe yield strength in MPa; value must be positive.',
    'Design Pressure [bar]': 'Required. Enter the design pressure in bar; zero or a positive value.',
    'Operating Temperature [degC]': 'Required. Enter the operating temperature in degrees C.',
    'Mechanism': (
        'Required. Dent w/crack uses a full-pressure laminate. An eligible external '
        'Dent no-crack uses component-pipe substrate load sharing. Legacy Dent is '
        'accepted only when upgrading an older batch workbook and becomes Dent w/crack.'
    ),
    'Defect Location': 'Required. Choose External or Internal.',
    'Defect Length [mm]': 'Required. Enter the defect length in millimetres; value must be positive.',
    'Defect Length Basis': (
        'For external corrosion, choose Actual defect length, Independent defects, '
        'or Enter manually as described on the Instructions worksheet.'
    ),
    'Repair Group ID': (
        'Required only for Enter manually. Use the same identifier on each linked '
        'Individual Defects row.'
    ),
    'Remaining Wall [mm]': 'Required. Enter the minimum remaining wall in millimetres; it cannot exceed nominal wall.',
    'Internal Corrosion Rate [mm/year]': (
        'Enter zero or a positive internal corrosion rate in mm/year. This value is '
        'required only for internal corrosion and may otherwise be blank.'
    ),
    'Design Life [years]': 'Required. Enter a whole number of years, at least one.',
    'Design Factor': 'Required. Enter a value from 0.10 through 1.00.',
    'Run Type A / Class 3 Check': 'Required. Choose Yes only when the additional Type A / Class 3 check is required.',
    'Installation Temperature [degC]': 'Required. Enter the installation temperature in degrees C.',
    'Component Type': 'Required. Choose Straight, Bend, Tee, Flange, or Reducer.',
    'Cyclic Derating Factor': 'Required. Enter a factor greater than zero and no greater than one.',
    'Axial Load Case': 'Required. Choose 0 for no axial load case or 1 for axial load case.',
    'Prowrap CF Cloth Width [mm]': (
        'Required. Enter a cloth width greater than the 50 mm stitch overlap; '
        '300 mm and 500 mm are approved configured widths.'
    ),
}

_THIN_GRAY = Side(style='thin', color='D9E1F2')
_COST_INPUT_COLOR = 'FFF2CC'

_DETAIL_HEADER_NOTES = {
    'Repair Group ID': 'Required. Link this individual defect to an Enter manually main row.',
    'Defect ID': 'Required. Use a unique defect identifier within each repair group.',
    'Individual longitudinal length [mm]': 'Required. Enter a positive individual defect length in millimetres.',
    'Remaining wall [mm]': 'Required. Enter the individual defect minimum remaining wall in millimetres.',
    'Separation exceeds 3t': 'Required. Choose Yes only when the next defect is separated by more than three nominal wall thicknesses.',
}


def create_template_workbook() -> bytes:
    """Return a ready-to-fill controlled batch workbook as ``.xlsx`` bytes."""
    workbook = Workbook()
    workbook.properties.title = 'PROWRAP CalcBatch v1.2'
    batch_info = workbook.active
    batch_info.title = 'Batch Information'
    data = workbook.create_sheet('Batch Input & Results')
    individual_defects = workbook.create_sheet('Individual Defects')
    cost = workbook.create_sheet('Cost Calculation')
    warnings = workbook.create_sheet('Warnings')
    summary = workbook.create_sheet('Summary')
    instructions = workbook.create_sheet('Instructions')
    lists = workbook.create_sheet('Lists')

    _build_batch_information(batch_info)
    _build_data_sheet(data)
    _build_individual_defects(individual_defects)
    _build_cost_calculation(cost)
    _build_warnings(warnings)
    _build_summary(summary)
    _build_instructions(instructions)
    _build_lists(workbook, lists)
    lists.sheet_state = 'hidden'

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _build_batch_information(worksheet) -> None:
    worksheet['A1'] = 'PROWRAP CalcBatch v1.2'
    worksheet['A1'].font = Font(name='Calibri', size=16, bold=True, color=INPUT_HEADER_COLOR)
    worksheet['A2'] = 'Enter the three values that apply to every defect row in this batch.'
    worksheet['A2'].alignment = Alignment(wrap_text=True)
    for row, label in enumerate(('Customer', 'Project Location', 'Report No'), start=3):
        worksheet.cell(row, 1, label)
        worksheet.cell(row, 2)
        apply_common_field_style(worksheet.cell(row, 1), worksheet.cell(row, 2))
        worksheet.cell(row, 2).comment = Comment('Enter this common batch value once.', 'PROTAP')
    worksheet.column_dimensions['A'].width = 24
    worksheet.column_dimensions['B'].width = 42
    worksheet.freeze_panes = 'A3'


def _build_data_sheet(worksheet) -> None:
    headers = INPUT_HEADERS + OUTPUT_HEADERS
    input_count = len(INPUT_HEADERS)
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(1, column, header)
        apply_header_style(
            cell, INPUT_HEADER_COLOR if column <= input_count else OUTPUT_HEADER_COLOR,
        )
        if column <= input_count:
            cell.comment = Comment(_HEADER_NOTES[header], 'PROTAP')

    worksheet.row_dimensions[1].height = HEADER_HEIGHT
    worksheet.freeze_panes = 'B2'
    end_column = worksheet.cell(1, len(headers)).coordinate.rstrip('1')
    table_ref = f'A1:{end_column}{MAX_ROWS + 1}'
    table = Table(displayName='BatchRows', ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    worksheet.add_table(table)

    input_cells = (
        worksheet.cell(row, column)
        for row in range(2, MAX_ROWS + 2)
        for column in range(1, input_count + 1)
    )
    unlock_cells(input_cells)
    for row in range(2, MAX_ROWS + 2):
        for column in range(1, len(headers) + 1):
            cell = worksheet.cell(row, column)
            cell.border = Border(bottom=_THIN_GRAY)
            if column >= input_count + 1:
                is_diagnostic_detail = headers[column - 1] in {
                    'B31G Detail', 'Type A Detail', 'Type B Detail',
                }
                cell.alignment = Alignment(
                    vertical='top',
                    wrap_text=not is_diagnostic_detail,
                    shrink_to_fit=is_diagnostic_detail,
                )

    _add_dropdowns(worksheet, _MAIN_DROPDOWN_NAMES, MAX_ROWS)
    _add_status_formatting(worksheet, input_count + 2)
    worksheet.protection.sheet = True
    worksheet.protection.autoFilter = False
    worksheet.protection.selectLockedCells = False
    worksheet.protection.selectUnlockedCells = False
    set_capped_column_widths(worksheet)
    for column in ('D', 'E', 'H', 'I', 'J', 'T', 'U', 'V', 'W', 'AF', 'AX', 'AY'):
        worksheet.column_dimensions[column].width = 28
    worksheet.column_dimensions['W'].width = 16


def _build_individual_defects(worksheet) -> None:
    """Build the protected 2,000-row linked-corrosion input table."""
    headers = DETAIL_INPUT_HEADERS + DETAIL_OUTPUT_HEADERS
    input_count = len(DETAIL_INPUT_HEADERS)
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(1, column, header)
        apply_header_style(
            cell, INPUT_HEADER_COLOR if column <= input_count else OUTPUT_HEADER_COLOR,
        )
        if column <= input_count:
            cell.comment = Comment(_DETAIL_HEADER_NOTES[header], 'PROTAP')

    worksheet.row_dimensions[1].height = HEADER_HEIGHT
    worksheet.freeze_panes = 'B2'
    table = Table(
        displayName='IndividualDefects',
        ref=f'A1:{worksheet.cell(1, len(headers)).column_letter}{MAX_DETAIL_ROWS + 1}',
    )
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    worksheet.add_table(table)

    unlock_cells(
        worksheet.cell(row, column)
        for row in range(2, MAX_DETAIL_ROWS + 2)
        for column in range(1, input_count + 1)
    )
    for row in range(2, MAX_DETAIL_ROWS + 2):
        for column in range(1, len(headers) + 1):
            cell = worksheet.cell(row, column)
            cell.border = Border(bottom=_THIN_GRAY)
            if column > input_count:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    _add_dropdowns(worksheet, _DETAIL_DROPDOWN_NAMES, MAX_DETAIL_ROWS)
    _add_status_formatting(worksheet, input_count + 2, MAX_DETAIL_ROWS)
    worksheet.protection.sheet = True
    worksheet.protection.autoFilter = False
    worksheet.protection.selectLockedCells = False
    worksheet.protection.selectUnlockedCells = False
    set_capped_column_widths(worksheet)


def _build_cost_calculation(worksheet) -> None:
    worksheet['A1'] = 'PROWRAP Cost Calculation'
    worksheet['A1'].font = Font(
        name='Calibri', size=16, bold=True, color=INPUT_HEADER_COLOR,
    )

    for address, label in COST_INPUTS:
        value_cell = worksheet[address]
        label_cell = worksheet.cell(value_cell.row, value_cell.column - 1, label)
        apply_common_field_style(label_cell, value_cell)
        value_cell.fill = PatternFill(fill_type='solid', fgColor=_COST_INPUT_COLOR)
        value_cell.number_format = '#,##0.00'

    validation = DataValidation(
        type='decimal', operator='greaterThanOrEqual', formula1='0',
        allow_blank=True, errorTitle='Enter zero or a positive number',
        error='Enter a numeric value greater than or equal to zero, or leave this cell blank.',
        showErrorMessage=True, errorStyle='stop',
    )
    for address, _ in COST_INPUTS:
        validation.add(address)
    worksheet.add_data_validation(validation)

    for column, header in enumerate(COST_TABLE_HEADERS, start=1):
        apply_header_style(worksheet.cell(5, column, header), OUTPUT_HEADER_COLOR)
    worksheet.row_dimensions[5].height = HEADER_HEIGHT

    table = Table(displayName='CostRows', ref='A5:V6')
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    worksheet.add_table(table)

    for row in range(6, MAX_ROWS + 6):
        for column in range(1, len(COST_TABLE_HEADERS) + 1):
            worksheet.cell(row, column).number_format = '#,##0.00'
        for column in (10, 15, 17):
            worksheet.cell(row, column).number_format = '#,##0'

    worksheet.freeze_panes = 'A6'
    worksheet.sheet_view.showGridLines = False
    worksheet.protection.sheet = True
    worksheet.protection.autoFilter = False
    worksheet.protection.selectLockedCells = False
    worksheet.protection.selectUnlockedCells = False
    set_capped_column_widths(worksheet)
    worksheet.merge_cells('A1:V1')


def _build_warnings(worksheet) -> None:
    worksheet['A1'] = 'Compliance Warning Register'
    worksheet['A1'].font = Font(
        name='Calibri', size=16, bold=True, color=INPUT_HEADER_COLOR,
    )
    worksheet['A2'] = (
        'Permanent warning references used in Batch Input & Results. '
        'Affected Excel rows use the original worksheet row numbers.'
    )
    worksheet.merge_cells('A2:C2')
    worksheet['A2'].alignment = Alignment(wrap_text=True, vertical='top')
    headings = (
        'Warning Code',
        'Warning Meaning / Required Action',
        'Affected Excel Rows',
    )
    for column, heading in enumerate(headings, start=1):
        apply_header_style(worksheet.cell(3, column, heading), OUTPUT_HEADER_COLOR)
    worksheet['A4'] = 'No compliance warnings were generated.'
    worksheet['A4'].font = Font(name='Calibri', italic=True, color='666666')
    worksheet['A4'].alignment = Alignment(vertical='top')
    worksheet.freeze_panes = 'A4'
    worksheet.sheet_view.showGridLines = False
    worksheet.column_dimensions['A'].width = 16
    worksheet.column_dimensions['B'].width = 90
    worksheet.column_dimensions['C'].width = 28
    worksheet.row_dimensions[2].height = 32
    worksheet.row_dimensions[3].height = HEADER_HEIGHT
    worksheet.protection.sheet = True
    worksheet.protection.autoFilter = False
    worksheet.protection.selectLockedCells = False


def _build_summary(worksheet) -> None:
    worksheet['A1'] = 'Batch Summary'
    worksheet['A1'].font = Font(name='Calibri', size=16, bold=True, color=INPUT_HEADER_COLOR)
    worksheet['A3'] = 'Customer'
    worksheet['A4'] = 'Project Location'
    worksheet['A5'] = 'Report No'
    worksheet['A7'] = 'Workbook Name'
    worksheet['A8'] = 'Processing Time [UTC]'
    worksheet['A10'] = 'Total Populated Rows'
    worksheet['A12'] = 'Status Counts'
    for row, status in enumerate(('OK', 'REVIEW REQUIRED', 'NOT REPAIRABLE', 'INPUT ERROR', 'SYSTEM ERROR'), start=13):
        worksheet.cell(row, 1, status)
    worksheet['A19'] = 'Type A Route Count'
    worksheet['A20'] = 'Type B Route Count'
    worksheet['A21'] = 'Rows with Compliance Warnings'
    worksheet['A22'] = 'Rows Requiring Engineering Review'
    worksheet['A24'] = 'Batch Engine Version'
    worksheet['A25'] = 'Pinned Source Revision'
    worksheet['A27'] = (
        'These results are preliminary screening outputs only and require competent engineering review '
        'before repair design, approval, procurement, or installation.'
    )
    worksheet['A27'].font = Font(name='Calibri', size=10, italic=True, color='9C0006')
    worksheet['A27'].alignment = Alignment(wrap_text=True, vertical='top')
    worksheet.merge_cells('A27:B29')
    for row in range(3, 26):
        if worksheet.cell(row, 1).value:
            worksheet.cell(row, 1).font = Font(name='Calibri', bold=True)
    worksheet.column_dimensions['A'].width = 35
    worksheet.column_dimensions['B'].width = 42
    worksheet.row_dimensions[27].height = 48
    worksheet.protection.sheet = True
    worksheet.protection.selectLockedCells = False
    worksheet.protection.selectUnlockedCells = False


def _build_instructions(worksheet) -> None:
    lines = (
        ('A1', 'PROWRAP CalcBatch v1.2 — Instructions', True),
        ('A3', '1. Complete Customer, Project Location, and Report No once on the Batch Information sheet.', False),
        ('A4', '2. Enter one main row per continuous repair on Batch Input & Results; the first input is Pipe OD [mm].', False),
        ('A5', f'3. Enter up to {MAX_ROWS} populated rows. Blank rows are ignored; partially populated rows receive INPUT ERROR.', False),
        ('A6', '4. Use the dropdown selections exactly as shown. Units are mm, MPa, bar, degC, years, m2, and kg where stated.', False),
        ('A7', '5. Internal Corrosion Rate [mm/year] is required only where Mechanism is Corrosion and Defect Location is Internal.', False),
        ('A8', '6. Prowrap CF Cloth Width must be greater than the fixed 50 mm stitch overlap. The approved configured widths are 300 mm and 500 mm; other valid widths require review.', False),
        ('A9', '7. Processed result rows show permanent warning codes only. Read their full meaning, required action, and affected rows on the Warnings worksheet.', False),
        ('A10', '8. On Cost Calculation, B3 (CF Cost / m2), E3 (Epoxy Cost / kg), and H3 (Price Multiplier) are highlighted and editable. They may be blank or retain values from a previously processed workbook.', False),
        ('A11', '9. Cost = Fabric Area x CF Cost / m2 + Epoxy Mass x Epoxy Cost / kg.', False),
        ('A12', '10. Price = Cost x Price Multiplier. No currency symbol is fixed, so use one consistent currency for both material rates.', False),
        ('A13', '11. The downloaded input template contains no formulas. A processed workbook contains only controlled Cost and Price formulas and may be safely uploaded again.', False),
        ('A14', '12. Previously downloaded controlled five-sheet, six-sheet, and seven-sheet workbooks remain accepted and are upgraded to the current eight-sheet output.', False),
        ('A15', '13. Dent w/crack uses a full-pressure laminate. An eligible external Dent no-crack uses component-pipe substrate load sharing. Dent no-crack selects a calculation basis; it is not a complete dent integrity or fatigue acceptance assessment. Legacy Dent is accepted only when upgrading an older batch workbook and becomes Dent w/crack.', False),
        ('A16', 'Status meanings', True),
        ('A17', 'OK — a valid result with no review warning.', False),
        ('A18', 'REVIEW REQUIRED — a numeric result exists, but an engineering or product-approval condition needs review.', False),
        ('A19', 'NOT REPAIRABLE — the Type B Formula 12 route has no repair solution for the requested case.', False),
        ('A20', 'INPUT ERROR — correct the indicated input and calculate again.', False),
        ('A21', 'SYSTEM ERROR — an unexpected processing issue occurred; retain the workbook and contact PROTAP.', False),
        ('A23', 'Material temperature basis: Tg = 110 degC, general qualified design limit = 90 degC, and long-life Class 3 Type B limit = 80 degC. The input template contains no formulas or macros. It is a controlled input template, not an engineering approval or certification.', False),
        ('A24', 'Linked corrosion modes: Actual defect length = continuous or interacting B31G length. Independent defects = 10 x 10 mm, each separated by more than 3t. t means nominal pipe wall thickness. Enter manually = leave main Remaining Wall blank and link detail rows with Repair Group ID. Defect Length remains the complete outer-to-outer continuous repair-zone span.', False),
    )
    for address, text, heading in lines:
        cell = worksheet[address]
        cell.value = text
        cell.font = Font(name='Calibri', size=14 if address == 'A1' else 11, bold=heading)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    worksheet.column_dimensions['A'].width = 115
    for row in range(3, 25):
        worksheet.row_dimensions[row].height = 32
    worksheet.row_dimensions[15].height = 48
    worksheet.row_dimensions[24].height = 64
    worksheet.row_dimensions[1].height = 28


def _build_lists(workbook, worksheet) -> None:
    for column, (name, (_, choices)) in enumerate(_CHOICES.items(), start=1):
        for row, value in enumerate(choices, start=1):
            worksheet.cell(row, column, value)
        letter = worksheet.cell(1, column).column_letter
        workbook.defined_names.add(DefinedName(
            name, attr_text=f"'Lists'!${letter}$1:${letter}${len(choices)}",
        ))


def _add_dropdowns(worksheet, choice_names, max_rows: int) -> None:
    """Add controlled selections by resolving each target from its semantic header."""
    for name in choice_names:
        header, _ = _CHOICES[name]
        letter = header_column_letter(worksheet, header)
        validation = DataValidation(
            type='list', formula1=f'={name}', allow_blank=True,
            errorTitle='Select a supported value',
            error='Choose a value from the controlled list.',
            showErrorMessage=True,
            errorStyle='stop',
        )
        validation.add(f'{letter}2:{letter}{max_rows + 1}')
        worksheet.add_data_validation(validation)


def _add_status_formatting(worksheet, status_column: int, max_rows: int = MAX_ROWS) -> None:
    letter = worksheet.cell(1, status_column).column_letter
    status_colors = {
        'OK': OK_COLOR,
        'REVIEW REQUIRED': REVIEW_REQUIRED_COLOR,
        'NOT REPAIRABLE': NOT_REPAIRABLE_COLOR,
        'INPUT ERROR': INPUT_ERROR_COLOR,
        'SYSTEM ERROR': SYSTEM_ERROR_COLOR,
    }
    for status, color in status_colors.items():
        rule = Rule(
            type='containsText', operator='containsText', text=status,
            dxf=DifferentialStyle(fill=PatternFill(fill_type='solid', fgColor=color)),
        )
        rule.formula = [f'NOT(ISERROR(SEARCH("{status}",{letter}2)))']
        worksheet.conditional_formatting.add(f'{letter}2:{letter}{max_rows + 1}', rule)
