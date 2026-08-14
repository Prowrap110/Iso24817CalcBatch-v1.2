"""Reusable visual and protection rules for PROWRAP batch workbooks."""

from copy import copy

from openpyxl.styles import Alignment, Font, PatternFill, Protection


INPUT_HEADER_COLOR = '1F4E78'
OUTPUT_HEADER_COLOR = '404040'
COMMON_FIELD_COLOR = 'D9EAF7'
OK_COLOR = 'C6E0B4'
REVIEW_REQUIRED_COLOR = 'FFD966'
NOT_REPAIRABLE_COLOR = 'F4CCCC'
INPUT_ERROR_COLOR = 'EA9999'
SYSTEM_ERROR_COLOR = 'B7B7B7'

WHITE = 'FFFFFF'
HEADER_HEIGHT = 42
MIN_COLUMN_WIDTH = 12
MAX_COLUMN_WIDTH = 38


def apply_header_style(cell, color: str) -> None:
    """Apply a readable, wrapped header style to one table heading."""
    cell.fill = PatternFill(fill_type='solid', fgColor=color)
    cell.font = Font(name='Calibri', size=10, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def apply_common_field_style(label_cell, value_cell) -> None:
    """Style one batch-level label/value pair and leave only the value editable."""
    label_cell.fill = PatternFill(fill_type='solid', fgColor=COMMON_FIELD_COLOR)
    label_cell.font = Font(name='Calibri', size=11, bold=True)
    label_cell.alignment = Alignment(vertical='center')
    value_cell.fill = PatternFill(fill_type='solid', fgColor='FFFFFF')
    value_cell.alignment = Alignment(vertical='center')
    value_cell.protection = Protection(locked=False)


def unlock_cells(cells) -> None:
    """Mark editable worksheet cells unlocked while preserving their style."""
    for cell in cells:
        cell.protection = Protection(locked=False, hidden=cell.protection.hidden)


def apply_wrapped_text(cells) -> None:
    """Make content readable without clipping long instructions or diagnostics."""
    for cell in cells:
        cell.alignment = copy(cell.alignment)
        cell.alignment = Alignment(
            horizontal=cell.alignment.horizontal,
            vertical=cell.alignment.vertical or 'top',
            wrap_text=True,
        )


def set_capped_column_widths(worksheet, minimum=MIN_COLUMN_WIDTH, maximum=MAX_COLUMN_WIDTH) -> None:
    """Fit visible content approximately while preventing impractically wide columns."""
    for column_cells in worksheet.iter_cols():
        column_letter = column_cells[0].column_letter
        longest = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=minimum,
        )
        worksheet.column_dimensions[column_letter].width = min(
            maximum, max(minimum, longest + 2)
        )
